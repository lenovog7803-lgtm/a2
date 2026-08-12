"""
Построение официальной формы Книги учёта доходов и расходов (КУДиР, Приложение 9
к Инструкции о порядке ведения учёта доходов и расходов) в openpyxl.

Разметка граф соответствует официальному бланку (nalog.gov.by):
  B — графа 1  (дата записи)
  C — графа 2  (наименование документа, его номер, дата)
  D — графа 3  (содержание хозяйственной операции)
  E — графа 4  (доходы, учитываемые в отчётном периоде — сумма)
  F — графа 5  (сумма налогов, сборов, уплаченная из выручки)
  G — графа 6  (освобождаемые доходы, сумма)
  H — графа 7  (иные поступления)
  I — графа 8  (расходы, приходящиеся на отчётный период)
  J — графа 9  (расходы по нормативу)
  K — графа 10 (иные расходы)
  L — графа 11 (примечание)
Столбец A в официальном бланке не используется (узкий отступ слева).

Эта система заполняет только графу 4 (доход-вознаграждение экспедитора,
E) и графу 11 (примечание, L) — остальные графы (F..K) остаются пустыми,
но участвуют в формулах ИТОГО, чтобы книга оставалась в полном
официальном формате.
"""
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

FONT_NAME = "Times New Roman CYR"

COLUMNS = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]

border_all = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

left_bottom = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
_left_center = Alignment(horizontal="left", vertical="center", wrap_text=True)
_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
_right_top = Alignment(horizontal="right", vertical="top", wrap_text=True)
_right_bottom = Alignment(horizontal="right", vertical="bottom")
_center_bottom = Alignment(horizontal="center", vertical="bottom", wrap_text=True)


def f(size, bold=False, italic=False):
    """Фабрика шрифта Times New Roman CYR."""
    return Font(name=FONT_NAME, size=size, bold=bold, italic=italic)


def set_widths(ws):
    widths = {
        "A": 1.2, "B": 11, "C": 19, "D": 24,
        "E": 10, "F": 15, "G": 12.7, "H": 11, "I": 13, "J": 10, "K": 13,
        "L": 12,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_header(ws, redaction_note):
    """Шапка формы (реквизиты бланка, название книги). Занимает строки 3-9."""
    ws.merge_cells("K3:L3")
    ws["K3"] = "Приложение 9"
    ws["K3"].font = f(8)
    ws["K3"].alignment = _right_top

    ws.merge_cells("H4:L4")
    ws["H4"] = (
        "к Инструкции о порядке ведения учета доходов и расходов\n"
        f"(в редакции постановления Министерства по налогам и сборам\nРеспублики Беларусь {redaction_note})"
    )
    ws["H4"].font = f(8)
    ws["H4"].alignment = _right_top

    ws["L5"] = "Форма"
    ws["L5"].font = f(8)
    ws["L5"].alignment = _right_bottom

    ws.merge_cells("B6:L6")
    ws["B6"] = "КНИГА"
    ws["B6"].font = f(11, bold=True)
    ws["B6"].alignment = _center_bottom

    ws.merge_cells("B7:L7")
    ws["B7"] = "учета доходов и расходов"
    ws["B7"].font = f(11, bold=True)
    ws["B7"].alignment = _center_bottom

    ws["L9"] = "(руб.)"
    ws["L9"].font = f(9)
    ws["L9"].alignment = _right_bottom


def build_table_head(ws):
    """Шапка таблицы с номерами и названиями граф. Занимает строки 11-14."""
    def merged(rng, value):
        ws.merge_cells(rng)
        cell = ws[rng.split(":")[0]]
        cell.value = value
        cell.font = f(9)
        cell.alignment = _center_wrap
        return cell

    merged("B11:B13", "Дата записи")
    merged("C11:C13", "Наименование документа, его номер, дата")
    merged("D11:D13", "Содержание хозяйственной операции")
    merged("E11:H11", "Доходы")
    merged("I11:K11", "Расходы")
    merged("L11:L13", "Примечание")

    merged("E12:F12", "доходы, учитываемые в отчетном периоде")
    merged("G12:G13", "освобождаемые доходы,\nсумма")
    merged("H12:H13", "иные поступления")
    merged("I12:I13", "расходы, приходящиеся на отчетный период")
    merged("J12:J13", "расходы по нормативу")
    merged("K12:K13", "иные расходы")

    ws["E13"] = "сумма"
    ws["E13"].font = f(9)
    ws["E13"].alignment = _center_wrap
    ws["F13"] = "сумма налогов, сборов, уплаченная из выручки"
    ws["F13"].font = f(9)
    ws["F13"].alignment = _center_wrap

    graph_numbers = {"B": 1, "C": 2, "D": 3, "E": 4, "F": 5, "G": 6, "H": 7, "I": 8, "J": 9, "K": 10, "L": 11}
    for col, num in graph_numbers.items():
        cell = ws[f"{col}14"]
        cell.value = num
        cell.font = f(9)
        cell.alignment = _center_wrap

    for row in (11, 12, 13, 14):
        for col in COLUMNS[1:]:
            ws[f"{col}{row}"].border = border_all


def style_data_row(ws, row, italic=False, fill=None):
    """Применяет границы, шрифт и заливку ко всей строке данных B..L."""
    pattern = PatternFill(start_color=fill, end_color=fill, fill_type="solid") if fill else None
    for col in COLUMNS[1:]:
        cell = ws[f"{col}{row}"]
        cell.border = border_all
        cell.font = f(9, italic=italic)
        cell.alignment = _left_center if col in ("C", "D", "L") else _center_wrap
        if pattern:
            cell.fill = pattern
