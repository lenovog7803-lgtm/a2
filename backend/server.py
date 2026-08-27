import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
logger.info("Starting server...")

from contextlib import asynccontextmanager
from fastapi import FastAPI, APIRouter, HTTPException, BackgroundTasks, Request, Depends, Query, WebSocket, WebSocketDisconnect
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os, uuid, asyncio, hashlib, secrets, time, html as _html
import jwt as _jwt
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, timezone, timedelta

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Google Sheets sync (lazy import to keep module-level safe even if libs missing)
try:
    from sheets_sync import trigger_sync as _sheets_trigger_sync, get_sync as _get_sheets_sync
except Exception as _e:  # pragma: no cover
    _sheets_trigger_sync = None
    _get_sheets_sync = None
    logging.getLogger(__name__).warning(f"sheets_sync import failed: {_e}")

try:
    from sheets_import import run_import as _sheets_run_import, get_last_status as _sheets_import_status, preview_import as _sheets_preview_import
except Exception as _e:  # pragma: no cover
    _sheets_run_import = None
    _sheets_import_status = None
    _sheets_preview_import = None
    logging.getLogger(__name__).warning(f"sheets_import import failed: {_e}")

try:
    from sheets_writer import (push_order as _sw_push_order, push_client as _sw_push_client,
                                push_carrier as _sw_push_carrier, delete_order as _sw_delete_order,
                                push_lead as _sw_push_lead, delete_lead as _sw_delete_lead, delete_client as _sw_delete_client)
except Exception as _e:  # pragma: no cover
    _sw_push_order = _sw_push_client = _sw_push_carrier = _sw_delete_order = _sw_push_lead = _sw_delete_lead = _sw_delete_client = None
    logging.getLogger(__name__).warning(f"sheets_writer import failed: {_e}")

try:
    from docs_gen import get_generator as _docs_get_generator, kind_to_field as _docs_kind_to_field, TEMPLATES as _DOC_TEMPLATES
except Exception as _e:  # pragma: no cover
    _docs_get_generator = None
    _docs_kind_to_field = None
    _DOC_TEMPLATES = {}
    logging.getLogger(__name__).warning(f"docs_gen import failed: {_e}")

try:
    from google_tasks import (create_google_task as _gt_create,
                               update_google_task as _gt_update,
                               delete_google_task as _gt_delete)
except Exception as _e:  # pragma: no cover
    _gt_create = _gt_update = _gt_delete = None
    logging.getLogger(__name__).warning(f"google_tasks import failed: {_e}")

try:
    from google_calendar import (create_calendar_event as _gc_create,
                                  update_calendar_event as _gc_update,
                                  delete_calendar_event as _gc_delete,
                                  create_simple_calendar_event as _gc_simple)
except Exception as _e:  # pragma: no cover
    _gc_create = _gc_update = _gc_delete = _gc_simple = None
    logging.getLogger(__name__).warning(f"google_calendar import failed: {_e}")

try:
    from oauth_google import (
        build_auth_url as _oauth_build_auth_url,
        fetch_token as _oauth_fetch_token,
        get_redirect_uri as _oauth_get_redirect,
        SCOPES as _OAUTH_SCOPES,
    )
except Exception as _e:  # pragma: no cover
    _oauth_build_auth_url = None
    _oauth_fetch_token = None
    _oauth_get_redirect = None
    _OAUTH_SCOPES = []
    logging.getLogger(__name__).warning(f"oauth_google import failed: {_e}")


async def _bg_sheets_sync():
    """Полная пересинхронизация ОТКЛЮЧЕНА (опасно — может перезаписать).
    Вместо этого — точечный upsert в _bg_push_order/_bg_push_client/_bg_push_carrier.
    """
    return  # no-op


async def _bg_push_order(order_obj: dict):
    if _sw_push_order is None:
        return
    try:
        await _sw_push_order(order_obj)
    except Exception as e:
        logging.getLogger(__name__).error(f"push_order bg failed: {e}")


async def _bg_push_client(client_obj: dict):
    if _sw_push_client is None:
        return
    try:
        await _sw_push_client(client_obj)
    except Exception as e:
        logging.getLogger(__name__).error(f"push_client bg failed: {e}")


async def _bg_push_carrier(carrier_obj: dict):
    if _sw_push_carrier is None:
        return
    try:
        await _sw_push_carrier(carrier_obj)
    except Exception as e:
        logging.getLogger(__name__).error(f"push_carrier bg failed: {e}")


async def _bg_push_lead(lead_obj: dict):
    if _sw_push_lead is None:
        return
    try:
        await _sw_push_lead(lead_obj)
    except Exception as e:
        logging.getLogger(__name__).error(f"push_lead bg failed: {e}", exc_info=True)


async def _bg_delete_lead(name: str):
    if not name or _sw_delete_lead is None:
        return
    try:
        await _sw_delete_lead(name)
    except Exception as e:
        logging.getLogger(__name__).error(f"delete_lead bg failed: {e}", exc_info=True)

async def _bg_delete_client(name: str):
    if not name or _sw_delete_client is None:
        return
    try:
        await _sw_delete_client(name)
    except Exception as e:
        logging.getLogger(__name__).error(f"delete_client bg failed: {e}", exc_info=True)


async def _bg_delete_order_row(order_number: str):
    if _sw_delete_order is None or not order_number:
        return
    try:
        await _sw_delete_order(order_number)
    except Exception as e:
        logging.getLogger(__name__).error(f"delete_order bg failed: {e}")


async def _auto_sync_loop():
    while True:
        await asyncio.sleep(600)  # каждые 10 минут
        # keep-alive ping
        try:
            import httpx
            async with httpx.AsyncClient() as client:
                await client.get("https://logistics-crm-backend.onrender.com/api/ping", timeout=10)
        except Exception:
            pass
        if _sheets_run_import is None:
            continue
        try:
            # skip_delete: background cycle only ever updates/creates fields
            # from the Sheet — never soft-deletes. Deletions only happen via
            # the director's manual, preview-then-confirm import.
            await _sheets_run_import(db, mode="merge", skip_delete=True)
        except Exception as e:
            logging.getLogger(__name__).error(f"auto sync failed: {e}")


async def _background_sheets_sync():
    """Первая синхронизация после старта — не блокирует ответ health-check'у.
    Даём Render/uvicorn несколько секунд, чтобы принять первые запросы, до похода в Google Sheets.
    """
    await asyncio.sleep(15)
    if _sheets_run_import is None:
        return
    try:
        result = await _sheets_run_import(db, mode="merge", skip_delete=True)
        logging.getLogger(__name__).info(f"[Sheets background sync] {result.get('imported')}")
    except Exception as e:
        logging.getLogger(__name__).error(f"[Sheets background sync] failed: {e}")


_INDEX_VERSION = 4  # bump when the index set below changes, to bypass the 24h skip below once


async def ensure_indexes():
    marker = await db.app_settings.find_one({"key": "indexes_ensured_at"})
    if marker and marker.get("version") == _INDEX_VERSION:
        last = datetime.fromisoformat(marker["value"])
        if datetime.now(timezone.utc) - last < timedelta(hours=24):
            logging.getLogger(__name__).info("Indexes checked recently, skipping")
            return

    await db.orders.create_index("id", background=True)
    await db.orders.create_index("order_number", background=True)
    await db.orders.create_index([("created_at", -1)], background=True)
    await db.orders.create_index("client_id", background=True)
    await db.orders.create_index("carrier_id", background=True)
    await db.orders.create_index([("deleted", 1), ("status", 1)], background=True)
    await db.orders.create_index("load_date", background=True)

    await db.clients.create_index("id", background=True)
    await db.clients.create_index("name", background=True)
    await db.clients.create_index([("deleted", 1), ("created_at", -1)], background=True)

    await db.carriers.create_index("id", background=True)
    await db.carriers.create_index("company_name", background=True)
    await db.carriers.create_index([("deleted", 1), ("created_at", -1)], background=True)

    await db.leads.create_index("id", background=True)
    await db.leads.create_index("phone", background=True)
    await db.leads.create_index([("stage", 1), ("next_call", 1)], background=True)
    await db.leads.create_index("industry", background=True)
    await db.leads.create_index("assigned_to", background=True)
    await db.leads.create_index([("deleted", 1), ("created_at", -1)], background=True)

    await db.tasks.create_index([("status", 1), ("deadline", 1)], background=True)
    await db.tasks.create_index([("created_at", -1)], background=True)
    # payment_reminder task lookup in _sync_payment_reminders (once/hour, per order+side)
    await db.tasks.create_index([("order_id", 1), ("type", 1), ("side", 1), ("status", 1)], background=True)
    await db.bot_subscribers.create_index("user_id", background=True)
    await db.reports.create_index([("period", 1), ("generated_at", -1)], background=True)
    await db.call_logs.create_index([("lead_id", 1), ("created_at", -1)], background=True)
    await db.call_logs.create_index([("created_by", 1), ("created_at", -1)], background=True)
    await db.payments_in.create_index("date", background=True)
    await db.payments_out.create_index("date", background=True)

    # _check_session_active() looks sessions up by "id" on literally every
    # authenticated request — this was the one lookup on that collection
    # with no supporting index, so it degraded as the (never-purged)
    # sessions collection grew over months of use.
    await db.sessions.create_index("id", background=True)
    await db.sessions.create_index("user_id", background=True)
    await db.sessions.create_index("active", background=True)
    await db.sessions.create_index([("active", 1), ("last_activity", -1)], background=True)
    await db.notifications.create_index([("read", 1), ("created_at", -1)], background=True)

    await db.kudir_entries.create_index([("entry_date", 1)], background=True)
    await db.kudir_entries.create_index("order_id", background=True)
    await db.kudir_entries.create_index("order_ids", background=True)

    await db.client_pp_ledger.create_index("client_id", background=True)
    await db.client_pp_ledger.create_index([("pp_date", 1)], background=True)

    try:
        await db.users.create_index("login", unique=True, background=True)
    except Exception as e:
        # Non-fatal: a duplicate login already in the data would make this
        # index build fail — logged for cleanup, doesn't block the rest.
        logging.getLogger(__name__).warning(f"users.login unique index failed (likely duplicate logins): {e}")

    await db.app_settings.update_one(
        {"key": "indexes_ensured_at"},
        {"$set": {"key": "indexes_ensured_at", "value": datetime.now(timezone.utc).isoformat(), "version": _INDEX_VERSION}},
        upsert=True,
    )
    logging.getLogger(__name__).info("Extended indexes ensured")


# ====== Backup helpers ======
_BACKUP_COLLECTIONS = ["orders", "clients", "carriers", "leads", "tasks", "users", "payments_in", "payments_out"]
_TRASH_COLLECTIONS = ["orders", "clients", "carriers", "leads"]


async def _create_backup(reason: str = "scheduled"):
    snapshot: dict = {}
    for cname in _BACKUP_COLLECTIONS:
        docs = await db[cname].find({}, {"_id": 0}).to_list(100000)
        snapshot[cname] = docs
    backup_doc = {
        "id": str(uuid.uuid4()),
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "created_at": now_iso(),
        "reason": reason,
        "counts": {k: len(v) for k, v in snapshot.items()},
        "collections": snapshot,
    }
    await db.backups.insert_one(backup_doc)
    # Keep only last 30 backups
    all_bk = await db.backups.find({}, {"_id": 1, "created_at": 1}).sort("created_at", -1).to_list(100)
    if len(all_bk) > 30:
        await db.backups.delete_many({"_id": {"$in": [b["_id"] for b in all_bk[30:]]}})
    logging.getLogger(__name__).info(f"Backup created: {backup_doc['id']} reason={reason} counts={backup_doc['counts']}")
    return backup_doc["id"]


async def _startup_backup():
    await asyncio.sleep(60)
    try:
        await _create_backup(reason="startup")
    except Exception as e:
        logging.getLogger(__name__).error(f"Startup backup failed: {e}")


async def _purge_old_trash():
    cutoff = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    for cname in _TRASH_COLLECTIONS:
        await db[cname].delete_many({"deleted": True, "deleted_at": {"$lt": cutoff}})
    logging.getLogger(__name__).info("Trash auto-purge completed")


async def _token_refresh_loop():
    """Refresh the stored Google OAuth token every 30 minutes so it never expires."""
    while True:
        try:
            token_doc = await db.oauth_tokens.find_one({"_id": "google"}, {"_id": 0})
            if not token_doc or not token_doc.get("refresh_token"):
                continue
            from oauth_google import make_user_credentials
            _creds, new_token = await asyncio.to_thread(make_user_credentials, token_doc)
            if new_token:
                await db.oauth_tokens.update_one(
                    {"_id": "google"},
                    {"$set": {"access_token": new_token, "updated_at": datetime.now(timezone.utc).isoformat()}},
                )
                logging.getLogger(__name__).info("Google OAuth token refreshed automatically")
        except Exception as e:
            logging.getLogger(__name__).error(f"_token_refresh_loop: refresh failed: {e}")
        await asyncio.sleep(1800)  # 30 minutes


async def _backup_loop():
    last_date = None
    while True:
        now = datetime.now(timezone.utc)
        if now.hour == 23 and now.minute == 0 and last_date != now.date():
            last_date = now.date()
            try:
                await _create_backup(reason="scheduled")
            except Exception as e:
                logging.getLogger(__name__).error(f"Nightly backup failed: {e}")
        await asyncio.sleep(60)


async def _trash_purge_loop():
    last_date = None
    while True:
        now = datetime.now(timezone.utc)
        if now.hour == 3 and now.minute == 0 and last_date != now.date():
            last_date = now.date()
            try:
                await _purge_old_trash()
            except Exception as e:
                logging.getLogger(__name__).error(f"Trash purge failed: {e}")
        await asyncio.sleep(60)


# ======================================================================
# А2 Инфо СРМ — отдельный Telegram-бот (не связан со старым ботом заявок).
# Токен в переменной окружения A2_INFO_BOT_TOKEN. Получатели (chat_id)
# хранятся в коллекции bot_subscribers, а не в env — чтобы подключать
# менеджеров без передеплоя.
#   bot_subscribers: {'id': str, 'user_id': str, 'telegram_chat_id': str, 'role': str}
# ======================================================================
async def send_telegram_a2info(chat_id: str, text: str) -> dict:
    """Возвращает {'ok': bool, ...} — чтобы вызывающий (ручной запуск отчёта)
    мог показать, дошло ли сообщение и с какой ошибкой."""
    import httpx
    token = os.environ.get('A2_INFO_BOT_TOKEN')
    if not token:
        logger.warning('[a2info bot] no token configured')
        return {'chat_id': chat_id, 'ok': False, 'error': 'A2_INFO_BOT_TOKEN не задан'}
    if not chat_id:
        return {'chat_id': chat_id, 'ok': False, 'error': 'нет chat_id'}
    async with httpx.AsyncClient() as _client:
        try:
            r = await _client.post(
                f'https://api.telegram.org/bot{token}/sendMessage',
                json={'chat_id': chat_id, 'text': text, 'parse_mode': 'HTML'},
                timeout=10,
            )
            body = {}
            try:
                body = r.json()
            except Exception:
                pass
            if not body.get('ok'):
                logger.error(f'[a2info bot] send to {chat_id} failed: {r.status_code} {r.text[:300]}')
            return {'chat_id': chat_id, 'ok': bool(body.get('ok')),
                    'status': r.status_code, 'error': body.get('description')}
        except Exception as e:
            logger.error(f'[a2info bot] send failed: {e}')
            return {'chat_id': chat_id, 'ok': False, 'error': str(e)}


async def notify_user(user_id: str, text: str) -> list:
    """Отправить уведомление конкретному пользователю во все его подписанные чаты."""
    if not user_id:
        return []
    subs = await db.bot_subscribers.find({'user_id': user_id}).to_list(10)
    out = []
    for s in subs:
        out.append(await send_telegram_a2info(s.get('telegram_chat_id'), text))
    return out


_REMINDER_WINDOW_DAYS = int(os.environ.get('PAYMENT_REMINDER_WINDOW_DAYS', '3'))


async def _run_payment_reminder_sync() -> dict:
    """Один проход: по каждой неоплаченной заявке, где до срока оплаты
    <= _REMINDER_WINDOW_DAYS дней (или уже просрочено), создаёт/обновляет
    одну задачу type=payment_reminder. Срок client-стороны = дата выгрузки,
    carrier-стороны = carrier_payment_deadline. Закрывается в mark_payment.
    Возвращает {'created': [task, ...], 'updated': n}."""
    now = datetime.now(_REPORT_TZ)
    today = now.date()
    today_str = now.strftime('%Y-%m-%d')
    created: list = []
    updated = 0

    orders = await db.orders.find({
        'deleted': {'$ne': True}, 'status': {'$ne': 'cancelled'},
        '$or': [
            {'client_paid': False},
            {'carrier_paid': False, 'carrier_payment_deadline': {'$ne': ''}},
        ],
    }).to_list(10000)

    for o in orders:
        for side in ('client', 'carrier'):
            if o.get(f'{side}_paid'):
                continue
            due = o.get('carrier_payment_deadline') if side == 'carrier' else o.get('unload_date')
            if not due:
                continue
            due_date = due[:10]
            try:
                d = datetime.fromisoformat(due_date).date()
            except Exception:
                continue
            days_left = (d - today).days
            if days_left > _REMINDER_WINDOW_DAYS:
                continue

            existing = await db.tasks.find_one({
                'order_id': o['id'], 'type': 'payment_reminder',
                'side': side, 'status': 'pending',
            })

            who = 'перевозчику' if side == 'carrier' else 'от клиента'
            if due_date <= today_str:
                overdue = (today - d).days
                title = (f"Оплати!: {who} по заявке {o['order_number']}"
                         + (f" (просрочено {overdue} дн.)" if overdue > 0 else ""))
            else:
                title = f"Через {days_left} дн. оплата {who} по заявке {o['order_number']}"

            if existing:
                if existing.get('title') != title or existing.get('due_date') != due_date:
                    await db.tasks.update_one(
                        {'id': existing['id']},
                        {'$set': {'title': title, 'due_date': due_date}},
                    )
                    updated += 1
            else:
                assigned = o.get('assigned_to') or o.get('created_by')
                task = {
                    'id': str(uuid.uuid4()), 'order_id': o['id'], 'order_number': o['order_number'],
                    'type': 'payment_reminder', 'task_type': 'payment', 'side': side,
                    'due_date': due_date, 'title': title, 'status': 'pending',
                    'assigned_user_id': assigned or None,
                    'created_by': assigned or '',
                    'created_at': now.isoformat(), 'completed_at': None,
                }
                await db.tasks.insert_one(task)
                created.append(task)

    return {'created': created, 'updated': updated}


async def _notify_new_payment_reminders(created: list) -> list:
    """Одно сводное сообщение в Telegram по новым напоминаниям об оплате —
    директорам-подписчикам и на дефолтный chat_id (вместо десятков пингов)."""
    if not created:
        return []
    rows = [f"•  {_esc(t['title'])}" for t in created[:40]]
    if len(created) > 40:
        rows.append(f"…и ещё {len(created) - 40}")
    text = f"🔔 <b>Новые напоминания об оплате · {len(created)}</b>\n" + "\n".join(rows)

    targets: list = list(_DEFAULT_REPORT_CHAT_IDS)
    directors = await db.users.find({'role': {'$in': list(DIRECTOR_ROLES)}}).to_list(20)
    for dd in directors:
        for s in await db.bot_subscribers.find({'user_id': dd['id']}).to_list(10):
            if s.get('telegram_chat_id'):
                targets.append(s['telegram_chat_id'])

    seen: set = set()
    out: list = []
    for cid in targets:
        if cid in seen:
            continue
        seen.add(cid)
        out.append(await send_telegram_a2info(cid, text))
    return out


async def _sync_payment_reminders():
    """Первый проход — почти сразу при старте, дальше раз в час."""
    await asyncio.sleep(25)
    while True:
        try:
            res = await _run_payment_reminder_sync()
            if res['created']:
                await _notify_new_payment_reminders(res['created'])
            if res['created'] or res['updated']:
                logger.info(f"[payment reminders] created {len(res['created'])}, updated {res['updated']}")
        except Exception as e:
            logger.error(f'[payment reminders] {e}')
        await asyncio.sleep(3600)


# ====== Отчёты: ежедневный / еженедельный / ежемесячный ======
# Всё в отчётах считается по местному времени (Минск, UTC+3, без летнего
# времени) — «сегодня» это сегодня по Минску, и рассылка в 21:00 по Минску.
_REPORT_TZ = timezone(timedelta(hours=3))
_REPORT_HOUR = 21


async def build_report(period: str) -> dict:
    now = datetime.now(_REPORT_TZ)  # местное время (Минск) — «сегодня» = сегодня по Минску
    if period == 'daily':
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == 'weekly':
        start = now - timedelta(days=7)
    else:
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    start_str = start.strftime('%Y-%m-%d')
    today_str = now.strftime('%Y-%m-%d')
    # Верхняя граница — конец сегодняшнего дня (будущие заявки не считаем).
    end_excl_str = (now + timedelta(days=1)).strftime('%Y-%m-%d')

    all_orders = await db.orders.find({'deleted': {'$ne': True}, 'status': {'$ne': 'cancelled'}}).to_list(10000)

    # Выручка/маржа — по заявкам, СОЗДАННым в периоде (по created_at).
    # created_at — ISO-строка (обычно UTC), сравниваем по датовому префиксу
    # YYYY-MM-DD, чего достаточно и устойчиво к разным форматам в базе.
    orders = [o for o in all_orders
              if start_str <= (o.get('created_at') or '')[:10] < end_excl_str]

    revenue = sum(float(o.get('client_rate') or 0) for o in orders)
    margin = sum(float(o.get('client_rate') or 0) - float(o.get('carrier_rate') or 0) for o in orders)
    # «Доставлено» — по дате выгрузки в периоде (а не по дате создания),
    # иначе в дневном отчёте почти всегда 0.
    delivered = sum(1 for o in all_orders if o.get('status') == 'done'
                    and start_str <= (o.get('unload_date') or '')[:10] < end_excl_str)
    overdue_carrier = [o for o in all_orders
                       if not o.get('carrier_paid') and o.get('carrier_payment_deadline', '')
                       and o.get('carrier_payment_deadline', '') < today_str]
    debtors = [o for o in all_orders if not o.get('client_paid') and o.get('status') == 'done']

    calls = await db.call_logs.find({'created_at': {'$gte': start_str}}).to_list(10000)
    won = sum(1 for c in calls if c.get('outcome') == 'won')

    # Завтрашние загрузки/выгрузки — какие заявки и контрагенты (для
    # ежедневного отчёта: "завтра выгрузка по заявке X — клиент …, перевозчик …").
    tomorrow_str = (now + timedelta(days=1)).strftime('%Y-%m-%d')

    def _leg(o):
        return {
            'order_number': o.get('order_number') or '',
            'client_name': o.get('client_name') or '',
            'carrier_name': o.get('carrier_name') or '',
            'route_from': o.get('route_from') or '',
            'route_to': o.get('route_to') or '',
        }

    tomorrow_loads = [_leg(o) for o in all_orders if (o.get('load_date') or '')[:10] == tomorrow_str]
    tomorrow_unloads = [_leg(o) for o in all_orders if (o.get('unload_date') or '')[:10] == tomorrow_str]

    client_totals: dict = {}
    for o in orders:
        n = o.get('client_name', '') or '—'
        client_totals[n] = client_totals.get(n, 0) + float(o.get('client_rate') or 0)
    top_clients = sorted(client_totals.items(), key=lambda x: -x[1])[:5]

    return {
        'id': str(uuid.uuid4()), 'period': period,
        'revenue': revenue, 'margin': margin, 'delivered': delivered,
        'orders_count': len(orders),
        'overdue_carrier_count': len(overdue_carrier),
        'overdue_carrier_sum': sum(float(o.get('carrier_rate') or 0) for o in overdue_carrier),
        'debtors_count': len(debtors),
        'debtors_sum': sum(float(o.get('client_rate') or 0) for o in debtors),
        'calls_total': len(calls), 'calls_won': won,
        'top_clients': [[n, s] for n, s in top_clients],
        'tomorrow_date': tomorrow_str,
        'tomorrow_loads': tomorrow_loads,
        'tomorrow_unloads': tomorrow_unloads,
        'generated_at': now.isoformat(),
    }


def _esc(s) -> str:
    """HTML-escape для parse_mode=HTML в Telegram — иначе '&' или '<' в
    названии клиента/перевозчика ломают всё сообщение (Telegram вернёт 400)."""
    return _html.escape(str(s if s is not None else ''))


def _byn(x) -> str:
    try:
        return f"{round(float(x or 0)):,}".replace(',', ' ') + ' BYN'
    except Exception:
        return f"{x} BYN"


_REPORT_HR = '━━━━━━━━━━━━━━'


def _fmt_leg(x: dict) -> str:
    route = ' → '.join(_esc(p) for p in (x.get('route_from'), x.get('route_to')) if p)
    line = f"•  <b>{_esc(x.get('order_number') or '—')}</b>" + (f"   {route}" if route else '')
    sub = []
    if x.get('client_name'):
        sub.append(f"клиент: {_esc(x['client_name'])}")
    if x.get('carrier_name'):
        sub.append(f"перевозчик: {_esc(x['carrier_name'])}")
    if sub:
        line += "\n     <i>" + '  ·  '.join(sub) + "</i>"
    return line


def format_report_text(report: dict) -> str:
    labels = {'daily': 'за сегодня', 'weekly': 'за неделю', 'monthly': 'за месяц'}
    period = report.get('period')
    delivered = report.get('delivered', 0)
    total_orders = report.get('orders_count', 0)

    L = [
        f"📊  <b>Отчёт {labels.get(period, period)}</b>",
        f"<i>{_esc(fmt_date_ru(report.get('generated_at', '')))}</i>",
        _REPORT_HR,
        f"💰  Выручка:  <b>{_byn(report.get('revenue'))}</b>",
        f"📈  Маржа:  <b>{_byn(report.get('margin'))}</b>",
        f"🆕  Создано заявок:  <b>{total_orders}</b>",
        f"✅  Доставлено:  <b>{delivered}</b>",
        "",
        f"⚠️  Просрочка перевозчикам:  <b>{report.get('overdue_carrier_count', 0)}</b>  ·  {_byn(report.get('overdue_carrier_sum'))}",
        f"🧾  Должники (клиенты):  <b>{report.get('debtors_count', 0)}</b>  ·  {_byn(report.get('debtors_sum'))}",
        f"📞  Звонки:  <b>{report.get('calls_total', 0)}</b>  ·  клиентами стали {report.get('calls_won', 0)}",
    ]

    top = report.get('top_clients') or []
    if top:
        L.append("")
        L.append("🏆  <b>Топ клиентов</b>")
        L += [f"{i}.  {_esc(n)}  —  {_byn(s)}" for i, (n, s) in enumerate(top, 1)]

    if period == 'daily':
        loads = report.get('tomorrow_loads') or []
        unloads = report.get('tomorrow_unloads') or []
        L.append(_REPORT_HR)
        L.append(f"📅  <b>Завтра — {_esc(fmt_date_ru(report.get('tomorrow_date', '')))}</b>")
        L.append("")
        L.append(f"📦  <b>Загрузки · {len(loads)}</b>")
        L += [_fmt_leg(x) for x in loads] or ["<i>нет</i>"]
        L.append("")
        L.append(f"🏁  <b>Выгрузки · {len(unloads)}</b>")
        L += [_fmt_leg(x) for x in unloads] or ["<i>нет</i>"]

    return '\n'.join(L)


# chat_id, куда отчёт уходит всегда — даже если в CRM никто не подписан
# (поле ввода chat_id из интерфейса убрано). Переопределяется переменной
# окружения A2_INFO_REPORT_CHAT_IDS (через запятую).
_DEFAULT_REPORT_CHAT_IDS = [c.strip() for c in os.environ.get(
    'A2_INFO_REPORT_CHAT_IDS', '558556324').split(',') if c.strip()]


async def send_scheduled_report(period: str) -> dict:
    report = await build_report(period)
    await db.reports.insert_one(dict(report))
    text = format_report_text(report)

    targets: list = list(_DEFAULT_REPORT_CHAT_IDS)
    directors = await db.users.find({'role': {'$in': list(DIRECTOR_ROLES)}}).to_list(20)
    for d in directors:
        for s in await db.bot_subscribers.find({'user_id': d['id']}).to_list(10):
            cid = s.get('telegram_chat_id')
            if cid:
                targets.append(cid)

    seen: set = set()
    delivery: list = []
    for cid in targets:
        if cid in seen:
            continue
        seen.add(cid)
        delivery.append(await send_telegram_a2info(cid, text))

    ok = sum(1 for d in delivery if d.get('ok'))
    logger.info(f'[report {period}] delivered to {ok}/{len(delivery)} chats')
    return {'report': report, 'delivery': delivery, 'sent': ok, 'targets': len(delivery)}


async def _report_scheduler():
    sent_today: set = set()
    while True:
        now = datetime.now(_REPORT_TZ)
        key_day = now.strftime('%Y-%m-%d')

        try:
            if now.hour == _REPORT_HOUR and now.minute < 5 and f'daily-{key_day}' not in sent_today:
                await send_scheduled_report('daily')
                sent_today.add(f'daily-{key_day}')

            if now.weekday() == 4 and now.hour == _REPORT_HOUR and now.minute < 5 and f'weekly-{key_day}' not in sent_today:
                await send_scheduled_report('weekly')
                sent_today.add(f'weekly-{key_day}')

            tomorrow = now + timedelta(days=1)
            if tomorrow.day == 1 and now.hour == _REPORT_HOUR and now.minute < 5 and f'monthly-{key_day}' not in sent_today:
                await send_scheduled_report('monthly')
                sent_today.add(f'monthly-{key_day}')
        except Exception as e:
            logger.error(f'[report scheduler] {e}')

        if len(sent_today) > 20:
            sent_today.clear()

        await asyncio.sleep(300)


async def _deferred_init():
    """Всё, что трогает базу, уезжает сюда — health-check не должен ждать ни строчки из этого."""
    try:
        count = await db.users.count_documents({"role": "admin"})
        if count == 0:
            admin = {
                "id": str(uuid.uuid4()),
                "name": "Администратор",
                "login": "admin",
                "password_hash": _hash_password("admin123"),
                "role": "admin",
                "created_at": now_iso(),
            }
            await db.users.insert_one(admin)
            logging.getLogger(__name__).info("Created default admin user (login=admin, password=admin123)")
    except Exception as e:
        logging.getLogger(__name__).error(f"admin user init failed: {e}")

    try:
        await _migrate_leads_stage()
    except Exception as e:
        logging.getLogger(__name__).error(f"_migrate_leads_stage failed: {e}")

    await asyncio.sleep(5)
    try:
        await ensure_indexes()
    except Exception as e:
        logging.getLogger(__name__).error(f"ensure_indexes failed: {e}")

    # Disabled permanently: this loop pulls the Sheet's (stale) order_number
    # values back over CRM on every cycle, which has now twice silently
    # reverted a manual order-number correction mid-session. Re-enable only
    # once the Sheet itself is kept in sync with any manual CRM renames.
    # asyncio.create_task(_background_sheets_sync())
    # asyncio.create_task(_auto_sync_loop())
    asyncio.create_task(_backup_loop())
    asyncio.create_task(_startup_backup())
    asyncio.create_task(_trash_purge_loop())
    asyncio.create_task(_token_refresh_loop())
    asyncio.create_task(_check_stale_managers())
    asyncio.create_task(_sync_payment_reminders())
    asyncio.create_task(_report_scheduler())

    # Debug: print last 30 carriers to inspect field names
    try:
        _debug_carriers = await db.carriers.find({"deleted": {"$ne": True}}).sort("created_at", -1).to_list(30)
        print(f"[startup] carriers count={len(_debug_carriers)}")
        for _c in _debug_carriers[:5]:
            print({k: v for k, v in _c.items() if k not in ['_id']})
    except Exception as _ce:
        print(f"[startup] carriers debug error: {_ce}")


@asynccontextmanager
async def lifespan(app: FastAPI):
    logger.info("Starting server...")
    asyncio.create_task(_deferred_init())
    yield
    client.close()


app = FastAPI(lifespan=lifespan)
api_router = APIRouter(prefix="/api")


@app.middleware("http")
async def log_slow_requests(request: Request, call_next):
    start = time.time()
    response = await call_next(request)
    duration = time.time() - start
    if duration > 0.5:
        logger.warning(f"[SLOW] {request.method} {request.url.path} took {duration:.2f}s")
    return response


# ====== Realtime (WebSocket broadcast) ======
class ConnectionManager:
    def __init__(self):
        self.active: list[WebSocket] = []

    async def connect(self, ws: WebSocket):
        await ws.accept()
        self.active.append(ws)

    def disconnect(self, ws: WebSocket):
        if ws in self.active:
            self.active.remove(ws)

    async def broadcast(self, event: dict):
        dead = []
        for ws in self.active:
            try:
                await ws.send_json(event)
            except Exception:
                dead.append(ws)
        for ws in dead:
            self.disconnect(ws)


manager = ConnectionManager()


@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            await websocket.receive_text()  # держим соединение, пинги от клиента игнорируем
    except WebSocketDisconnect:
        manager.disconnect(websocket)
    except Exception:
        manager.disconnect(websocket)


TAX_RATE = 0.20  # 20% — для расчёта прибыли (маржа − налог)

JWT_SECRET = os.environ.get("JWT_SECRET", "crm-secret-key-change-in-production-2024")
JWT_ALGORITHM = "HS256"
JWT_EXPIRE_HOURS = 24 * 7

_http_bearer = HTTPBearer(auto_error=False)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


_RU_MONTHS = ["января", "февраля", "марта", "апреля", "мая", "июня",
              "июля", "августа", "сентября", "октября", "ноября", "декабря"]


def fmt_date_ru(date_str: str) -> str:
    """YYYY-MM-DD -> '3 января 2026'. Возвращает исходную строку, если формат неожиданный."""
    if not date_str:
        return ""
    try:
        y, m, d = date_str[:10].split("-")
        return f"{int(d)} {_RU_MONTHS[int(m) - 1]} {y}"
    except Exception:
        return date_str


def _hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    key = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 100000)
    return f"{salt}${key.hex()}"


def _verify_password(password: str, password_hash: str) -> bool:
    try:
        salt, key_hex = password_hash.split("$", 1)
        key = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 100000)
        return key.hex() == key_hex
    except Exception:
        return False


def _create_token(user_id: str, role: str, session_id: Optional[str] = None) -> str:
    payload = {
        "sub": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRE_HOURS),
    }
    if session_id:
        payload["sid"] = session_id
    return _jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


# Debounces last_activity writes for active sessions — pinged on ~every
# authenticated request, but we only want to touch the DB once a minute per
# session, not on every single API call.
_last_activity_touch: dict = {}


async def _touch_session(sid: str):
    now_ts = datetime.now(timezone.utc)
    last = _last_activity_touch.get(sid)
    if last and (now_ts - last).total_seconds() < 60:
        return
    _last_activity_touch[sid] = now_ts
    try:
        await db.sessions.update_one({"id": sid}, {"$set": {"last_activity": now_ts.isoformat()}})
    except Exception:
        pass


async def _check_session_active(sid: Optional[str]) -> bool:
    # Tokens issued before session tracking existed carry no "sid" — treat
    # them as always valid so already-logged-in users aren't forced out by
    # this deploy. Real session enforcement kicks in on their next login.
    if not sid:
        return True
    session = await db.sessions.find_one({"id": sid}, {"_id": 0, "active": 1})
    if not session or not session.get("active", False):
        return False
    asyncio.create_task(_touch_session(sid))
    return True


async def _get_user_from_token(credentials: HTTPAuthorizationCredentials = Depends(_http_bearer)) -> Optional[dict]:
    if not credentials:
        return None
    try:
        payload = _jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            return None
        if not await _check_session_active(payload.get("sid")):
            return None
        return await db.users.find_one({"id": user_id}, {"_id": 0})
    except Exception:
        return None


async def _require_user(credentials: HTTPAuthorizationCredentials = Depends(_http_bearer)) -> dict:
    if not credentials:
        raise HTTPException(401, "Требуется авторизация")
    try:
        payload = _jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(401, "Неверный токен")
        if not await _check_session_active(payload.get("sid")):
            raise HTTPException(401, "Сессия завершена")
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not user:
            raise HTTPException(401, "Пользователь не найден")
        return user
    except HTTPException:
        raise
    except _jwt.ExpiredSignatureError:
        raise HTTPException(401, "Токен истёк")
    except Exception:
        raise HTTPException(401, "Неверный токен")


DIRECTOR_ROLES = ("admin", "director")


def require_director(current_user: dict = Depends(_require_user)) -> dict:
    if current_user.get("role") not in DIRECTOR_ROLES:
        raise HTTPException(403, "Доступно только директору")
    return current_user


# ====== Models ======
class Client(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    contact_person: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    unp: Optional[str] = ""
    inn: Optional[str] = ""
    kpp: Optional[str] = ""
    director: Optional[str] = ""
    basis: Optional[str] = "Устава"
    legal_address: Optional[str] = ""
    postal_address: Optional[str] = ""
    bank_name: Optional[str] = ""
    bank_account: Optional[str] = ""
    bank_bik: Optional[str] = ""
    bank_corr_account: Optional[str] = ""
    payment_terms: Optional[str] = ""
    cargo_types: Optional[str] = ""
    directions: Optional[str] = ""
    notes: Optional[str] = ""
    city: Optional[str] = ""
    address: Optional[str] = ""
    total_revenue: Optional[float] = 0
    orders_count: Optional[int] = 0
    created_at: str = Field(default_factory=now_iso)


class ClientPayload(BaseModel):
    name: str
    contact_person: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    unp: Optional[str] = ""
    inn: Optional[str] = ""
    kpp: Optional[str] = ""
    director: Optional[str] = ""
    basis: Optional[str] = "Устава"
    legal_address: Optional[str] = ""
    postal_address: Optional[str] = ""
    bank_name: Optional[str] = ""
    bank_account: Optional[str] = ""
    bank_bik: Optional[str] = ""
    bank_corr_account: Optional[str] = ""
    payment_terms: Optional[str] = ""
    cargo_types: Optional[str] = ""
    directions: Optional[str] = ""
    notes: Optional[str] = ""
    city: Optional[str] = ""
    address: Optional[str] = ""


class Carrier(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_name: str
    driver_name: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    unp: Optional[str] = ""
    address: Optional[str] = ""
    rs: Optional[str] = ""
    bik: Optional[str] = ""
    bank: Optional[str] = ""
    inn: Optional[str] = ""
    kpp: Optional[str] = ""
    legal_address: Optional[str] = ""
    postal_address: Optional[str] = ""
    director: Optional[str] = ""
    basis: Optional[str] = ""
    bank_name: Optional[str] = ""
    bank_account: Optional[str] = ""
    bank_bik: Optional[str] = ""
    bank_corr_account: Optional[str] = ""
    vehicle_type: Optional[str] = ""
    plate: Optional[str] = ""
    capacity_tons: Optional[float] = 0
    capacity_m3: Optional[float] = 0
    cargo_types: Optional[str] = ""
    regions: Optional[str] = ""
    rating: Optional[float] = 5.0
    notes: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class CarrierPayload(BaseModel):
    company_name: str
    driver_name: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    unp: Optional[str] = ""
    address: Optional[str] = ""
    rs: Optional[str] = ""
    bik: Optional[str] = ""
    bank: Optional[str] = ""
    inn: Optional[str] = ""
    kpp: Optional[str] = ""
    legal_address: Optional[str] = ""
    postal_address: Optional[str] = ""
    director: Optional[str] = ""
    basis: Optional[str] = ""
    bank_name: Optional[str] = ""
    bank_account: Optional[str] = ""
    bank_bik: Optional[str] = ""
    bank_corr_account: Optional[str] = ""
    vehicle_type: Optional[str] = ""
    plate: Optional[str] = ""
    capacity_tons: Optional[float] = 0
    capacity_m3: Optional[float] = 0
    cargo_types: Optional[str] = ""
    regions: Optional[str] = ""
    rating: Optional[float] = 5.0
    notes: Optional[str] = ""


class Order(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    order_number: str
    client_id: Optional[str] = ""
    client_name: Optional[str] = ""
    carrier_id: Optional[str] = ""
    carrier_name: Optional[str] = ""
    route_from: str
    route_to: str
    route_from_address: Optional[str] = ""
    route_to_address: Optional[str] = ""
    load_date: Optional[str] = ""
    unload_date: Optional[str] = ""
    driver_name: Optional[str] = ""
    driver_phone: Optional[str] = ""
    vehicle_type: Optional[str] = ""
    vehicle_plate: Optional[str] = ""
    vehicle_info: Optional[str] = ""
    client_rate: float = 0
    carrier_rate: float = 0
    status: str = "new"
    client_paid: bool = False
    carrier_paid: bool = False
    client_paid_date: Optional[str] = ""
    carrier_paid_date: Optional[str] = ""
    # 4 раздельных статуса по документам
    docs_to_client_sent: bool = False       # Документы Клиенту — отправлены
    docs_from_client_received: bool = False  # Документы от Клиента — получены
    docs_to_carrier_sent: bool = False       # Документы Перевозчику — отправлены/получены им
    docs_from_carrier_received: bool = False # Документы от Перевозчика — получены/они отправили
    docs_to_client_date: Optional[str] = ""
    docs_from_client_date: Optional[str] = ""
    docs_to_carrier_date: Optional[str] = ""
    docs_from_carrier_date: Optional[str] = ""
    cargo: Optional[str] = ""
    weight_tons: Optional[float] = 0
    notes: Optional[str] = ""
    # Ссылки на сгенерированные документы Google Docs
    doc_url_client: Optional[str] = ""
    doc_url_carrier: Optional[str] = ""
    doc_url_act: Optional[str] = ""
    calendar_event_id: Optional[str] = ""
    calendar_event_url: Optional[str] = ""
    assigned_to: Optional[str] = ""
    created_by: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)
    is_overdue: Optional[bool] = False
    carrier_payment_days: Optional[int] = 20
    carrier_payment_deadline: Optional[str] = ""
    carrier_payment_reminder_date: Optional[str] = ""
    client_pp_number: Optional[str] = ""
    client_pp_date: Optional[str] = ""
    carrier_pp_number: Optional[str] = ""
    carrier_pp_date: Optional[str] = ""
    carrier_act_number: Optional[str] = ""  # номер акта перевозчика — вводится при получении документов
    carrier_act_date: Optional[str] = ""
    # Несколько частичных ПП на одну заявку. Пусто/отсутствует у старых
    # заявок — тогда список собирается на лету из client_pp_number/
    # client_paid_date/client_rate, см. _get_payments().
    client_payments: Optional[List[dict]] = None
    carrier_payments: Optional[List[dict]] = None
    # Marks a side as paid in cash — no PP number/date exists to record, so
    # the "missing ПП" and "amount doesn't match" checks skip these orders
    # instead of flagging a cash payment as an error.
    client_cash: Optional[bool] = False
    carrier_cash: Optional[bool] = False


class OrderPayload(BaseModel):
    order_number: Optional[str] = None
    client_id: Optional[str] = ""
    client_name: Optional[str] = ""
    carrier_id: Optional[str] = ""
    carrier_name: Optional[str] = ""
    route_from: str
    route_to: str
    route_from_address: Optional[str] = ""
    route_to_address: Optional[str] = ""
    load_date: Optional[str] = ""
    unload_date: Optional[str] = ""
    driver_name: Optional[str] = ""
    driver_phone: Optional[str] = ""
    vehicle_type: Optional[str] = ""
    vehicle_plate: Optional[str] = ""
    vehicle_info: Optional[str] = ""
    client_rate: float = 0
    carrier_rate: float = 0
    status: str = "new"
    client_paid: bool = False
    carrier_paid: bool = False
    client_paid_date: Optional[str] = ""
    carrier_paid_date: Optional[str] = ""
    docs_to_client_sent: bool = False
    docs_from_client_received: bool = False
    docs_to_carrier_sent: bool = False
    docs_from_carrier_received: bool = False
    docs_to_client_date: Optional[str] = ""
    docs_from_client_date: Optional[str] = ""
    docs_to_carrier_date: Optional[str] = ""
    docs_from_carrier_date: Optional[str] = ""
    cargo: Optional[str] = ""
    weight_tons: Optional[float] = 0
    notes: Optional[str] = ""
    doc_url_client: Optional[str] = ""
    doc_url_carrier: Optional[str] = ""
    doc_url_act: Optional[str] = ""
    assigned_to: Optional[str] = ""
    carrier_payment_days: Optional[int] = 20


class OrderUpdate(BaseModel):
    order_number: Optional[str] = None
    client_id: Optional[str] = None
    client_name: Optional[str] = None
    carrier_id: Optional[str] = None
    carrier_name: Optional[str] = None
    route_from: Optional[str] = None
    route_to: Optional[str] = None
    route_from_address: Optional[str] = None
    route_to_address: Optional[str] = None
    load_date: Optional[str] = None
    unload_date: Optional[str] = None
    driver_name: Optional[str] = None
    driver_phone: Optional[str] = None
    vehicle_type: Optional[str] = None
    vehicle_plate: Optional[str] = None
    vehicle_info: Optional[str] = None
    client_rate: Optional[float] = None
    carrier_rate: Optional[float] = None
    status: Optional[str] = None
    client_paid: Optional[bool] = None
    carrier_paid: Optional[bool] = None
    client_paid_date: Optional[str] = None
    carrier_paid_date: Optional[str] = None
    docs_to_client_sent: Optional[bool] = None
    docs_from_client_received: Optional[bool] = None
    docs_to_carrier_sent: Optional[bool] = None
    docs_from_carrier_received: Optional[bool] = None
    docs_to_client_date: Optional[str] = None
    docs_from_client_date: Optional[str] = None
    docs_to_carrier_date: Optional[str] = None
    docs_from_carrier_date: Optional[str] = None
    cargo: Optional[str] = None
    weight_tons: Optional[float] = None
    notes: Optional[str] = None
    doc_url_client: Optional[str] = None
    doc_url_carrier: Optional[str] = None
    doc_url_act: Optional[str] = None
    calendar_event_id: Optional[str] = None
    calendar_event_url: Optional[str] = None
    assigned_to: Optional[str] = None
    carrier_payment_days: Optional[int] = None
    carrier_payment_deadline: Optional[str] = None
    carrier_payment_reminder_date: Optional[str] = None
    client_pp_number: Optional[str] = None
    client_pp_date: Optional[str] = None
    carrier_pp_number: Optional[str] = None
    carrier_pp_date: Optional[str] = None
    carrier_act_number: Optional[str] = None
    carrier_act_date: Optional[str] = None
    client_payments: Optional[List[dict]] = None
    carrier_payments: Optional[List[dict]] = None
    client_cash: Optional[bool] = None
    carrier_cash: Optional[bool] = None


class Lead(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    company: Optional[str] = ""
    phone: str
    contact_person: Optional[str] = ""
    contact_position: Optional[str] = ""
    email: Optional[str] = ""
    website: Optional[str] = ""
    industry: Optional[str] = ""
    city: Optional[str] = ""
    region: Optional[str] = ""
    stage: str = "new"
    last_contact: Optional[str] = ""
    last_call: Optional[str] = ""
    next_call: Optional[str] = None
    notes: Optional[str] = ""
    directions: Optional[str] = ""
    call_notes: Optional[List] = Field(default_factory=list)
    call_attempts: Optional[int] = 0
    total_calls: Optional[int] = 0
    cadence_step: Optional[int] = 0
    first_call_at: Optional[str] = None
    last_call_at: Optional[str] = None
    won_at: Optional[str] = None
    lost_reason: Optional[str] = None
    client_id: Optional[str] = None
    assigned_to: Optional[str] = ""
    assigned_at: Optional[str] = None
    created_at: str = Field(default_factory=now_iso)
    updated_at: Optional[str] = None
    stage_changed_at: Optional[str] = None


class LeadPayload(BaseModel):
    name: str
    company: Optional[str] = ""
    phone: str
    contact_person: Optional[str] = ""
    contact_position: Optional[str] = ""
    email: Optional[str] = ""
    website: Optional[str] = ""
    industry: Optional[str] = ""
    city: Optional[str] = ""
    region: Optional[str] = ""
    stage: str = "new"
    last_contact: Optional[str] = ""
    next_call: Optional[str] = None
    notes: Optional[str] = ""
    directions: Optional[str] = ""
    call_notes: Optional[List] = Field(default_factory=list)
    assigned_to: Optional[str] = ""


class LeadUpdate(BaseModel):
    name: Optional[str] = None
    company: Optional[str] = None
    phone: Optional[str] = None
    contact_person: Optional[str] = None
    contact_position: Optional[str] = None
    email: Optional[str] = None
    website: Optional[str] = None
    industry: Optional[str] = None
    city: Optional[str] = None
    region: Optional[str] = None
    stage: Optional[str] = None
    last_contact: Optional[str] = None
    next_call: Optional[str] = None
    notes: Optional[str] = None
    directions: Optional[str] = None
    call_notes: Optional[List] = None
    lost_reason: Optional[str] = None
    assigned_to: Optional[str] = None


# ====== Lead call module: stages, cadence, scripts ======
LEAD_STAGES = [
    {"id": "new",         "label": "Новый",         "order": 1, "color": "#8A93A0", "active": True},
    {"id": "reached",     "label": "Дозвонился",    "order": 2, "color": "#1366F0", "active": True},
    {"id": "interested",  "label": "Заинтересован", "order": 3, "color": "#1366F0", "active": True},
    {"id": "thinking",    "label": "Думает",        "order": 4, "color": "#D97706", "active": True},
    {"id": "kp_sent",     "label": "КП отправлено", "order": 5, "color": "#7C3AED", "active": True},
    {"id": "negotiation", "label": "Переговоры",    "order": 6, "color": "#7C3AED", "active": True},
    {"id": "won",         "label": "Клиент",        "order": 7, "color": "#1E9E5A", "active": False},
    {"id": "no_contact",  "label": "Нет контакта",  "order": 8, "color": "#8A93A0", "active": False},
    {"id": "lost",        "label": "Отказ",         "order": 9, "color": "#E0473B", "active": False},
]
LEAD_STAGE_BY_ID = {s["id"]: s for s in LEAD_STAGES}

CADENCE = [
    {"step": 1, "action": "call",  "delay_days": 0, "label": "Первый звонок"},
    {"step": 2, "action": "call",  "delay_days": 2, "label": "Второй звонок"},
    {"step": 3, "action": "kp",    "delay_days": 1, "label": "Отправить КП"},
    {"step": 4, "action": "call",  "delay_days": 3, "label": "Звонок после КП"},
    {"step": 5, "action": "close", "delay_days": 5, "label": "Закрытие — решение"},
]


def next_cadence_touch(current_step: int):
    """Вернуть следующий шаг каденции и дату."""
    nxt = current_step + 1
    step_cfg = next((s for s in CADENCE if s["step"] == nxt), None)
    if not step_cfg:
        return None, None
    d = datetime.now(timezone.utc) + timedelta(days=step_cfg["delay_days"])
    d = d.replace(hour=10, minute=0, second=0, microsecond=0)
    return nxt, d.isoformat()


LOST_REASONS = [
    "Дорого",
    "Есть свой перевозчик",
    "Не занимаемся логистикой",
    "Не тот профиль груза",
    "Не наши направления",
    "Просят не звонить",
    "Компания не работает",
    "Другое",
]

DEFAULT_SCRIPTS = {
    "new": "Добрый день! Меня зовут [имя], компания А2 Групп, занимаемся грузоперевозками Беларусь–Россия.\n\nПодскажите, с кем можно поговорить по вопросам логистики?\n\n— Вы отправляете грузы в Россию или из России?\n— Кто сейчас возит?\n— Какие направления чаще всего?",
    "reached": "Спасибо что уделили время.\n\n— Какие объёмы отправляете в месяц?\n— Какой тип груза?\n— Что не устраивает в текущем перевозчике?\n\nМы работаем по маршрутам Москва–Минск и обратно, свои проверенные перевозчики, документы в порядке, оплата по факту доставки.",
    "interested": "Давайте я подготовлю расчёт под ваши направления.\n\n— На какие маршруты посчитать?\n— Какой средний вес отправки?\n— Тент, реф или изотерм?\n\nПришлю КП сегодня-завтра на почту.",
    "thinking": "Звоню уточнить, посмотрели наше предложение?\n\n— Что смущает по цене или условиям?\n— С чем сравниваете?\n\nГотов обсудить условия, если по цене не проходим — скажите ориентир.",
    "kp_sent": "Отправлял вам КП [дата], хотел уточнить — дошло?\n\n— Успели посмотреть цифры?\n— Есть вопросы по условиям оплаты или срокам?\n\nЕсли по каким-то направлениям цена не подходит — скажите, посмотрю что можно сделать.",
    "negotiation": "По условиям договорились, давайте закрепим.\n\n— Когда планируете первую отправку?\n— Пришлю договор на согласование, нужны реквизиты.\n— Кто будет контактным лицом по заявкам?",
}

LEAD_STAGE_MIGRATION = {
    "new": "new", "thinking": "thinking", "sent_kp": "kp_sent", "callback": "reached",
    "won": "won", "lost": "lost", "no_contact": "no_contact",
}


async def _migrate_leads_stage():
    """Одноразовая (идемпотентная) миграция поля status -> stage у лидов."""
    try:
        docs = await db.leads.find({"status": {"$exists": True}}, {"_id": 1, "status": 1}).to_list(100000)
        for d in docs:
            new_stage = LEAD_STAGE_MIGRATION.get(d.get("status"), d.get("status") or "new")
            await db.leads.update_one({"_id": d["_id"]}, {"$set": {"stage": new_stage}, "$unset": {"status": ""}})
        if docs:
            logging.getLogger(__name__).info(f"_migrate_leads_stage: migrated {len(docs)} leads")
    except Exception as e:
        logging.getLogger(__name__).error(f"_migrate_leads_stage failed: {e}")


# ===== CRUD helper =====
def make_crud(prefix: str, collection: str, ModelCls, PayloadCls, sync_to_sheets: bool = False, user_filter: bool = False, soft_delete: bool = False):
    @api_router.get(f"/{prefix}", response_model=List[ModelCls])
    async def list_items(current_user: Optional[dict] = Depends(_get_user_from_token)):
        filter_q: dict = {}
        if soft_delete:
            filter_q["deleted"] = {"$ne": True}
        if user_filter and current_user and current_user.get("role") == "manager":
            perms = current_user.get("permissions") or {}
            if not perms.get("can_view_all_leads"):
                filter_q["assigned_to"] = current_user["id"]
        docs = await db[collection].find(filter_q, {"_id": 0}).sort("created_at", -1).to_list(50000)
        return [ModelCls(**d) for d in docs]

    @api_router.post(f"/{prefix}", response_model=ModelCls)
    async def create_item(payload: PayloadCls, background_tasks: BackgroundTasks):
        obj = ModelCls(**payload.dict())
        await db[collection].insert_one(obj.dict())
        if sync_to_sheets:
            if collection == "clients":
                background_tasks.add_task(_bg_push_client, obj.dict())
            elif collection == "carriers":
                background_tasks.add_task(_bg_push_carrier, obj.dict())
            elif collection == "leads":
                background_tasks.add_task(_bg_push_lead, obj.dict())
        return obj

    @api_router.get(f"/{prefix}/{{item_id}}", response_model=ModelCls)
    async def get_item(item_id: str):
        q: dict = {"id": item_id}
        if soft_delete:
            q["deleted"] = {"$ne": True}
        doc = await db[collection].find_one(q, {"_id": 0})
        if not doc:
            try:
                from bson import ObjectId
                q2: dict = {"_id": ObjectId(item_id)}
                if soft_delete:
                    q2["deleted"] = {"$ne": True}
                doc = await db[collection].find_one(q2, {"_id": 0})
            except Exception:
                pass
        if not doc:
            raise HTTPException(404, "Not found")
        return ModelCls(**doc)

    @api_router.put(f"/{prefix}/{{item_id}}", response_model=ModelCls)
    async def update_item(item_id: str, payload: PayloadCls, background_tasks: BackgroundTasks,
                          current_user: Optional[dict] = Depends(_get_user_from_token)):
        old_status = None
        if collection == "leads":
            old_doc = await db[collection].find_one({"id": item_id}, {"_id": 0})
            old_status = (old_doc or {}).get("stage")
        upd = payload.dict(exclude_none=True)
        if collection == "leads" and "stage" in upd and upd["stage"] != old_status:
            upd["stage_changed_at"] = datetime.now(timezone.utc).isoformat()
        await db[collection].update_one({"id": item_id}, {"$set": upd})
        doc = await db[collection].find_one({"id": item_id}, {"_id": 0})
        if not doc:
            raise HTTPException(404, "Not found")
        if collection == "leads":
            new_status = doc.get("stage")
            if new_status != old_status:
                action = "status_change"
            elif any(k in upd for k in ("call_notes", "last_contact", "next_call")):
                action = "call"
            else:
                action = "update"
            await db.lead_activity.insert_one({
                "id": str(uuid.uuid4()),
                "lead_id": item_id,
                "user_id": (current_user or {}).get("id", ""),
                "action": action,
                "old_status": old_status,
                "new_status": new_status,
                "timestamp": datetime.now(timezone.utc),
                "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            })
        if sync_to_sheets:
            if collection == "clients":
                background_tasks.add_task(_bg_push_client, doc)
            elif collection == "carriers":
                background_tasks.add_task(_bg_push_carrier, doc)
            elif collection == "leads":
                background_tasks.add_task(_bg_push_lead, doc)
        return ModelCls(**doc)

    @api_router.delete(f"/{prefix}/{{item_id}}")
    async def delete_item(item_id: str, background_tasks: BackgroundTasks):
        doc = await db[collection].find_one({"id": item_id}, {"_id": 0})
        if soft_delete:
            if doc:
                if sync_to_sheets and collection == "leads":
                    background_tasks.add_task(_bg_delete_lead, doc.get("name", ""))
                if sync_to_sheets and collection == "clients":
                    background_tasks.add_task(_bg_delete_client, doc.get("name", ""))
                deleted_at = now_iso()
                print(f"[delete_{collection}] id={item_id} setting deleted=True deleted_at={deleted_at}")
                result = await db[collection].update_one({"id": item_id}, {"$set": {"deleted": True, "deleted_at": deleted_at}})
                print(f"[delete_{collection}] modified_count={result.modified_count}")
            else:
                print(f"[delete_{collection}] id={item_id} not found")
            return {"ok": True}
        if doc:
            if sync_to_sheets and collection == "leads":
                background_tasks.add_task(_bg_delete_lead, doc.get("name", ""))
            if sync_to_sheets and collection == "clients":
                background_tasks.add_task(_bg_delete_client, doc.get("name", ""))
        await db[collection].delete_one({"id": item_id})
        return {"ok": True}


@api_router.get("/clients", response_model=List[Client])
async def list_clients(_: Optional[dict] = Depends(_get_user_from_token)):
    docs = await db.clients.find({"deleted": {"$ne": True}}, {"_id": 0}).sort("created_at", -1).to_list(50000)
    agg = await db.orders.aggregate([
        {"$match": {"deleted": {"$ne": True}}},
        {"$group": {"_id": "$client_name", "total_revenue": {"$sum": "$client_rate"}, "orders_count": {"$sum": 1}}}
    ]).to_list(10000)
    stats = {s["_id"]: s for s in agg if s["_id"]}
    for doc in docs:
        s = stats.get(doc.get("name", ""), {})
        doc["total_revenue"] = s.get("total_revenue", 0)
        doc["orders_count"] = s.get("orders_count", 0)
    return [Client(**d) for d in docs]

make_crud("clients", "clients", Client, ClientPayload, sync_to_sheets=True, soft_delete=True)
make_crud("carriers", "carriers", Carrier, CarrierPayload, sync_to_sheets=True, soft_delete=True)


@api_router.post("/clients/{client_id}/generate_acts")
async def generate_all_acts(client_id: str, current_user: dict = Depends(_require_user)):
    """Сгенерировать акты+счета для всех незакрытых заявок клиента."""
    if _docs_get_generator is None:
        raise HTTPException(500, "Docs generation недоступна")

    client_doc = await db.clients.find_one({"id": client_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not client_doc:
        try:
            from bson import ObjectId
            client_doc = await db.clients.find_one({"_id": ObjectId(client_id), "deleted": {"$ne": True}}, {"_id": 0})
        except Exception:
            pass
    if not client_doc:
        raise HTTPException(404, "Клиент не найден")

    # Получаем токен один раз здесь, на основном event loop — дальше он идёт
    # параметром в gen.generate/gen.create_combined_doc, которые выполняются
    # в asyncio.to_thread. Раньше credentials строились лениво внутри самого
    # потока через отдельный event loop, что падало с "Task ... attached to
    # a different loop" (Motor-клиент db привязан к основному loop).
    token_doc = await db.oauth_tokens.find_one({"_id": "google"}, {"_id": 0})

    client_name = client_doc.get("name", "")
    orders = await db.orders.find(
        {
            "$or": [{"client_id": client_id}, {"client_name": client_name}],
            "deleted": {"$ne": True},
            "status": {"$ne": "cancelled"},
        },
        {"_id": 0},
    ).to_list(10000)

    gen = _docs_get_generator()
    results = []
    errors = []

    for order in orders:
        order_number = order.get("order_number", "")
        order_id_val = order.get("id", "")

        carrier_doc = None
        if order.get("carrier_id"):
            carrier_doc = await db.carriers.find_one({"id": order["carrier_id"]}, {"_id": 0})
        if not carrier_doc and order.get("carrier_name"):
            carrier_doc = await db.carriers.find_one({"company_name": order["carrier_name"]}, {"_id": 0})

        order_urls: dict = {}
        had_error = False
        for kind in ("client", "act"):
            field = _docs_kind_to_field(kind)
            existing = order.get(field) or ""
            if existing:
                order_urls[kind] = existing
                continue
            try:
                url = await asyncio.to_thread(gen.generate, kind, order, client_doc, carrier_doc, token_doc)
                await db.orders.update_one({"id": order_id_val}, {"$set": {field: url}})
                order_urls[kind] = url
            except Exception as e:
                had_error = True
                errors.append({"order_number": order_number, "kind": kind, "error": str(e)})

        if not had_error:
            results.append({"order_number": order_number, **order_urls})

    combined_url = None
    act_items = [r for r in results if r.get("act")]
    if act_items:
        try:
            combined_url = await asyncio.to_thread(gen.create_combined_doc, act_items, client_name, token_doc)
            print(f"[generate_all_acts] combined doc: {combined_url}")
        except Exception as e:
            print(f"[generate_all_acts] combine error: {e}")

    print(f"[generate_all_acts] client={client_name} orders={len(orders)} ok={len(results)} err={len(errors)}")
    return {
        "url": combined_url,
        "created": len(results),
        "errors": len(errors),
        "client_name": client_name,
        "error_details": errors,
    }


@api_router.get("/leads/activity/stats")
async def leads_activity_stats():
    today = datetime.now(timezone.utc)
    days = [(today - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(6, -1, -1)]
    docs = await db.lead_activity.find({"date": {"$gte": days[0]}}, {"_id": 0}).to_list(10000)
    counts: dict = {d: 0 for d in days}
    for doc in docs:
        dt = doc.get("date", "")
        if dt in counts:
            counts[dt] += 1
    return [{"date": k, "count": v} for k, v in sorted(counts.items())]


@api_router.get("/leads/industries")
async def leads_industries():
    pipeline = [
        {"$match": {"deleted": {"$ne": True}, "industry": {"$nin": [None, ""]}}},
        {"$group": {"_id": "$industry", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    rows = await db.leads.aggregate(pipeline).to_list(200)
    return {"industries": [{"name": r["_id"], "count": r["count"]} for r in rows]}


@api_router.get("/leads", response_model=List[Lead])
async def list_leads_filtered(industry: Optional[str] = None, current_user: dict = Depends(_require_user)):
    filter_q: dict = {"deleted": {"$ne": True}}
    if current_user and current_user.get("role") == "manager":
        perms = current_user.get("permissions") or {}
        if not perms.get("can_view_all_leads"):
            # Own leads plus the unclaimed pool — not just own, so a manager
            # can still find and claim leads nobody has picked up yet.
            filter_q["$or"] = [{"assigned_to": current_user["id"]}, {"assigned_to": None}, {"assigned_to": ""}]
    if industry:
        filter_q["industry"] = industry
    docs = await db.leads.find(filter_q, {"_id": 0}).sort("created_at", -1).to_list(50000)
    return [Lead(**d) for d in docs]


def _leads_manager_filter(current_user: Optional[dict]) -> dict:
    base: dict = {"deleted": {"$ne": True}}
    if current_user and current_user.get("role") == "manager":
        perms = current_user.get("permissions") or {}
        if not perms.get("can_view_all_leads"):
            base["$or"] = [{"assigned_to": current_user["id"]}, {"assigned_to": None}, {"assigned_to": ""}]
    return base


@api_router.post("/leads/{lead_id}/claim")
async def claim_lead(lead_id: str, current_user: dict = Depends(_require_user)):
    lead = await db.leads.find_one({"id": lead_id, "deleted": {"$ne": True}})
    if not lead:
        raise HTTPException(404, "Лид не найден")
    if lead.get("assigned_to") and lead["assigned_to"] != current_user["id"]:
        raise HTTPException(400, "Лид уже закреплён за другим менеджером")
    await db.leads.update_one({"id": lead_id}, {"$set": {
        "assigned_to": current_user["id"], "assigned_at": datetime.now(timezone.utc).isoformat(),
    }})
    return {"ok": True}


@api_router.get("/leads/queue")
async def get_call_queue(industry: Optional[str] = None, limit: int = 50,
                         current_user: dict = Depends(_require_user)):
    now = datetime.now(timezone.utc).isoformat()
    today_end = datetime.now(timezone.utc).replace(hour=23, minute=59, second=59).isoformat()

    active_stages = [s["id"] for s in LEAD_STAGES if s["active"]]
    base = {**_leads_manager_filter(current_user), "stage": {"$in": active_stages}}
    if industry:
        base["industry"] = industry

    overdue = await db.leads.find(
        {**base, "next_call": {"$ne": None, "$lt": now}}, {"_id": 0}
    ).sort("next_call", 1).to_list(limit)

    today = await db.leads.find(
        {**base, "next_call": {"$gte": now, "$lte": today_end}}, {"_id": 0}
    ).sort("next_call", 1).to_list(limit)

    hot_stages = ["interested", "thinking", "kp_sent", "negotiation"]
    hot = await db.leads.find(
        {**base, "stage": {"$in": hot_stages}, "next_call": None}, {"_id": 0}
    ).sort("last_call_at", 1).to_list(limit)

    fresh = await db.leads.find(
        {**base, "stage": "new", "next_call": None}, {"_id": 0}
    ).sort("created_at", 1).to_list(limit)

    seen, queue = set(), []
    for bucket, tag in [(overdue, "overdue"), (today, "today"), (hot, "hot"), (fresh, "new")]:
        for l in bucket:
            if l["id"] in seen:
                continue
            seen.add(l["id"])
            l["queue_reason"] = tag
            queue.append(l)
            if len(queue) >= limit:
                break
        if len(queue) >= limit:
            break

    return {
        "queue": queue,
        "counts": {
            "overdue": len(overdue),
            "today": len(today),
            "hot": len(hot),
            "new": len(fresh),
        },
    }


@api_router.put("/leads/{item_id}", response_model=Lead)
async def update_lead(item_id: str, payload: LeadUpdate, background_tasks: BackgroundTasks,
                      current_user: Optional[dict] = Depends(_get_user_from_token)):
    old_doc = await db.leads.find_one({"id": item_id}, {"_id": 0})
    if not old_doc:
        raise HTTPException(404, "Not found")

    p_dict = payload.dict(exclude_none=True)
    p_dict.pop("call_notes", None)  # never overwrite notes via general PUT; use POST /call_notes
    p_dict["updated_at"] = now_iso()

    await db.leads.update_one({"id": item_id}, {"$set": p_dict})
    doc = await db.leads.find_one({"id": item_id}, {"_id": 0})

    background_tasks.add_task(_bg_push_lead, doc)
    return Lead(**doc)


class CallNotePayload(BaseModel):
    text: str
    author: Optional[str] = ""


@api_router.post("/leads/{item_id}/call_notes", response_model=Lead)
async def add_call_note(item_id: str, payload: CallNotePayload,
                        current_user: Optional[dict] = Depends(_get_user_from_token)):
    doc = await db.leads.find_one({"id": item_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Not found")
    note = {
        "text": payload.text,
        "date": datetime.now(timezone.utc).isoformat(),
        "author": payload.author or (current_user or {}).get("name", ""),
    }
    await db.leads.update_one(
        {"id": item_id},
        {
            "$push": {"call_notes": {"$each": [note], "$position": 0}},
            "$set": {"last_call": datetime.now(timezone.utc).strftime("%Y-%m-%d")},
        }
    )
    doc = await db.leads.find_one({"id": item_id}, {"_id": 0})
    return Lead(**doc)


@api_router.delete("/leads/{item_id}/call_notes/{note_index}", response_model=Lead)
async def delete_call_note(item_id: str, note_index: int,
                           current_user: Optional[dict] = Depends(_get_user_from_token)):
    doc = await db.leads.find_one({"id": item_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Not found")
    notes = doc.get("call_notes") or []
    if note_index < 0 or note_index >= len(notes):
        raise HTTPException(400, "Invalid note index")
    notes.pop(note_index)
    await db.leads.update_one({"id": item_id}, {"$set": {"call_notes": notes}})
    doc = await db.leads.find_one({"id": item_id}, {"_id": 0})
    return Lead(**doc)


async def _bg_call_calendar(title: str, date_str: str, description: str):
    if _gc_simple is None:
        return
    try:
        token_doc = await db.oauth_tokens.find_one({"_id": "google"}, {"_id": 0})
        if not token_doc:
            logging.getLogger(__name__).warning("_bg_call_calendar: no Google token in DB")
            return
        await asyncio.to_thread(_gc_simple, title, date_str, description, token_doc)
    except Exception as e:
        logging.getLogger(__name__).error(f"_bg_call_calendar failed: {e}")


@api_router.post("/leads/{item_id}/call")
async def log_call(item_id: str, payload: dict, background_tasks: BackgroundTasks,
                   current_user: Optional[dict] = Depends(_get_user_from_token)):
    lead = await db.leads.find_one({"id": item_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Лид не найден")

    outcome = payload.get("outcome")
    comment = (payload.get("comment") or "").strip()
    lost_reason = payload.get("lost_reason")
    manual_next = payload.get("next_call")
    now = datetime.now(timezone.utc)

    if outcome != "no_answer" and not comment:
        raise HTTPException(400, "Комментарий обязателен")
    if outcome == "lost" and not lost_reason:
        raise HTTPException(400, "Укажите причину отказа")

    attempts = int(lead.get("call_attempts") or 0)
    cadence_step = int(lead.get("cadence_step") or 0)
    upd: dict = {
        "last_call_at": now.isoformat(),
        "last_call": now.strftime("%Y-%m-%d"),
        "total_calls": int(lead.get("total_calls") or 0) + 1,
        "updated_at": now.isoformat(),
    }
    if not lead.get("first_call_at"):
        upd["first_call_at"] = now.isoformat()

    next_call = None

    if outcome == "no_answer":
        attempts += 1
        upd["call_attempts"] = attempts
        if attempts >= 5:
            upd["stage"] = "no_contact"
            upd["next_call"] = None
        else:
            d = (now + timedelta(days=1)).replace(hour=10, minute=0, second=0, microsecond=0)
            next_call = d.isoformat()
            upd["next_call"] = next_call
    else:
        upd["call_attempts"] = 0
        upd["stage"] = outcome
        upd["last_call_note"] = comment

        if outcome in ("won", "lost"):
            upd["next_call"] = None
            if outcome == "won":
                upd["won_at"] = now.isoformat()
            if outcome == "lost":
                upd["lost_reason"] = lost_reason
        else:
            if manual_next:
                next_call = manual_next
                upd["next_call"] = next_call
            else:
                new_step, auto_date = next_cadence_touch(cadence_step)
                if new_step:
                    upd["cadence_step"] = new_step
                    next_call = auto_date
                    upd["next_call"] = next_call

    if "stage" in upd and upd["stage"] != lead.get("stage"):
        upd["stage_changed_at"] = now.isoformat()

    await db.leads.update_one({"id": item_id}, {"$set": upd})

    log = {
        "id": str(uuid.uuid4()),
        "lead_id": item_id,
        "lead_name": lead.get("name", ""),
        "outcome": outcome,
        "comment": comment,
        "lost_reason": lost_reason,
        "next_call": next_call,
        "attempt_no": attempts,
        "cadence_step": upd.get("cadence_step", cadence_step),
        "duration_sec": payload.get("duration_sec"),
        "created_by": (current_user or {}).get("id", ""),
        "created_at": now.isoformat(),
    }
    await db.call_logs.insert_one(dict(log))
    # Keep legacy per-user activity feed (admin stats/leaderboards) alive.
    await db.lead_activity.insert_one({
        "id": str(uuid.uuid4()),
        "lead_id": item_id,
        "user_id": (current_user or {}).get("id", ""),
        "action": "call",
        "old_status": lead.get("stage"),
        "new_status": upd.get("stage", lead.get("stage")),
        "timestamp": now,
        "date": now.strftime("%Y-%m-%d"),
    })

    if next_call:
        task_obj = {
            "id": str(uuid.uuid4()),
            "title": f"Перезвонить: {lead.get('name', '')}",
            "task_type": "call",
            "due_date": next_call[:10],
            "due_time": next_call[11:16],
            "status": "pending",
            "description": f"{lead.get('phone', '')}. {comment}".strip(". "),
            "created_by": (current_user or {}).get("id", ""),
            "created_at": now_iso(),
            "google_task_id": None,
        }
        await db.tasks.insert_one(task_obj)
        background_tasks.add_task(_bg_gt_create, task_obj)
        background_tasks.add_task(_bg_call_calendar, task_obj["title"], task_obj["due_date"], task_obj["description"])

    doc = await db.leads.find_one({"id": item_id}, {"_id": 0})
    background_tasks.add_task(_bg_push_lead, doc)

    await manager.broadcast({"type": "lead_updated", "lead_id": item_id})
    return {
        "ok": True,
        "log": log,
        "stage": upd.get("stage", lead.get("stage")),
        "attempts": upd.get("call_attempts"),
        "next_call": next_call,
        "ask_create_client": outcome == "won" and not lead.get("client_id"),
        "lead": doc,
    }


@api_router.get("/leads/{item_id}/calls")
async def call_history(item_id: str):
    logs = await db.call_logs.find({"lead_id": item_id}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return {"calls": logs}


@api_router.post("/leads/{item_id}/convert")
async def convert_lead_to_client(item_id: str, current_user: Optional[dict] = Depends(_get_user_from_token)):
    lead = await db.leads.find_one({"id": item_id}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Лид не найден")
    if lead.get("client_id"):
        return {"client_id": lead["client_id"], "already": True}

    client = {
        "id": str(uuid.uuid4()),
        "name": lead.get("company") or lead.get("name", ""),
        "phone": lead.get("phone", ""),
        "email": lead.get("email", ""),
        "contact_person": lead.get("contact_person", ""),
        "city": lead.get("city", ""),
        "created_at": now_iso(),
    }
    await db.clients.insert_one(dict(client))

    now = datetime.now(timezone.utc)
    lead_upd: dict = {"client_id": client["id"]}
    # Converting to a client is a win even if it didn't come through the
    # call-outcome flow — without this, stats/session activity never see it.
    if lead.get("stage") != "won":
        lead_upd["stage"] = "won"
        lead_upd["won_at"] = now.isoformat()
        lead_upd["stage_changed_at"] = now.isoformat()
    if not lead.get("assigned_to") and current_user:
        lead_upd["assigned_to"] = current_user["id"]
        lead_upd["assigned_at"] = now.isoformat()
    await db.leads.update_one({"id": item_id}, {"$set": lead_upd})

    if current_user:
        await db.lead_activity.insert_one({
            "id": str(uuid.uuid4()), "lead_id": item_id, "user_id": current_user["id"],
            "action": "convert", "old_status": lead.get("stage"), "new_status": lead_upd.get("stage", lead.get("stage")),
            "timestamp": now, "date": now.strftime("%Y-%m-%d"),
        })

    return {"client_id": client["id"], "client": client}


@api_router.get("/leads/scripts")
async def get_scripts(current_user: Optional[dict] = Depends(_get_user_from_token)):
    doc = await db.settings.find_one({"key": "call_scripts"}, {"_id": 0})
    return {"scripts": (doc or {}).get("value") or DEFAULT_SCRIPTS}


@api_router.put("/leads/scripts")
async def save_scripts(payload: dict, current_user: Optional[dict] = Depends(_get_user_from_token)):
    await db.settings.update_one(
        {"key": "call_scripts"},
        {"$set": {"key": "call_scripts", "value": payload.get("scripts", {})}},
        upsert=True,
    )
    return {"ok": True}


def _safe_float(v) -> float:
    # Some legacy imported orders have non-numeric junk in rate fields
    # (blank strings, stray text) — this must never 500 a whole report.
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _analytics_period_start(period: str) -> datetime:
    now = datetime.now(timezone.utc)
    if period == "week":
        return now - timedelta(days=7)
    elif period == "month":
        return now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif period == "quarter":
        qm = (now.month - 1) // 3 * 3 + 1
        return now.replace(month=qm, day=1, hour=0, minute=0, second=0, microsecond=0)
    return datetime(2020, 1, 1, tzinfo=timezone.utc)


async def _compute_leads_analytics(logs: list, leads: list) -> dict:
    DAILY_GOAL = 45
    by_day: dict = {}
    for l in logs:
        d = l["created_at"][:10]
        by_day[d] = by_day.get(d, 0) + 1
    calls_by_day = [{"date": d, "calls": c, "goal": DAILY_GOAL} for d, c in sorted(by_day.items())]

    stage_counts: dict = {}
    for l in leads:
        s = l.get("stage", "new")
        stage_counts[s] = stage_counts.get(s, 0) + 1

    funnel = []
    prev = None
    for st in LEAD_STAGES:
        if st["id"] in ("no_contact", "lost"):
            continue
        cnt = stage_counts.get(st["id"], 0)
        reached_or_further = sum(
            stage_counts.get(s["id"], 0) for s in LEAD_STAGES
            if s["order"] >= st["order"] and s["id"] not in ("no_contact", "lost")
        )
        conv = round(reached_or_further / prev * 100, 1) if prev else 100.0
        funnel.append({
            "stage": st["id"], "label": st["label"], "color": st["color"],
            "count": cnt, "cumulative": reached_or_further, "conversion": conv,
        })
        prev = reached_or_further or prev

    reasons: dict = {}
    for l in leads:
        if l.get("stage") == "lost" and l.get("lost_reason"):
            reasons[l["lost_reason"]] = reasons.get(l["lost_reason"], 0) + 1
    lost_reasons = sorted(
        [{"reason": k, "count": v} for k, v in reasons.items()],
        key=lambda x: -x["count"]
    )

    ind: dict = {}
    for l in leads:
        i = l.get("industry") or "Не указана"
        ind.setdefault(i, {"total": 0, "won": 0, "lost": 0, "called": 0})
        ind[i]["total"] += 1
        if l.get("stage") == "won":
            ind[i]["won"] += 1
        if l.get("stage") == "lost":
            ind[i]["lost"] += 1
        if (l.get("total_calls") or 0) > 0:
            ind[i]["called"] += 1
    by_industry = sorted([
        {
            "industry": k,
            "total": v["total"],
            "called": v["called"],
            "won": v["won"],
            "lost": v["lost"],
            "conversion": round(v["won"] / v["called"] * 100, 1) if v["called"] else 0.0,
        }
        for k, v in ind.items()
    ], key=lambda x: -x["conversion"])

    converted = [l for l in leads if l.get("client_id")]
    cold_client_ids = [l["client_id"] for l in converted]
    cold_orders = await db.orders.find(
        {"client_id": {"$in": cold_client_ids}, "deleted": {"$ne": True}}, {"_id": 0}
    ).to_list(100000) if cold_client_ids else []
    cold_revenue = sum(float(o.get("client_rate") or 0) for o in cold_orders)
    cold_margin = sum(
        float(o.get("client_rate") or 0) - float(o.get("carrier_rate") or 0)
        for o in cold_orders
    )

    return {
        "calls_by_day": calls_by_day,
        "daily_goal": DAILY_GOAL,
        "total_calls": len(logs),
        "funnel": funnel,
        "lost_reasons": lost_reasons,
        "by_industry": by_industry,
        "cold_leads": {
            "converted_count": len(converted),
            "orders_count": len(cold_orders),
            "revenue": cold_revenue,
            "margin": cold_margin,
        },
    }


@api_router.get("/leads/analytics")
async def leads_analytics(period: str = "month", current_user: Optional[dict] = Depends(_get_user_from_token)):
    start_str = _analytics_period_start(period).isoformat()
    logs = await db.call_logs.find({"created_at": {"$gte": start_str}}, {"_id": 0}).to_list(100000)
    leads = await db.leads.find({"deleted": {"$ne": True}}, {"_id": 0}).to_list(100000)
    return await _compute_leads_analytics(logs, leads)


@api_router.get("/leads/calls_by_day")
async def calls_by_day(date: str, current_user: Optional[dict] = Depends(_get_user_from_token)):
    start = f"{date}T00:00:00"
    end = f"{date}T23:59:59"
    logs = await db.call_logs.find(
        {"created_at": {"$gte": start, "$lte": end}}, {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    return {"calls": logs, "date": date}


# ====== Task models ======
class Task(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: Optional[str] = ""
    task_type: str = "other"  # call / meeting / docs / order / other
    due_date: Optional[str] = ""
    due_time: Optional[str] = ""
    status: str = "pending"  # pending / done
    google_task_id: Optional[str] = None
    created_by: Optional[str] = ""
    assigned_user_id: Optional[str] = None
    created_at: str = Field(default_factory=now_iso)


class TaskCreate(BaseModel):
    title: str
    description: Optional[str] = ""
    task_type: str = "other"
    due_date: Optional[str] = ""
    due_time: Optional[str] = ""
    status: str = "pending"
    assigned_user_id: Optional[str] = None


class TaskUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    task_type: Optional[str] = None
    due_date: Optional[str] = None
    due_time: Optional[str] = None
    status: Optional[str] = None
    google_task_id: Optional[str] = None
    assigned_user_id: Optional[str] = None


# ====== Google Tasks background helpers ======
async def _bg_gt_create(task_obj: dict):
    if _gt_create is None:
        return
    try:
        token_doc = await db.oauth_tokens.find_one({"_id": "google"}, {"_id": 0})
        if not token_doc:
            logging.getLogger(__name__).warning("_bg_gt_create: no Google token in DB")
            return
        gid = await asyncio.to_thread(_gt_create, task_obj, token_doc)
        if gid:
            await db.tasks.update_one({"id": task_obj["id"]}, {"$set": {"google_task_id": gid}})
            logging.getLogger(__name__).info(f"_bg_gt_create: synced task {task_obj.get('id')} → gid={gid}")
        else:
            logging.getLogger(__name__).warning(f"_bg_gt_create: create_google_task returned None for task {task_obj.get('id')}")
    except Exception as e:
        logging.getLogger(__name__).error(f"_bg_gt_create failed: {e}", exc_info=True)


async def _bg_gt_update(task_obj: dict):
    if _gt_update is None or not task_obj.get("google_task_id"):
        return
    try:
        token_doc = await db.oauth_tokens.find_one({"_id": "google"}, {"_id": 0})
        if not token_doc:
            return
        await asyncio.to_thread(_gt_update, task_obj["google_task_id"], task_obj, token_doc)
    except Exception as e:
        logging.getLogger(__name__).error(f"_bg_gt_update failed: {e}")


async def _bg_gt_delete(google_task_id: str):
    if _gt_delete is None or not google_task_id:
        return
    try:
        token_doc = await db.oauth_tokens.find_one({"_id": "google"}, {"_id": 0})
        if not token_doc:
            return
        await asyncio.to_thread(_gt_delete, google_task_id, token_doc)
    except Exception as e:
        logging.getLogger(__name__).error(f"_bg_gt_delete failed: {e}")


# ====== Google Calendar background helpers ======
async def _bg_cal_create(order_obj: dict):
    if _gc_create is None:
        return
    try:
        token_doc = await db.oauth_tokens.find_one({"_id": "google"}, {"_id": 0})
        if not token_doc:
            return
        result = await asyncio.to_thread(_gc_create, order_obj, token_doc)
        if result:
            await db.orders.update_one(
                {"id": order_obj["id"]},
                {"$set": {"calendar_event_id": result["event_id"], "calendar_event_url": result["html_link"]}},
            )
    except Exception as e:
        logging.getLogger(__name__).error(f"_bg_cal_create failed: {e}")


async def _bg_cal_update(order_doc: dict):
    if _gc_update is None or not order_doc.get("calendar_event_id"):
        return
    try:
        token_doc = await db.oauth_tokens.find_one({"_id": "google"}, {"_id": 0})
        if not token_doc:
            return
        await asyncio.to_thread(_gc_update, order_doc["calendar_event_id"], order_doc, token_doc)
    except Exception as e:
        logging.getLogger(__name__).error(f"_bg_cal_update failed: {e}")


async def _bg_cal_delete(event_id: str):
    if _gc_delete is None or not event_id:
        return
    try:
        token_doc = await db.oauth_tokens.find_one({"_id": "google"}, {"_id": 0})
        if not token_doc:
            return
        await asyncio.to_thread(_gc_delete, event_id, token_doc)
    except Exception as e:
        logging.getLogger(__name__).error(f"_bg_cal_delete failed: {e}")


def add_business_days(start_date: datetime, days: int) -> datetime:
    current = start_date
    added = 0
    while added < days:
        current += timedelta(days=1)
        if current.weekday() < 5:
            added += 1
    return current


async def _bg_carrier_payment_calendar(deadline_str: str, reminder_str: str, carrier_name: str,
                                        order_number: str, carrier_rate: float, reminder_days_before: int,
                                        deadline_display: str):
    try:
        from google_calendar import create_simple_calendar_event as _create_simple
    except Exception as _ie:
        logging.getLogger(__name__).warning(f"create_simple_calendar_event unavailable: {_ie}")
        return
    try:
        token_doc = await db.oauth_tokens.find_one({"_id": "google"}, {"_id": 0})
        if not token_doc:
            logging.getLogger(__name__).warning("_bg_carrier_payment_calendar: no Google token in DB")
            return
        r1 = await asyncio.to_thread(
            _create_simple,
            f"💰 Оплатить перевозчика: {carrier_name}",
            deadline_str,
            f"Заявка {order_number}, сумма {carrier_rate} BYN",
            token_doc,
        )
        logging.getLogger(__name__).info(f"carrier payment calendar event (deadline): {r1}")
        r2 = await asyncio.to_thread(
            _create_simple,
            f"⚠️ Через {reminder_days_before} дня оплатить: {carrier_name}",
            reminder_str,
            f"Срок оплаты: {deadline_display}",
            token_doc,
        )
        logging.getLogger(__name__).info(f"carrier payment calendar event (reminder): {r2}")
    except Exception as e:
        logging.getLogger(__name__).error(f"_bg_carrier_payment_calendar failed: {e}", exc_info=True)


def _close_carrier_google_tasks_sync(carrier_name: str, token_doc: dict) -> list:
    """Sync: mark payment tasks done in Google Tasks and ✅ calendar events."""
    from google_tasks import _build_service as _gt_build_svc, get_or_create_tasklist, PAYMENT_LIST_NAME
    from google_calendar import _build_service as _cal_build_svc
    from datetime import datetime, timedelta, timezone as _tz
    results = []
    # 1. Close Google Tasks in "Оплаты перевозчиков"
    try:
        gt_service = _gt_build_svc(token_doc)
        list_id = get_or_create_tasklist(gt_service, PAYMENT_LIST_NAME)
        tasks_resp = gt_service.tasks().list(tasklist=list_id, showCompleted=False).execute()
        for task in tasks_resp.get("items", []):
            title = task.get("title", "")
            if carrier_name.lower() in title.lower():
                updated = dict(task)
                updated["status"] = "completed"
                gt_service.tasks().update(tasklist=list_id, task=task["id"], body=updated).execute()
                results.append(f"Google Task completed: {title}")
    except Exception as e:
        results.append(f"Google Tasks error: {e}")
    # 2. Mark Google Calendar events with ✅
    try:
        cal_service = _cal_build_svc(token_doc)
        now = datetime.now(_tz.utc)
        events_resp = cal_service.events().list(
            calendarId="primary",
            q=carrier_name,
            timeMin=(now - timedelta(days=90)).isoformat(),
            timeMax=(now + timedelta(days=90)).isoformat(),
        ).execute()
        for event in events_resp.get("items", []):
            summary = event.get("summary", "")
            if (("оплатить" in summary.lower() or "оплата" in summary.lower())
                    and carrier_name.lower() in summary.lower()
                    and not summary.startswith("✅")):
                updated_event = dict(event)
                updated_event["summary"] = "✅ " + summary
                cal_service.events().update(
                    calendarId="primary", eventId=event["id"], body=updated_event
                ).execute()
                results.append(f"Calendar event updated: {summary}")
    except Exception as e:
        results.append(f"Calendar error: {e}")
    return results


async def _bg_close_carrier_payment_on_paid(carrier_name: str, order_id: str):
    if not carrier_name:
        return
    try:
        token_doc = await db.oauth_tokens.find_one({"_id": "google"}, {"_id": 0})
        if not token_doc:
            logging.getLogger(__name__).warning("_bg_close_carrier_payment_on_paid: no Google token")
            return
        results = await asyncio.to_thread(_close_carrier_google_tasks_sync, carrier_name, token_doc)
        for r in results:
            print(f"[carrier_paid] {r}")
    except Exception as e:
        logging.getLogger(__name__).error(f"_bg_close_carrier_payment_on_paid: {e}", exc_info=True)


# ====== Task CRUD endpoints ======
@api_router.get("/tasks", response_model=List[Task])
async def list_tasks(status: Optional[str] = "pending", current_user: dict = Depends(_require_user)):
    # Default to pending-only so the main task view (and the dashboard
    # overdue widget / counts) never shows completed rows. Pass status=all
    # for the dedicated "Завершённые" tab, or status=done explicitly.
    filter_q: dict = {}
    if status and status != "all":
        filter_q["status"] = status
    if current_user.get("role") == "manager":
        filter_q["$or"] = [{"created_by": current_user["id"]}, {"assigned_user_id": current_user["id"]}]
    docs = await db.tasks.find(filter_q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return [Task(**d) for d in docs]


@api_router.post("/tasks/{task_id}/assign")
async def assign_task(task_id: str, payload: dict, current_user: dict = Depends(require_director)):
    await db.tasks.update_one({"id": task_id}, {"$set": {"assigned_user_id": payload.get("user_id")}})
    return {"ok": True}


@api_router.get("/tasks/{task_id}", response_model=Task)
async def get_task(task_id: str):
    doc = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Task not found")
    return Task(**doc)


@api_router.post("/tasks", response_model=Task)
async def create_task(payload: TaskCreate, background_tasks: BackgroundTasks,
                      current_user: Optional[dict] = Depends(_get_user_from_token)):
    task_data = payload.dict()
    if current_user:
        task_data["created_by"] = current_user["id"]
    obj = Task(**task_data)
    await db.tasks.insert_one(obj.dict())
    background_tasks.add_task(_bg_gt_create, obj.dict())
    return obj


@api_router.put("/tasks/{task_id}", response_model=Task)
async def update_task(task_id: str, payload: TaskUpdate, background_tasks: BackgroundTasks):
    await db.tasks.update_one({"id": task_id}, {"$set": payload.dict(exclude_none=True)})
    doc = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Not found")
    background_tasks.add_task(_bg_gt_update, doc)
    return Task(**doc)


@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, background_tasks: BackgroundTasks):
    doc = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if doc and doc.get("google_task_id"):
        background_tasks.add_task(_bg_gt_delete, doc["google_task_id"])
    await db.tasks.delete_one({"id": task_id})
    return {"ok": True}


# ====== А2 Инфо СРМ — подписка на Telegram-бота ======
@api_router.post("/bot/subscribe")
async def subscribe_bot(payload: dict, current_user: dict = Depends(_require_user)):
    """Пользователь вводит свой chat_id в CRM, чтобы получать уведомления
    (напоминания об оплате — менеджерам по их заявкам, отчёты — директорам)."""
    chat_id = str(payload.get("chat_id") or "").strip()
    if not chat_id:
        raise HTTPException(400, "chat_id обязателен")
    await db.bot_subscribers.update_one(
        {"user_id": current_user["id"]},
        {
            "$set": {
                "user_id": current_user["id"],
                "telegram_chat_id": chat_id,
                "role": current_user.get("role"),
            },
            "$setOnInsert": {"id": str(uuid.uuid4())},
        },
        upsert=True,
    )
    return {"ok": True}


@api_router.get("/bot/subscription")
async def get_bot_subscription(current_user: dict = Depends(_require_user)):
    sub = await db.bot_subscribers.find_one({"user_id": current_user["id"]}, {"_id": 0})
    return {"subscribed": bool(sub), "chat_id": sub.get("telegram_chat_id") if sub else None}


# ====== Отчёты — история (скрытая страница, только директор) ======
@api_router.get("/reports")
async def list_reports(period: Optional[str] = None, current_user: dict = Depends(require_director)):
    q = {"period": period} if period else {}
    reports = await db.reports.find(q, {"_id": 0}).sort("generated_at", -1).to_list(100)
    return {"reports": reports}


@api_router.post("/reports/run")
async def run_report_now(period: str = "daily", current_user: dict = Depends(require_director)):
    """Сформировать отчёт вручную и разослать в Telegram. Возвращает статус
    доставки по каждому chat_id (дошло / ошибка), чтобы было видно, почему
    отчёт не приходит — нет токена, chat_id не нажал /start и т.п."""
    if period not in ("daily", "weekly", "monthly"):
        raise HTTPException(400, "period must be daily|weekly|monthly")
    result = await send_scheduled_report(period)
    return {
        "ok": True,
        "report": result.get("report"),
        "sent": result.get("sent", 0),
        "targets": result.get("targets", 0),
        "delivery": result.get("delivery", []),
        "token_configured": bool(os.environ.get("A2_INFO_BOT_TOKEN")),
    }


@api_router.post("/tasks/sync_payment_reminders")
async def sync_payment_reminders_now(current_user: dict = Depends(require_director)):
    """Прогнать синхронизацию задач-напоминаний об оплате прямо сейчас
    (обычно раз в час). Возвращает, сколько создано/обновлено и ушло ли
    сводное сообщение в Telegram."""
    res = await _run_payment_reminder_sync()
    delivery = await _notify_new_payment_reminders(res["created"]) if res["created"] else []
    return {
        "ok": True,
        "created": len(res["created"]),
        "updated": res["updated"],
        "sent": sum(1 for d in delivery if d.get("ok")),
        "token_configured": bool(os.environ.get("A2_INFO_BOT_TOKEN")),
    }


_ORDERS_LIGHT_PROJECTION = {
    "_id": 0, "id": 1, "order_number": 1, "client_id": 1, "client_name": 1,
    "carrier_id": 1, "carrier_name": 1, "route_from": 1, "route_to": 1, "cargo": 1,
    "status": 1, "client_paid": 1, "carrier_paid": 1, "client_paid_date": 1,
    "carrier_paid_date": 1, "client_rate": 1, "carrier_rate": 1, "load_date": 1,
    "unload_date": 1, "weight_tons": 1, "client_pp_number": 1, "client_payments": 1,
    "carrier_pp_number": 1, "carrier_pp_date": 1, "carrier_payments": 1,
    "client_cash": 1, "carrier_cash": 1,
    "docs_to_client_sent": 1, "docs_from_client_received": 1, "docs_to_carrier_sent": 1,
    "docs_from_carrier_received": 1,
}


@api_router.get("/orders", response_model=List[Order])
async def list_orders(current_user: dict = Depends(_require_user), limit: int = 2000,
                       client_id: Optional[str] = None, carrier_id: Optional[str] = None,
                       light: bool = False):
    filter_q: dict = {"deleted": {"$ne": True}}
    if current_user.get("role") == "manager":
        perms = current_user.get("permissions") or {}
        if not perms.get("can_view_all_orders"):
            filter_q["$or"] = [{"assigned_to": current_user["id"]}, {"assigned_to": None}, {"assigned_to": ""}]
    # Pushed down to Mongo instead of fetched-then-filtered client-side —
    # CarrierDetail/ClientDetail already filter the result by these ids on
    # their end (kept as a harmless no-op there), but previously that meant
    # every visit to either page pulled the *entire* orders collection just
    # to display one counterparty's handful of rows.
    if client_id:
        filter_q["client_id"] = client_id
    if carrier_id:
        filter_q["carrier_id"] = carrier_id
    import re as _re
    # light=true is opt-in, used only by the list-overview screens (Orders
    # list, dashboard's preloaded orders) — the full unprojected shape stays
    # the default for every other caller (order detail, client/carrier
    # detail, the mobile app) so nothing else silently loses fields.
    projection = _ORDERS_LIGHT_PROJECTION if light else {"_id": 0}
    _t0 = time.time()
    docs = await db.orders.find(filter_q, projection).to_list(min(max(limit, 1), 2000))
    _t1 = time.time()

    seen: set = set()
    unique_docs = []
    for d in docs:
        key = d.get("order_number") or str(d.get("_id", ""))
        if key not in seen:
            seen.add(key)
            unique_docs.append(d)

    _num_pat = _re.compile(r"\d+")
    def _order_sort_key(d):
        m = _num_pat.search(d.get("order_number") or "")
        return int(m.group()) if m else 0

    docs = sorted(unique_docs, key=_order_sort_key, reverse=True)

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    result = []
    for d in docs:
        unload = (d.get("unload_date") or "")[:10]
        status = d.get("status", "")
        d["is_overdue"] = bool(unload and unload < today and status not in ("done", "delivered", "cancelled"))
        result.append(Order(**d))
    _t2 = time.time()
    if _t2 - _t0 > 0.5:
        logger.warning(f"[PROFILE] /orders count={len(docs)} db_fetch={_t1-_t0:.2f}s process+pydantic={_t2-_t1:.2f}s")
    return result


@api_router.get("/orders/next_number")
async def next_order_number():
    import re as _re
    year = datetime.now(timezone.utc).year
    all_docs = await db.orders.find({"deleted": {"$ne": True}}, {"_id": 0, "order_number": 1}).to_list(5000)
    pattern = _re.compile(r"[ЗЗз3]\s*[-–—]\s*(\d+)\s*/\s*(\d{4})", _re.IGNORECASE)
    max_num = 0
    for d in all_docs:
        m = pattern.search(d.get("order_number", "") or "")
        if not m:
            continue
        n, y = int(m.group(1)), int(m.group(2))
        if y == year and n > max_num:
            max_num = n
    return {"order_number": f"З-{max_num + 1:03d}/{year}"}


@api_router.post("/orders", response_model=Order)
async def create_order(payload: OrderPayload, background_tasks: BackgroundTasks,
                       current_user: Optional[dict] = Depends(_get_user_from_token)):
    order_data = payload.dict()
    order_data["created_by"] = current_user["id"] if current_user else "admin"
    if not order_data.get("order_number"):
        import re as _re
        year = datetime.now(timezone.utc).year
        _pat = _re.compile(r"[ЗЗз3]\s*[-–—]\s*(\d+)\s*/\s*(\d{4})", _re.IGNORECASE)
        candidate = None
        # Two orders created close together used to be able to compute the
        # same "next number" and both get inserted with it — the sheets sync
        # would then later silently overwrite one of them entirely (matched
        # by that shared number, not by id). Re-checking right before insert
        # and retrying narrows that race; it isn't a hard DB-level guarantee
        # (no unique index — some already-duplicated numbers exist), but
        # combined with the sheets-import fix a collision can no longer
        # destroy data even if one still slips through.
        for _attempt in range(5):
            all_nums = await db.orders.find(
                {"deleted": {"$ne": True}}, {"_id": 0, "order_number": 1}
            ).to_list(10000)
            max_num = 0
            for _d in all_nums:
                _m = _pat.search(_d.get("order_number", "") or "")
                if _m and int(_m.group(2)) == year and int(_m.group(1)) > max_num:
                    max_num = int(_m.group(1))
            candidate = f"З-{max_num + 1:03d}/{year}"
            if not await db.orders.find_one({"order_number": candidate}, {"_id": 1}):
                break
        order_data["order_number"] = candidate
    obj = Order(**order_data)
    await db.orders.insert_one(obj.dict())
    background_tasks.add_task(_bg_push_order, obj.dict())
    background_tasks.add_task(_bg_cal_create, obj.dict())
    return obj


@api_router.get("/orders/{order_id}", response_model=Order)
async def get_order(order_id: str):
    doc = await db.orders.find_one({"id": order_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Order not found")
    return Order(**doc)


@api_router.put("/orders/{order_id}", response_model=Order)
async def update_order(order_id: str, payload: OrderUpdate, background_tasks: BackgroundTasks,
                       current_user: Optional[dict] = Depends(_get_user_from_token)):
    old_doc = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not old_doc:
        raise HTTPException(404, "Order not found")
    update_data = payload.dict(exclude_unset=True)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if update_data.get("client_paid") is True and "client_paid_date" not in update_data:
        if not old_doc.get("client_paid"):
            update_data["client_paid_date"] = today
    if update_data.get("carrier_paid") is True and "carrier_paid_date" not in update_data:
        if not old_doc.get("carrier_paid"):
            update_data["carrier_paid_date"] = today
    if update_data.get("carrier_paid") is True and not old_doc.get("carrier_paid"):
        _order_number = old_doc.get("order_number", "")
        _carrier_name = old_doc.get("carrier_name", "")
        _or_clauses: list = [{"order_id": order_id}]
        if _carrier_name:
            _or_clauses.append({"title": {"$regex": _carrier_name, "$options": "i"}})
        await db.tasks.update_many(
            {"status": {"$ne": "done"}, "$or": _or_clauses},
            {"$set": {"status": "done", "completed_at": now_iso()}}
        )
        logging.getLogger(__name__).info(f"Closed payment tasks for order {_order_number}")
        background_tasks.add_task(_bg_close_carrier_payment_on_paid, _carrier_name, order_id)
    carrier_already_paid = update_data.get("carrier_paid") or old_doc.get("carrier_paid")
    if update_data.get("docs_from_carrier_received") and not old_doc.get("docs_from_carrier_received") and not carrier_already_paid:
        days = update_data.get("carrier_payment_days") or old_doc.get("carrier_payment_days") or 20
        received_date = datetime.now(timezone.utc)
        payment_deadline = add_business_days(received_date, days)
        reminder_days_before = 5 if days > 15 else 3
        reminder_date = add_business_days(received_date, days - reminder_days_before)
        deadline_str = payment_deadline.strftime("%Y-%m-%d")
        reminder_str = reminder_date.strftime("%Y-%m-%d")
        deadline_display = payment_deadline.strftime("%d.%m.%Y")
        update_data["carrier_payment_deadline"] = deadline_str
        update_data["carrier_payment_reminder_date"] = reminder_str
        carrier_name = old_doc.get("carrier_name") or ""
        order_number = old_doc.get("order_number") or ""
        carrier_rate = old_doc.get("carrier_rate") or 0
        _created_by = (current_user or {}).get("id", "")
        for task_fields in [
            {
                "id": str(uuid.uuid4()),
                "title": f"Оплатить перевозчика: {carrier_name} — заявка {order_number}",
                "task_type": "payment",
                "due_date": deadline_str,
                "due_time": "10:00",
                "status": "pending",
                "description": f"Заявка {order_number}, сумма: {carrier_rate} BYN. Срок: {deadline_display}",
                "order_id": order_id,
                "created_by": _created_by,
                "created_at": now_iso(),
                "google_task_id": None,
            },
            {
                "id": str(uuid.uuid4()),
                "title": f"⚠️ Через {reminder_days_before} дня оплатить перевозчика: {carrier_name}",
                "task_type": "reminder",
                "due_date": reminder_str,
                "due_time": "09:00",
                "status": "pending",
                "description": f"Срок оплаты: {deadline_display}. Заявка {order_number}",
                "order_id": order_id,
                "created_by": _created_by,
                "created_at": now_iso(),
                "google_task_id": None,
            },
        ]:
            obj = Task(**task_fields)
            await db.tasks.insert_one(obj.dict())
            background_tasks.add_task(_bg_gt_create, obj.dict())
        background_tasks.add_task(
            _bg_carrier_payment_calendar,
            deadline_str, reminder_str, carrier_name, order_number,
            carrier_rate, reminder_days_before, deadline_display,
        )
    if update_data:
        await db.orders.update_one({"id": order_id}, {"$set": update_data})
        # Log changes
        actor = (current_user or {}).get("name") or "admin"
        log_entries = []
        skip_log = {"calendar_event_id", "calendar_event_url", "doc_url_client", "doc_url_carrier", "doc_url_act",
                    "client_paid_date", "carrier_paid_date",
                    "client_pp_number", "client_pp_date", "carrier_pp_number", "carrier_pp_date"}
        for field, new_val in update_data.items():
            if field in skip_log:
                continue
            old_val = old_doc.get(field)
            if old_val != new_val:
                log_entries.append({
                    "id": str(uuid.uuid4()),
                    "order_id": order_id,
                    "order_number": old_doc.get("order_number", ""),
                    "field": field,
                    "old_value": str(old_val) if old_val is not None else "",
                    "new_value": str(new_val) if new_val is not None else "",
                    "timestamp": now_iso(),
                    "user": actor,
                })
        if log_entries:
            await db.order_logs.insert_many(log_entries)
    doc = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Order not found")
    # Retroactive KUDiR sync: mark_payment() syncs these at the moment of
    # payment, but the "дозаполнить ПП" screen for already-paid orders saves
    # through this generic endpoint instead — so a PP number/date filled in
    # later (or a payment unmarked) here needs to trigger the same
    # create/remove logic. Safe to call on every save; it just recomputes.
    await _sync_kudir_row(order_id, "client")
    await _sync_kudir_row(order_id, "carrier")
    background_tasks.add_task(_bg_push_order, doc)
    background_tasks.add_task(_bg_cal_update, doc)
    await manager.broadcast({"type": "order_updated", "order_id": order_id, "patch": update_data})
    return Order(**doc)


async def _kudir_payment_group(order: dict, pp_field: str, date_field: str) -> list:
    """Заявки, оплаченные ОДНИМ и тем же платёжным поручением (тот же номер
    ПП + та же дата) — банк нередко закрывает несколько заявок одним
    переводом, и такая группа должна лечь в КУДиР одной строкой, а не по
    строке на каждую заявку."""
    import re as _re
    date_str = (order.get(date_field) or '')[:10]
    if not order.get(pp_field) or not date_str:
        return [order]
    group = await db.orders.find({
        pp_field: order[pp_field],
        date_field: {'$regex': f'^{_re.escape(date_str)}'},
        'deleted': {'$ne': True}, 'status': {'$ne': 'cancelled'},
    }, {'_id': 0}).to_list(500)
    return group or [order]


async def _create_kudir_income_row(order: dict):
    """Строка дохода в КУДиР — создаётся/пересчитывается при отметке оплаты
    клиентом с номером ПП. Заявки с одинаковым client_pp_number + датой
    схлопываются в одну строку с суммарной маржой (как в акте сверки)."""
    import re as _re
    if not order.get('client_pp_number') or not order.get('client_paid_date'):
        return
    date_str = order['client_paid_date'][:10]
    group = await _kudir_payment_group(order, 'client_pp_number', 'client_paid_date')
    group_ids = [o['id'] for o in group]
    order_numbers = sorted(o.get('order_number') or '' for o in group)
    orders_label = ', '.join(n for n in order_numbers if n)
    # Each order's margin is floored at zero BEFORE summing — otherwise a
    # loss on one order in the group would silently offset a gain on
    # another (they only share this row because they share a PP+date, not
    # because they're related), understating the group's real income.
    total_margin = sum(max(0.0, float(o.get('client_rate') or 0) - float(o.get('carrier_rate') or 0)) for o in group)
    total_paid = sum(float(o.get('client_rate') or 0) for o in group)
    client_names = sorted({o.get('client_name') or '' for o in group} - {''})
    client_label = client_names[0] if len(client_names) <= 1 else ', '.join(client_names)
    carrier_notes = sorted({
        f"{float(o.get('carrier_rate') or 0):.2f} — перевозчику {o.get('carrier_name')}"
        for o in group if o.get('carrier_name')
    })

    # Client-side act — this business's own act numbering is fixed 1:1 to
    # the order number (акт № совпадает с номером заявки), unlike the
    # carrier's act which is the carrier's own separate document.
    def _client_act_part(o):
        d = (o.get('unload_date') or o.get('load_date') or '')[:10]
        num = o.get('order_number') or ''
        if _re.match(r'^\d{4}-\d{2}-\d{2}$', d):
            return f"№ {num} от {fmt_date_ru(d)}"
        return f"№ {num}"
    client_act_parts = [p for p in (_client_act_part(o) for o in group) if p]
    client_act_ref = f"; Акт {', '.join(client_act_parts)}" if client_act_parts else ''

    entry = {
        'id': str(uuid.uuid4()),
        'order_id': order['id'], 'order_ids': group_ids,
        'order_number': orders_label,
        'row_type': 'income',
        'entry_date': date_str,
        'document_ref': f"Плат. поручение № {order['client_pp_number']} от {fmt_date_ru(date_str)}{client_act_ref}",
        'content': f"Оплата от {client_label} на сумму {total_paid:.2f} BYN по заявк{'е' if len(group) == 1 else 'ам'} № {orders_label}",
        'income_amount': total_margin,
        'note': '; '.join(carrier_notes),
        'created_at': now_iso(),
    }
    # A row someone hand-corrected in the book (PATCH /kudir/entries) must
    # survive the next resync instead of being silently overwritten — skip
    # both the delete and the reinsert when one already exists for this group.
    if await db.kudir_entries.find_one({'row_type': 'income', 'order_id': {'$in': group_ids}, 'manually_edited': True}):
        return
    await db.kudir_entries.delete_many({'row_type': 'income', 'order_id': {'$in': group_ids}})
    await db.kudir_entries.insert_one(entry)


async def _create_kudir_transit_row(order: dict):
    """Строка транзита в КУДиР — создаётся/пересчитывается при отметке
    оплаты перевозчику с номером ПП. Заявки с одинаковым
    carrier_pp_number + датой схлопываются в одну строку."""
    if not order.get('carrier_pp_number') or not order.get('carrier_paid_date'):
        return
    date_str = order['carrier_paid_date'][:10]
    group = await _kudir_payment_group(order, 'carrier_pp_number', 'carrier_paid_date')
    group_ids = [o['id'] for o in group]
    order_numbers = sorted(o.get('order_number') or '' for o in group)
    orders_label = ', '.join(n for n in order_numbers if n)
    carrier_names = sorted({o.get('carrier_name') or '' for o in group} - {''})
    carrier_label = carrier_names[0] if len(carrier_names) <= 1 else ', '.join(carrier_names)
    total_transit = sum(float(o.get('carrier_rate') or 0) for o in group)

    def _act_part(o):
        if o.get('carrier_act_number') and o.get('carrier_act_date'):
            return f"№ {o['carrier_act_number']} от {fmt_date_ru(o['carrier_act_date'][:10])}"
        if o.get('carrier_act_number'):
            return f"№ {o['carrier_act_number']}"
        return None
    act_parts = [p for p in (_act_part(o) for o in group) if p]
    act_ref = f"; Акт {', '.join(act_parts)}" if act_parts else ''

    entry = {
        'id': str(uuid.uuid4()),
        'order_id': order['id'], 'order_ids': group_ids,
        'order_number': orders_label,
        'row_type': 'transit',
        'entry_date': date_str,
        'document_ref': f"Плат. поручение № {order['carrier_pp_number']} от {fmt_date_ru(date_str)}{act_ref}",
        'content': f"Перечислено перевозчику {carrier_label} на сумму {total_transit:.2f} BYN по заявк{'е' if len(group) == 1 else 'ам'} № {orders_label}",
        'income_amount': None,
        'note': '',
        'created_at': now_iso(),
    }
    # See the matching guard in _create_kudir_income_row — a hand-corrected
    # row must not get clobbered by the next auto-sync.
    if await db.kudir_entries.find_one({'row_type': 'transit', 'order_id': {'$in': group_ids}, 'manually_edited': True}):
        return
    await db.kudir_entries.delete_many({'row_type': 'transit', 'order_id': {'$in': group_ids}})
    await db.kudir_entries.insert_one(entry)


_LEGACY_PAYMENT_ID_PREFIX = 'legacy-'


def _get_payments(order: dict, side: str) -> list:
    """A side's payments as a list, whether the order already has the new
    `{side}_payments` array or is an old order that only has the single
    client_pp_number/client_paid_date/client_rate fields. Old orders get a
    synthesized one-item list so nothing entered before this feature shipped
    is lost — the synthesized entry's id is prefixed so callers can tell it
    apart from a real, persisted payment."""
    field = f'{side}_payments'
    payments = order.get(field)
    if payments:
        return payments
    paid_field = f'{side}_paid'
    pp_field = f'{side}_pp_number'
    date_field = f'{side}_paid_date'
    pp_date_field = f'{side}_pp_date'
    rate_field = f'{side}_rate'
    if order.get(paid_field) and (order.get(pp_field) or order.get(date_field)):
        return [{
            'id': f"{_LEGACY_PAYMENT_ID_PREFIX}{order.get('id','')}-{side}",
            'pp_number': order.get(pp_field) or '',
            'pp_date': order.get(pp_date_field) or order.get(date_field) or '',
            'amount': float(order.get(rate_field) or 0),
        }]
    return []


async def _sync_kudir_rows_for_payments(order: dict, side: str):
    """One KUDiR row per payment entry in `{side}_payments`. Carrier-side
    rows stay one-row-per-payment-per-order — zero-amount audit lines,
    matching the existing transit-row design (see _create_kudir_transit_row).

    Client-side rows merge across orders: a single bank transfer often
    closes several orders at once (same PP number + date recorded on each
    order's own payments array independently) — those get combined into one
    row with the summed amount/margin and a per-order carrier-cost
    breakdown, matching how the printed act reads, instead of one row per
    order for what was really one payment."""
    if side != 'client':
        row_type = 'transit'
        payments = [p for p in (order.get('carrier_payments') or [])
                    if not str(p.get('id', '')).startswith(_LEGACY_PAYMENT_ID_PREFIX)]
        existing_rows = await db.kudir_entries.find(
            {'row_type': row_type, 'order_id': order['id'], 'payment_id': {'$exists': True}}, {'_id': 0}
        ).to_list(500)
        locked_payment_ids = {r['payment_id'] for r in existing_rows if r.get('manually_edited')}
        delete_filter = {'row_type': row_type, 'order_id': order['id']}
        if locked_payment_ids:
            delete_filter['payment_id'] = {'$exists': True, '$nin': list(locked_payment_ids)}
        else:
            delete_filter['payment_id'] = {'$exists': True}
        await db.kudir_entries.delete_many(delete_filter)
        for p in payments:
            if p.get('id') in locked_payment_ids:
                continue
            date_str = (p.get('pp_date') or '')[:10]
            if not date_str:
                continue
            amount = float(p.get('amount') or 0)
            await db.kudir_entries.insert_one({
                'id': str(uuid.uuid4()),
                'payment_id': p.get('id'),
                'order_id': order['id'], 'order_ids': [order['id']],
                'order_number': order.get('order_number', ''),
                'row_type': row_type,
                'entry_date': date_str,
                'document_ref': f"Плат. поручение № {p.get('pp_number','')} от {fmt_date_ru(date_str)}",
                'created_at': now_iso(),
                'content': (f"Перечислено перевозчику {order.get('carrier_name','')} на сумму {amount:.2f} BYN "
                            f"по заявке № {order.get('order_number','')}"),
                'income_amount': None,
                'note': '',
            })
        return

    import re as _re
    payments = [p for p in (order.get('client_payments') or [])
                if not str(p.get('id', '')).startswith(_LEGACY_PAYMENT_ID_PREFIX)]
    if not payments:
        # Still clear out this order's own rows (e.g. its last payment was
        # just deleted) — but never touch a group row another sibling order
        # still legitimately owns, so scope this to rows where this order is
        # the ONLY member.
        await db.kudir_entries.delete_many(
            {'row_type': 'income', 'order_ids': [order['id']]}
        )
        return

    groups = []
    seen_keys = set()
    for p in payments:
        date_str = (p.get('pp_date') or '')[:10]
        pp_number = (p.get('pp_number') or '').strip()
        if not date_str:
            continue
        key = (pp_number, date_str)
        if key in seen_keys:
            continue
        seen_keys.add(key)

        members = [(order, p)]
        if pp_number:
            siblings = await db.orders.find({
                'id': {'$ne': order['id']},
                'client_payments': {'$elemMatch': {
                    'pp_number': pp_number, 'pp_date': {'$regex': f'^{_re.escape(date_str)}'},
                }},
                'deleted': {'$ne': True}, 'status': {'$ne': 'cancelled'},
            }, {'_id': 0}).to_list(50)
            for sib in siblings:
                for sp in (sib.get('client_payments') or []):
                    if (sp.get('pp_number') or '').strip() == pp_number and (sp.get('pp_date') or '')[:10] == date_str:
                        members.append((sib, sp))
                        break
        groups.append((pp_number, date_str, members))

    for pp_number, date_str, members in groups:
        # This order can belong to several *different* groups at once (one
        # payment grouped with order B, another payment grouped with C+D) —
        # deleting by "any row touching any member order" would also wipe
        # out those other groups' rows the moment they share so much as one
        # member with this one, without regenerating them (they're not in
        # `members` here). Scope the delete to exactly this group's own row
        # — same date + same document reference — instead.
        document_ref = f"Плат. поручение № {pp_number} от {fmt_date_ru(date_str)}" if pp_number else ''
        delete_q = {'row_type': 'income', 'entry_date': date_str}
        if pp_number:
            delete_q['document_ref'] = document_ref
        else:
            delete_q['order_ids'] = [mo['id'] for mo, _ in members]
        # A row someone hand-corrected in the book (PATCH /kudir/entries)
        # must survive this resync — leave this specific group alone.
        if await db.kudir_entries.find_one({**delete_q, 'manually_edited': True}):
            continue
        await db.kudir_entries.delete_many(delete_q)
        total_amount = sum(float(mp.get('amount') or 0) for _, mp in members)
        order_numbers = sorted(mo.get('order_number') or '' for mo, _ in members)
        orders_label = ', '.join(n for n in order_numbers if n)
        client_names = sorted({mo.get('client_name') or '' for mo, _ in members} - {''})
        client_label = client_names[0] if len(client_names) <= 1 else ', '.join(client_names)
        carrier_notes = sorted({
            f"{float(mo.get('carrier_rate') or 0):.2f} — перевозчику {mo.get('carrier_name')}"
            for mo, _ in members if mo.get('carrier_name')
        })
        total_margin = 0.0
        partial = False
        for mo, mp in members:
            rate = float(mo.get('client_rate') or 0)
            margin = max(0.0, rate - float(mo.get('carrier_rate') or 0))
            amt = float(mp.get('amount') or 0)
            share = (amt / rate) if rate else 0
            total_margin += margin * share
            if abs(amt - rate) > 0.01 or len(mo.get('client_payments') or []) > 1:
                partial = True

        await db.kudir_entries.insert_one({
            'id': str(uuid.uuid4()),
            'payment_id': members[0][1].get('id') if len(members) == 1 else None,
            'order_id': members[0][0]['id'],
            'order_ids': [mo['id'] for mo, _ in members],
            'order_number': orders_label,
            'row_type': 'income',
            'entry_date': date_str,
            'document_ref': document_ref,
            'created_at': now_iso(),
            'content': (f"{'Частичная оплата' if partial else 'Оплата'} от {client_label} на сумму {total_amount:.2f} BYN "
                        f"по заявк{'е' if len(members) == 1 else 'ам'} № {orders_label}"),
            'income_amount': round(total_margin, 2),
            'note': '; '.join(carrier_notes),
        })


async def _sync_kudir_row(order_id: str, side: str):
    """Keeps a KUDiR row in sync with an order's current payment state —
    creates/recomputes it when paid+PP+date are all present, and removes
    the order from any existing row when they're not (payment unmarked, or
    the PP/date got cleared). Called from both mark_payment and the generic
    order-update endpoint so unmarking a payment there also clears the
    book, not just marking one adds to it."""
    row_type = 'income' if side == 'client' else 'transit'
    paid_field = 'client_paid' if side == 'client' else 'carrier_paid'
    pp_field = 'client_pp_number' if side == 'client' else 'carrier_pp_number'
    date_field = 'client_paid_date' if side == 'client' else 'carrier_paid_date'
    create_fn = _create_kudir_income_row if side == 'client' else _create_kudir_transit_row

    order = await db.orders.find_one({'id': order_id}, {'_id': 0})
    if not order:
        return

    # Multi-payment orders (payments/{side} endpoints) skip the legacy
    # single-PP merge-by-group logic entirely — they get one row per
    # payment instead, kept in sync here so the generic order-update
    # endpoint's retroactive resync (below) doesn't clobber them.
    if order.get(f'{side}_payments'):
        await _sync_kudir_rows_for_payments(order, side)
        return

    if order.get(paid_field) and order.get(pp_field) and order.get(date_field):
        await create_fn(order)
        return

    existing = await db.kudir_entries.find_one({'row_type': row_type, 'order_ids': order_id})
    if not existing:
        return
    if existing.get('manually_edited'):
        return
    await db.kudir_entries.delete_one({'id': existing['id']})
    remaining_ids = [oid for oid in existing.get('order_ids', []) if oid != order_id]
    if remaining_ids:
        remaining_order = await db.orders.find_one({'id': remaining_ids[0]}, {'_id': 0})
        if remaining_order:
            await create_fn(remaining_order)


class PaymentMark(BaseModel):
    paid: bool
    pp_number: Optional[str] = None
    pp_date: Optional[str] = None


@api_router.post("/orders/{order_id}/mark_payment")
async def mark_payment(order_id: str, side: str, payload: PaymentMark, background_tasks: BackgroundTasks,
                        current_user: Optional[dict] = Depends(_get_user_from_token)):
    if side not in ("client", "carrier"):
        raise HTTPException(400, "side must be 'client' or 'carrier'")

    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Заявка не найдена")

    paid_field = "client_paid" if side == "client" else "carrier_paid"
    date_field = "client_paid_date" if side == "client" else "carrier_paid_date"
    pp_field = "client_pp_number" if side == "client" else "carrier_pp_number"
    pp_date_field = "client_pp_date" if side == "client" else "carrier_pp_date"

    pay_date = payload.pp_date or now_iso()[:10]
    order_update = OrderUpdate(**{
        paid_field: payload.paid,
        date_field: pay_date if payload.paid else None,
        pp_field: (payload.pp_number or None) if payload.paid else None,
        pp_date_field: (payload.pp_date or None) if payload.paid else None,
    })
    # Reuses update_order()'s existing side effects: task closing, carrier
    # payment deadline calc, order_logs audit trail, calendar/sheet sync.
    await update_order(order_id, order_update, background_tasks, current_user)
    # Belt-and-suspenders direct write for the PP fields specifically — seen
    # cases where they didn't make it onto the order via the OrderUpdate
    # round-trip above even though the payments_in record below always got
    # them right, leaving the order itself showing no PP number.
    if payload.paid:
        await db.orders.update_one({"id": order_id}, {"$set": {
            pp_field: payload.pp_number or "", pp_date_field: payload.pp_date or pay_date,
        }})

    if payload.paid:
        amount = order.get("client_rate") if side == "client" else order.get("carrier_rate")
        collection = db.payments_in if side == "client" else db.payments_out
        record = {
            "id": str(uuid.uuid4()),
            "pp_number": payload.pp_number or "",
            "date": pay_date,
            "amount": amount or 0,
            "order_id": order_id,
            "order_number": order.get("order_number") or "",
            "notes": f"Заявка {order.get('order_number') or order_id}",
            "created_at": now_iso(),
        }
        if side == "client":
            record["client_id"] = order.get("client_id") or ""
            record["client_name"] = order.get("client_name") or ""
        else:
            record["carrier_id"] = order.get("carrier_id") or ""
            record["carrier_name"] = order.get("carrier_name") or ""
        await collection.insert_one(record)

        # Auto-close the payment-reminder task for this order+side the moment
        # the payment is marked (mirrors _sync_payment_reminders, which
        # creates/updates it on a schedule).
        await db.tasks.update_many(
            {"order_id": order_id, "type": "payment_reminder", "side": side, "status": "pending"},
            {"$set": {"status": "done", "completed_at": now_iso()}},
        )

    # Syncs the KUDiR row either way: creates/recomputes it when paid with a
    # PP+date, and removes it (or shrinks a merged group) when unmarked.
    await _sync_kudir_row(order_id, side)

    await manager.broadcast({"type": "payment_marked", "order_id": order_id, "side": side})
    return {
        "ok": True,
        paid_field: payload.paid,
        date_field: pay_date if payload.paid else None,
        pp_field: payload.pp_number if payload.paid else None,
        pp_date_field: payload.pp_date if payload.paid else None,
    }


class PaymentEntry(BaseModel):
    pp_number: str
    pp_date: str
    amount: float


@api_router.post("/orders/{order_id}/payments/{side}")
async def add_payment(order_id: str, side: str, payload: PaymentEntry, background_tasks: BackgroundTasks,
                       current_user: Optional[dict] = Depends(_get_user_from_token)):
    """Adds one partial-payment PP to an order's side. Several of these can
    add up to the order's rate (client_rate/carrier_rate) — the side is only
    flagged paid once the sum matches; each PP still gets its own KUDiR row
    (see _sync_kudir_rows_for_payments) so partial payments show up in the
    book as they land instead of waiting for the last one."""
    if side not in ("client", "carrier"):
        raise HTTPException(400, "side must be 'client' or 'carrier'")
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Заявка не найдена")

    field = f"{side}_payments"
    rate_field = f"{side}_rate"
    paid_field = f"{side}_paid"
    date_field = f"{side}_paid_date"
    expected = float(order.get(rate_field) or 0)

    payments = [p for p in _get_payments(order, side)
                if not str(p.get("id", "")).startswith(_LEGACY_PAYMENT_ID_PREFIX)]
    payments.append({"id": str(uuid.uuid4()), **payload.dict()})
    total = sum(float(p.get("amount") or 0) for p in payments)
    fully_paid = abs(total - expected) < 0.01

    set_data = {field: payments, paid_field: fully_paid}
    if fully_paid and not order.get(date_field):
        set_data[date_field] = payload.pp_date
    await db.orders.update_one({"id": order_id}, {"$set": set_data})
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0})

    await _sync_kudir_rows_for_payments(updated, side)

    background_tasks.add_task(_bg_push_order, updated)
    await manager.broadcast({"type": "payment_marked", "order_id": order_id, "side": side})
    return {"ok": True, "payments": payments, "total": total, "expected": expected, "fully_paid": fully_paid}


@api_router.delete("/orders/{order_id}/payments/{side}/{payment_id}")
async def delete_payment(order_id: str, side: str, payment_id: str, background_tasks: BackgroundTasks,
                          current_user: Optional[dict] = Depends(_get_user_from_token)):
    if side not in ("client", "carrier"):
        raise HTTPException(400, "side must be 'client' or 'carrier'")
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Заявка не найдена")

    field = f"{side}_payments"
    rate_field = f"{side}_rate"
    paid_field = f"{side}_paid"
    expected = float(order.get(rate_field) or 0)

    payments = [p for p in _get_payments(order, side)
                if p.get("id") != payment_id and not str(p.get("id", "")).startswith(_LEGACY_PAYMENT_ID_PREFIX)]
    total = sum(float(p.get("amount") or 0) for p in payments)
    fully_paid = bool(payments) and abs(total - expected) < 0.01

    await db.orders.update_one({"id": order_id}, {"$set": {field: payments, paid_field: fully_paid}})
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0})

    await _sync_kudir_rows_for_payments(updated, side)

    background_tasks.add_task(_bg_push_order, updated)
    await manager.broadcast({"type": "payment_marked", "order_id": order_id, "side": side})
    return {"ok": True, "payments": payments, "total": total, "expected": expected, "fully_paid": fully_paid}


def _has_pp(order: dict, side: str) -> bool:
    """True if the side has a real PP number — either the legacy single
    field, or at least one entry in the payments list with a non-blank
    pp_number (a payments list can exist with every pp_number left blank,
    e.g. saved via 'дозаполнить позже' — that still counts as missing)."""
    if order.get(f'{side}_pp_number'):
        return True
    return any(p.get('pp_number') for p in (order.get(f'{side}_payments') or []))


@api_router.get("/kudir/missing_pp")
async def missing_pp_orders(current_user: dict = Depends(require_director)):
    import re as _re
    # Broad query (any paid order) — precise "actually missing a PP" check
    # happens in Python below via _has_pp, since that requires inspecting
    # the payments list contents, not just whether the field is empty.
    candidates = await db.orders.find({
        'deleted': {'$ne': True}, 'status': {'$ne': 'cancelled'},
        '$or': [{'client_paid': True}, {'carrier_paid': True}],
    }, {'_id': 0}).to_list(10000)

    orders = []
    for o in candidates:
        missing = []
        if o.get('client_paid') and not _has_pp(o, 'client'):
            missing.append('client_pp')
        if o.get('carrier_paid') and not _has_pp(o, 'carrier'):
            missing.append('carrier_pp')
        if o.get('carrier_paid') and (o.get('docs_from_client_received') or o.get('docs_from_carrier_received')) and not o.get('carrier_act_number'):
            missing.append('act')
        if not missing:
            continue
        o['missing_fields'] = missing
        orders.append(o)

    # Order by заявка number (год, номер) — "по порядку с 1 заявки" — falling
    # back to created_at for order_number formats the usual "З-NNN/YYYY"
    # regex doesn't match, so those still sort somewhere stable instead of
    # crashing the endpoint.
    _order_num_pat = _re.compile(r"[ЗЗз3]\s*[-–—]\s*(\d+)\s*/\s*(\d{4})", _re.IGNORECASE)

    def _sort_key(o):
        m = _order_num_pat.search(o.get('order_number', '') or '')
        if m:
            return (0, int(m.group(2)), int(m.group(1)))
        return (1, o.get('created_at', '') or '', 0)

    orders.sort(key=_sort_key)

    return {'orders': orders, 'total': len(orders)}


@api_router.get("/kudir/entries")
async def list_kudir_entries(date_from: Optional[str] = None, date_to: Optional[str] = None,
                              current_user: dict = Depends(require_director)):
    q: dict = {}
    if date_from or date_to:
        q['entry_date'] = {}
        if date_from:
            q['entry_date']['$gte'] = date_from
        if date_to:
            q['entry_date']['$lte'] = date_to
    entries = await db.kudir_entries.find(q, {'_id': 0}).sort('entry_date', 1).to_list(20000)
    total_income = sum(e.get('income_amount') or 0 for e in entries)
    return {'entries': entries, 'total_income': total_income, 'count': len(entries)}


_KUDIR_EXTRA_COLUMNS = {
    # графа 5..10 официального бланка (Приложение 9) — не заполняются
    # автосинхронизацией (только графа 4/E — доход, и графа 11/L —
    # примечание — считаются из заявки), но доступны для ручного заполнения
    # прямо в CRM, чтобы веб-вид один в один совпадал с печатной формой.
    'tax_from_revenue',  # графа 5 (F) — сумма налогов из выручки
    'exempt_income',     # графа 6 (G) — освобождаемые доходы
    'other_income',      # графа 7 (H) — иные поступления
    'expense_period',    # графа 8 (I) — расходы за отчётный период
    'expense_norm',      # графа 9 (J) — расходы по нормативу
    'expense_other',     # графа 10 (K) — иные расходы
}


@api_router.post("/kudir/entries")
async def create_kudir_entry(payload: dict, current_user: dict = Depends(require_director)):
    """A hand-added row not tied to any order — for a real bank payment
    (found while reconciling the statement) that has no matching order to
    attach it to, e.g. because that order was deleted after the payment was
    recorded. Locked (manually_edited) from creation since there's no order
    to ever auto-resync it from."""
    allowed = {'content', 'note', 'document_ref', 'income_amount', 'entry_date', 'order_number', 'row_type'} | _KUDIR_EXTRA_COLUMNS
    entry = {k: v for k, v in payload.items() if k in allowed}
    entry.setdefault('row_type', 'income')
    if not entry.get('entry_date'):
        raise HTTPException(400, "entry_date обязателен")
    entry.update({
        'id': str(uuid.uuid4()),
        'order_id': None, 'order_ids': [],
        'order_number': entry.get('order_number', ''),
        'created_at': now_iso(),
        'manually_edited': True, 'edited_at': now_iso(),
    })
    await db.kudir_entries.insert_one(entry)
    return {'ok': True, 'id': entry['id']}


@api_router.patch("/kudir/entries/{entry_id}")
async def update_kudir_entry(entry_id: str, payload: dict, current_user: dict = Depends(require_director)):
    allowed = {'content', 'note', 'document_ref', 'income_amount', 'entry_date'} | _KUDIR_EXTRA_COLUMNS
    upd = {k: v for k, v in payload.items() if k in allowed}
    # Marks the row as hand-corrected so the order-driven auto-sync
    # (_create_kudir_income_row / _create_kudir_transit_row /
    # _sync_kudir_rows_for_payments) leaves it alone from now on instead of
    # silently deleting and rebuilding it from the order the next time
    # anything about that order's payment is touched.
    if upd:
        upd['manually_edited'] = True
        upd['edited_at'] = now_iso()
        await db.kudir_entries.update_one({'id': entry_id}, {'$set': upd})
    return {'ok': True}


@api_router.delete("/kudir/entries/{entry_id}")
async def delete_kudir_entry(entry_id: str, current_user: dict = Depends(require_director)):
    """Removes a book row outright — for duplicate/erroneous rows the
    auto-sync won't touch (rows created without a payment_id, e.g. by the
    legacy grouped-PP path, are never cleaned up automatically). Zeroing a
    row's amount via PATCH leaves the row itself behind; this is for when
    the row shouldn't exist at all."""
    await db.kudir_entries.delete_one({'id': entry_id})
    return {'ok': True}


@api_router.post("/kudir/entries/{entry_id}/unlock")
async def unlock_kudir_entry(entry_id: str, current_user: dict = Depends(require_director)):
    """Reverts a hand-corrected row back to auto-sync — the next payment
    change on its order(s) will recompute and overwrite it again."""
    await db.kudir_entries.update_one({'id': entry_id}, {'$unset': {'manually_edited': '', 'edited_at': ''}})
    return {'ok': True}


@api_router.post("/kudir/resync")
async def resync_kudir(current_user: dict = Depends(require_director)):
    """One-off sweep: re-runs _sync_kudir_rows_for_payments for every order
    still on the payments-array format, so the cross-order PP+date grouping
    added there applies to rows created before that logic existed instead of
    only to orders touched going forward."""
    orders = await db.orders.find(
        {'deleted': {'$ne': True}, 'client_payments': {'$exists': True, '$ne': []}}, {'_id': 0}
    ).to_list(5000)
    for o in orders:
        await _sync_kudir_rows_for_payments(o, 'client')
    return {'ok': True, 'orders_processed': len(orders)}


@api_router.get("/kudir/export")
async def export_kudir(year: int, quarter: Optional[int] = None, token: Optional[str] = None,
                        current_user: Optional[dict] = Depends(_get_user_from_token)):
    # Opened via plain browser navigation (so the file downloads itself) —
    # that means no Authorization header, so a `token` query param is
    # accepted as a fallback and validated the same way the header would be.
    user = current_user
    if not user and token:
        try:
            token_payload = _jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            user = await db.users.find_one({"id": token_payload.get("sub")}, {"_id": 0})
        except Exception:
            user = None
    if not user or user.get("role") not in DIRECTOR_ROLES:
        raise HTTPException(403, "Доступно только директору")

    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from kudir_template import build_header, build_table_head, style_data_row, border_all, f as font_fn, left_bottom, set_widths

    quarters = [quarter] if quarter else [1, 2, 3, 4]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}" + (f" Q{quarter}" if quarter else "")
    set_widths(ws)
    build_header(ws, "29.03.2024 N 11")
    build_table_head(ws)

    row = 15
    prev_ytd_row = None

    for q in quarters:
        q_start_month = (q - 1) * 3 + 1
        date_from = f"{year}-{q_start_month:02d}-01"
        q_end_month = q_start_month + 2
        last_day = 31 if q_end_month in (1, 3, 5, 7, 8, 10, 12) else (30 if q_end_month != 2 else 29)
        date_to = f"{year}-{q_end_month:02d}-{last_day}"

        entries = await db.kudir_entries.find({
            'entry_date': {'$gte': date_from, '$lte': date_to}
        }, {'_id': 0}).sort('entry_date', 1).to_list(20000)

        quarter_start_row = row
        for e in entries:
            style_data_row(ws, row)
            ws[f"B{row}"] = fmt_date_ru(e['entry_date'])
            ws[f"C{row}"] = e['document_ref']
            ws[f"D{row}"] = e['content']
            if e.get('income_amount') is not None:
                ws[f"E{row}"] = round(e['income_amount'], 2)
            ws[f"L{row}"] = e.get('note', '')
            row += 1
        quarter_end_row = row - 1

        ws[f"B{row}"] = f"Итого за {q}-й квартал"
        ws[f"B{row}"].font = font_fn(10)
        ws[f"B{row}"].alignment = left_bottom
        ws.merge_cells(f"B{row}:D{row}")
        for col in "EFGHIJK":
            ws[f"{col}{row}"] = f"=SUM({col}{quarter_start_row}:{col}{quarter_end_row})" if quarter_end_row >= quarter_start_row else 0
            ws[f"{col}{row}"].border = border_all
            ws[f"{col}{row}"].font = font_fn(9)
        ws[f"B{row}"].border = border_all
        quarter_total_row = row
        row += 1

        ws[f"B{row}"] = "Итого с начала календарного года"
        ws[f"B{row}"].font = font_fn(10)
        ws[f"B{row}"].alignment = left_bottom
        ws.merge_cells(f"B{row}:D{row}")
        for col in "EFGHIJK":
            ws[f"{col}{row}"] = f"={col}{quarter_total_row}" if prev_ytd_row is None else f"={col}{prev_ytd_row}+{col}{quarter_total_row}"
            ws[f"{col}{row}"].border = border_all
            ws[f"{col}{row}"].font = font_fn(9)
        ws[f"B{row}"].border = border_all
        prev_ytd_row = row
        row += 2

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Книга учёта — {year}" + (f" кв.{quarter}" if quarter else " год") + ".xlsx"
    # Header values must be latin-1 — the Cyrillic name only fits via the
    # RFC 5987 filename* form, with an ASCII fallback for older clients.
    from urllib.parse import quote
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"kudir_{year}.xlsx\"; filename*=UTF-8''{quote(filename)}"}
    )


@api_router.post("/orders/{order_id}/sync_doc_urls")
async def sync_order_doc_urls(order_id: str):
    """Read doc URLs for this order from Google Sheet and update MongoDB."""
    doc = await db.orders.find_one({"id": order_id}, {"_id": 0, "order_number": 1})
    if not doc:
        raise HTTPException(404, "Order not found")
    order_number = doc.get("order_number", "")
    if not order_number:
        return {"ok": False, "message": "No order_number"}

    if _sheets_run_import is None:
        return {"ok": False, "message": "Sheets not configured"}

    try:
        def _fetch_doc_urls():
            from sheets_import import SheetsImporter
            imp = SheetsImporter()
            rows = imp._read_tab("Заказы")
            for row in rows[4:]:
                if len(row) > 0 and row[0].strip() == order_number.strip():
                    return {
                        "doc_url_client":  row[34].strip() if len(row) > 34 else "",
                        "doc_url_carrier": row[35].strip() if len(row) > 35 else "",
                        "doc_url_act":     row[36].strip() if len(row) > 36 else "",
                    }
            return None

        urls = await asyncio.to_thread(_fetch_doc_urls)
        if not urls:
            return {"ok": False, "message": f"Order {order_number} not found in sheet"}

        update = {k: v for k, v in urls.items() if v}
        if update:
            await db.orders.update_one({"id": order_id}, {"$set": update})
        return {"ok": True, **urls}
    except Exception as e:
        logging.getLogger(__name__).error(f"sync_doc_urls failed: {e}", exc_info=True)
        raise HTTPException(500, str(e))


@api_router.get("/orders/{order_id}/logs")
async def get_order_logs(order_id: str):
    logs = await db.order_logs.find({"order_id": order_id}, {"_id": 0}).sort("timestamp", -1).to_list(1000)
    return logs


@api_router.delete("/orders/{order_id}")
async def delete_order(order_id: str, background_tasks: BackgroundTasks,
                        current_user: dict = Depends(require_director)):
    # Primary lookup by UUID `id` string field (how the app stores orders)
    doc = await db.orders.find_one({"id": order_id})
    if not doc:
        # Fallback: order imported from Sheets may only have order_number
        doc = await db.orders.find_one({"order_number": order_id})
    if not doc:
        logging.getLogger(__name__).warning(f"[delete_order] {order_id!r} not found")
        raise HTTPException(status_code=404, detail="Заявка не найдена")

    # Already soft-deleted — idempotent, return success
    if doc.get("deleted"):
        return {"ok": True}

    deleted_at = now_iso()
    result = await db.orders.update_one(
        {"_id": doc["_id"]},  # use Mongo _id for exact match, no ambiguity
        {"$set": {"deleted": True, "deleted_at": deleted_at}},
    )
    logging.getLogger(__name__).info(
        f"[delete_order] order_number={doc.get('order_number')!r} "
        f"id={doc.get('id')!r} matched={result.matched_count} modified={result.modified_count}"
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=500, detail="Не удалось удалить заявку")

    if doc.get("order_number"):
        background_tasks.add_task(_bg_delete_order_row, doc["order_number"])
    if doc.get("calendar_event_id"):
        background_tasks.add_task(_bg_cal_delete, doc["calendar_event_id"])
    return {"ok": True}


@api_router.post("/orders/{order_id}/duplicate", response_model=Order)
async def duplicate_order(order_id: str, background_tasks: BackgroundTasks,
                           current_user: dict = Depends(_require_user)):
    doc = await db.orders.find_one({"id": order_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Order not found")
    import re as _re
    year = datetime.now(timezone.utc).year
    all_docs = await db.orders.find({"deleted": {"$ne": True}}, {"_id": 0, "order_number": 1}).to_list(5000)
    pattern = _re.compile(r"[ЗЗз3]\s*[-–—]\s*(\d+)\s*/\s*(\d{4})", _re.IGNORECASE)
    max_num = 0
    for d in all_docs:
        m = pattern.search(d.get("order_number", "") or "")
        if not m:
            continue
        n, y = int(m.group(1)), int(m.group(2))
        if y == year and n > max_num:
            max_num = n
    next_number = f"З-{max_num + 1:03d}/{year}"
    exclude = {"_id", "order_number", "created_at", "status", "client_paid", "carrier_paid",
               "client_paid_date", "carrier_paid_date", "calendar_event_id",
               "letter_to_client_sent", "letter_from_client_received", "is_overdue",
               "carrier_payment_deadline", "carrier_payment_reminder_date"}
    new_data = {k: v for k, v in doc.items() if k not in exclude}
    new_data["id"] = str(uuid.uuid4())
    new_data["order_number"] = next_number
    new_data["status"] = "new"
    new_data["client_paid"] = False
    new_data["carrier_paid"] = False
    new_data["client_paid_date"] = ""
    new_data["carrier_paid_date"] = ""
    new_data["calendar_event_id"] = ""
    new_data["created_at"] = now_iso()
    new_data["created_by"] = current_user["id"]
    obj = Order(**new_data)
    await db.orders.insert_one(obj.dict())
    background_tasks.add_task(_bg_push_order, obj.dict())
    background_tasks.add_task(_bg_cal_create, obj.dict())
    return obj


# ====== OAuth Google (для генерации Docs от имени пользователя) ======
from fastapi.responses import HTMLResponse, RedirectResponse


@api_router.get("/auth/google/start")
async def auth_google_start(request: Request = None):  # type: ignore  # noqa
    """Возвращает auth_url для перехода пользователя на consent screen Google."""
    if _oauth_build_auth_url is None:
        raise HTTPException(500, "OAuth не настроен")
    try:
        auth_url, state = _oauth_build_auth_url()
        redirect_uri = _oauth_get_redirect()
        return {"auth_url": auth_url, "state": state, "redirect_uri": redirect_uri}
    except Exception as e:
        raise HTTPException(500, f"OAuth start failed: {e}")


@api_router.get("/auth/google/callback")
async def auth_google_callback(code: str = "", state: str = "", error: str = ""):
    """Принимает code от Google, обменивает на refresh_token и сохраняет в Mongo."""
    if error:
        return HTMLResponse(_callback_html("Ошибка авторизации: " + error, success=False))
    if not code:
        return HTMLResponse(_callback_html("Не получен код авторизации", success=False))
    if _oauth_fetch_token is None:
        return HTMLResponse(_callback_html("OAuth не настроен на сервере", success=False))

    try:
        token = _oauth_fetch_token(code=code, state=state)
        # Google omits refresh_token on repeat consent (unless prompt=consent forced it);
        # keep the previously stored one instead of nulling it out on upsert.
        existing = await db.oauth_tokens.find_one({"_id": "google"})
        refresh_token = token.get("refresh_token") or (existing.get("refresh_token") if existing else None)
        token_doc = {
            "_id": "google",
            "access_token": token.get("access_token"),
            "refresh_token": refresh_token,
            "expiry": token.get("expires_at"),
            "scopes": _OAUTH_SCOPES,
            "saved_at": now_iso(),
        }
        # сохраняем (upsert)
        await db.oauth_tokens.replace_one({"_id": "google"}, token_doc, upsert=True)
        # сбросим кеш сервисов в docs_gen
        if _docs_get_generator is not None:
            try:
                _docs_get_generator()._drive = None
                _docs_get_generator()._docs = None
            except Exception:
                pass
        # Redirect away from the code-bearing URL so a page refresh / back-forward
        # replay can't resend the already-consumed authorization code to Google
        # (that resend is what produces invalid_grant).
        return RedirectResponse(url="/api/auth/google/callback/done?ok=1")
    except Exception as e:
        logging.getLogger(__name__).error(f"OAuth callback failed: {e}", exc_info=True)
        from urllib.parse import quote
        return RedirectResponse(url=f"/api/auth/google/callback/done?ok=0&msg={quote(str(e))}")


@api_router.get("/auth/google/callback/done")
async def auth_google_callback_done(ok: int = 1, msg: str = ""):
    """Статическая страница результата — не трогает code/state, безопасна для перезагрузки."""
    if ok:
        return HTMLResponse(_callback_html("Авторизация успешна! Можно закрыть это окно и вернуться в приложение.", success=True))
    return HTMLResponse(_callback_html(msg or "Ошибка авторизации", success=False))


@api_router.get("/auth/google/status")
async def auth_google_status():
    doc = await db.oauth_tokens.find_one({"_id": "google"}, {"_id": 0, "access_token": 0})
    if not doc:
        return {"connected": False}
    return {
        "connected": bool(doc.get("refresh_token")),
        "saved_at": doc.get("saved_at"),
        "scopes": doc.get("scopes", []),
    }


@api_router.delete("/auth/google")
async def auth_google_disconnect():
    await db.oauth_tokens.delete_one({"_id": "google"})
    if _docs_get_generator is not None:
        try:
            _docs_get_generator()._drive = None
            _docs_get_generator()._docs = None
        except Exception:
            pass
    return {"ok": True}


def _callback_html(msg: str, success: bool = True) -> str:
    color = "#0A8C3E" if success else "#D4351C"
    return f"""<!doctype html>
<html><head><meta charset='utf-8'><title>Google Auth</title>
<style>
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
       background: #0A0A0C; color: #fff; display: flex; align-items: center; justify-content: center;
       min-height: 100vh; margin: 0; }}
.box {{ max-width: 480px; padding: 40px; text-align: center; }}
.icon {{ width: 64px; height: 64px; margin: 0 auto 16px; border-radius: 50%;
        background: {color}; display: flex; align-items: center; justify-content: center;
        font-size: 32px; }}
h1 {{ font-size: 22px; font-weight: 600; margin: 0 0 8px; }}
p {{ color: #A1A1AA; line-height: 1.5; }}
</style></head>
<body>
<div class='box'>
  <div class='icon'>{('✓' if success else '✕')}</div>
  <h1>{('Готово' if success else 'Не получилось')}</h1>
  <p>{msg}</p>
</div></body></html>
"""


# ====== User Auth & Management ======
class _LoginPayload(BaseModel):
    login: str
    password: str


class _CreateUserPayload(BaseModel):
    name: str
    login: str
    password: str
    role: str = "manager"
    permissions: Optional[dict] = None
    daily_call_goal: Optional[int] = 45


LOGIN_LOCK_THRESHOLD = 5
LOGIN_LOCK_MINUTES = 15


@api_router.post("/auth/login")
async def auth_login(payload: _LoginPayload, request: Request):
    user = await db.users.find_one({"login": payload.login}, {"_id": 0})
    if not user:
        raise HTTPException(401, "Неверный логин или пароль")

    if user.get("status") == "suspended":
        raise HTTPException(403, "Доступ приостановлен. Обратитесь к директору.")

    locked_until = user.get("locked_until")
    if locked_until:
        try:
            if datetime.now(timezone.utc) < datetime.fromisoformat(locked_until):
                raise HTTPException(429, "Слишком много попыток. Попробуйте через 15 минут.")
        except ValueError:
            pass

    if not _verify_password(payload.password, user.get("password_hash", "")):
        attempts = int(user.get("failed_attempts") or 0) + 1
        upd: dict = {"failed_attempts": attempts}
        if attempts >= LOGIN_LOCK_THRESHOLD:
            upd["locked_until"] = (datetime.now(timezone.utc) + timedelta(minutes=LOGIN_LOCK_MINUTES)).isoformat()
        await db.users.update_one({"id": user["id"]}, {"$set": upd})
        raise HTTPException(401, "Неверный логин или пароль")

    await db.users.update_one({"id": user["id"]}, {"$set": {"failed_attempts": 0, "locked_until": None}})

    session_id = str(uuid.uuid4())
    now_dt = datetime.now(timezone.utc)
    await db.sessions.insert_one({
        "id": session_id, "user_id": user["id"], "user_login": user["login"], "role": user["role"],
        "device_info": request.headers.get("user-agent", ""),
        "ip": request.client.host if request.client else "",
        "created_at": now_dt.isoformat(), "last_activity": now_dt.isoformat(), "active": True,
    })

    token = _create_token(user["id"], user["role"], session_id)
    safe_user = {k: v for k, v in user.items() if k != "password_hash"}

    if user["role"] == "manager":
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()), "type": "login",
            "message": f"{user.get('name', user['login'])} ({user['login']}) вошёл в систему",
            "created_at": now_dt.isoformat(), "read": False,
        })

    return {"token": token, "user": safe_user, "session_id": session_id}


@api_router.get("/auth/me")
async def auth_me(current_user: dict = Depends(_require_user)):
    return {k: v for k, v in current_user.items() if k != "password_hash"}


@api_router.get("/users")
async def list_users(role: Optional[str] = None, current_user: dict = Depends(require_director)):
    filter_q = {"role": role} if role else {}
    users = await db.users.find(filter_q, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users


@api_router.post("/users")
async def create_user(payload: _CreateUserPayload, current_user: dict = Depends(require_director)):
    existing = await db.users.find_one({"login": payload.login})
    if existing:
        raise HTTPException(400, "Логин уже занят")
    user = {
        "id": str(uuid.uuid4()),
        "name": payload.name,
        "login": payload.login,
        "password_hash": _hash_password(payload.password),
        "role": payload.role,
        "permissions": payload.permissions or {},
        "status": "active",
        "daily_call_goal": payload.daily_call_goal if payload.role == "manager" else None,
        "failed_attempts": 0,
        "locked_until": None,
        "created_at": now_iso(),
    }
    await db.users.insert_one(user)
    return {k: v for k, v in user.items() if k not in ("password_hash", "_id")}


@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(require_director)):
    if user_id == current_user.get("id"):
        raise HTTPException(400, "Нельзя удалить себя")
    target = await db.users.find_one({"id": user_id})
    if target and target.get("role") in DIRECTOR_ROLES:
        admin_count = await db.users.count_documents({"role": {"$in": list(DIRECTOR_ROLES)}})
        if admin_count <= 1:
            raise HTTPException(400, "Нельзя удалить последнего администратора")
    await db.users.delete_one({"id": user_id})
    await db.sessions.update_many({"user_id": user_id}, {"$set": {"active": False}})
    return {"ok": True}


@api_router.get("/users/activity_summary")
async def get_users_activity_summary(current_user: dict = Depends(require_director)):
    now_dt = datetime.now(timezone.utc)
    month_start = f"{now_dt.year}-{now_dt.month:02d}-01"
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    # Count orders_created per user
    created_agg = await db.orders.aggregate([
        {"$match": {"created_at": {"$gte": month_start}, "deleted": {"$ne": True}}},
        {"$group": {"_id": "$created_by", "count": {"$sum": 1}}},
    ]).to_list(1000)
    orders_created_by_user = {item["_id"]: item["count"] for item in created_agg}
    # Count orders_assigned per user
    assigned_agg = await db.orders.aggregate([
        {"$match": {"created_at": {"$gte": month_start}, "deleted": {"$ne": True}}},
        {"$group": {"_id": "$assigned_to", "count": {"$sum": 1}}},
    ]).to_list(1000)
    orders_assigned_by_user = {item["_id"]: item["count"] for item in assigned_agg}
    # Count orders_total (created_by OR assigned_to, unique per user)
    all_month_orders = await db.orders.find(
        {"created_at": {"$gte": month_start}, "deleted": {"$ne": True}},
        {"_id": 0, "created_by": 1, "assigned_to": 1},
    ).to_list(10000)
    orders_total_by_user: dict = {}
    for o in all_month_orders:
        for uid in set(filter(None, [o.get("created_by"), o.get("assigned_to")])):
            orders_total_by_user[uid] = orders_total_by_user.get(uid, 0) + 1
    leads_agg = await db.leads.aggregate([
        {"$group": {"_id": "$assigned_to", "lead_ids": {"$push": "$id"}}},
    ]).to_list(1000)
    leads_by_user = {item["_id"]: item["lead_ids"] for item in leads_agg}
    start_30 = (now_dt - timedelta(days=29)).strftime("%Y-%m-%d")
    activity_docs = await db.lead_activity.find(
        {"date": {"$gte": start_30}}, {"_id": 0, "lead_id": 1}
    ).to_list(50000)
    activity_by_lead: dict = {}
    for doc in activity_docs:
        lid = doc.get("lead_id")
        if lid:
            activity_by_lead[lid] = activity_by_lead.get(lid, 0) + 1
    result = []
    for u in users:
        uid = u["id"]
        user_leads = leads_by_user.get(uid, [])
        calls = sum(activity_by_lead.get(lid, 0) for lid in user_leads)
        orders_created = orders_created_by_user.get(uid, 0)
        orders_assigned = orders_assigned_by_user.get(uid, 0)
        orders_total = orders_total_by_user.get(uid, 0)
        result.append({
            **u,
            "orders_month": orders_total,
            "orders_created": orders_created,
            "orders_assigned": orders_assigned,
            "orders_total": orders_total,
            "calls_month": calls,
        })
    return result


class _UpdateUserPayload(BaseModel):
    name: Optional[str] = None
    login: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None
    permissions: Optional[dict] = None
    status: Optional[str] = None
    daily_call_goal: Optional[int] = None


@api_router.put("/users/{user_id}")
async def update_user(user_id: str, payload: _UpdateUserPayload, current_user: dict = Depends(require_director)):
    update_data: dict = {}
    if payload.name is not None:
        update_data["name"] = payload.name
    if payload.login is not None:
        existing = await db.users.find_one({"login": payload.login, "id": {"$ne": user_id}})
        if existing:
            raise HTTPException(400, "Логин уже занят")
        update_data["login"] = payload.login
    if payload.password:
        update_data["password_hash"] = _hash_password(payload.password)
    if payload.role is not None:
        if user_id == current_user.get("id"):
            raise HTTPException(400, "Нельзя изменить свою роль")
        update_data["role"] = payload.role
    if payload.permissions is not None:
        if user_id == current_user.get("id"):
            raise HTTPException(400, "Нельзя изменить свои права")
        update_data["permissions"] = payload.permissions
    if payload.status is not None:
        if payload.status == "suspended" and user_id == current_user.get("id"):
            raise HTTPException(400, "Нельзя заблокировать себя")
        update_data["status"] = payload.status
        if payload.status == "suspended":
            await db.sessions.update_many({"user_id": user_id}, {"$set": {"active": False}})
    if payload.daily_call_goal is not None:
        update_data["daily_call_goal"] = payload.daily_call_goal
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(404, "Пользователь не найден")
    return user


async def _session_activity_summary(user_id: str, since: str) -> dict:
    calls = await db.call_logs.count_documents({"created_by": user_id, "created_at": {"$gte": since}})
    # Activity in the leads section isn't only phone calls — a lead can also
    # be marked won by converting it straight to a client, which doesn't
    # touch call_logs at all. Counted separately so "0 calls" during a
    # session that only did conversions doesn't read as "did nothing".
    won = await db.leads.count_documents({"assigned_to": user_id, "stage": "won", "won_at": {"$gte": since}})
    return {"calls": calls, "won": won}


@api_router.get("/admin/sessions")
async def list_sessions(current_user: dict = Depends(require_director)):
    sessions = await db.sessions.find({"active": True}, {"_id": 0}).sort("last_activity", -1).to_list(200)
    for s in sessions:
        s["activity_summary"] = await _session_activity_summary(s["user_id"], s["created_at"])
    return {"sessions": sessions}


@api_router.post("/admin/sessions/{session_id}/logout")
async def force_logout(session_id: str, current_user: dict = Depends(require_director)):
    await db.sessions.update_one({"id": session_id}, {"$set": {"active": False}})
    return {"ok": True}


async def _compute_manager_stats(period: str, query: Optional[dict] = None) -> list:
    now = datetime.now(timezone.utc)
    if period == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "week":
        start = now - timedelta(days=7)
    elif period == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        start = datetime(2020, 1, 1, tzinfo=timezone.utc)
    start_str = start.isoformat()

    managers = await db.users.find(query or {}, {"_id": 0, "password_hash": 0}).to_list(200)
    result = []
    for m in managers:
        total_calls = await db.call_logs.count_documents({"created_by": m["id"], "created_at": {"$gte": start_str}})
        leads_in_work = await db.leads.count_documents({
            "assigned_to": m["id"], "stage": {"$nin": ["won", "lost", "no_contact"]},
        })
        won = await db.leads.count_documents({"assigned_to": m["id"], "stage": "won", "won_at": {"$gte": start_str}})
        called_leads = await db.leads.count_documents({"assigned_to": m["id"], "total_calls": {"$gt": 0}})
        overdue = await db.leads.count_documents({
            "assigned_to": m["id"], "next_call": {"$ne": None, "$lt": now.isoformat()},
            "stage": {"$nin": ["won", "lost", "no_contact"]},
        })
        conversion = round(won / called_leads * 100, 1) if called_leads else 0.0
        goal_pct = round(total_calls / (m.get("daily_call_goal") or 45) * 100, 1) if period == "today" else None

        result.append({
            "id": m["id"], "name": m["name"], "login": m["login"],
            "calls": total_calls, "goal_pct": goal_pct,
            "leads_in_work": leads_in_work, "won": won, "conversion": conversion,
            "overdue": overdue,
        })
    return result


@api_router.get("/admin/manager_stats")
async def manager_stats(period: str = "today", current_user: dict = Depends(require_director)):
    result = await _compute_manager_stats(period)
    return {"managers": result}


@api_router.get("/leads/leaderboard")
async def leads_leaderboard(period: str = "today", current_user: dict = Depends(_require_user)):
    result = await _compute_manager_stats(period, query={"role": "manager"})
    result.sort(key=lambda r: r["calls"], reverse=True)
    return {"leaderboard": result, "period": period}


async def _my_dashboard_month(uid: str, month_str: str) -> dict:
    my, mm = int(month_str[:4]), int(month_str[5:7])
    start = f"{month_str}-01"
    end = f"{my + 1}-01-01" if mm == 12 else f"{my}-{mm + 1:02d}-01"
    calls, won = await asyncio.gather(
        db.call_logs.count_documents({"created_by": uid, "created_at": {"$gte": start, "$lt": end}}),
        db.leads.count_documents({"assigned_to": uid, "stage": "won", "won_at": {"$gte": start, "$lt": end}}),
    )
    return {"month": month_str, "calls": calls, "won": won}


@api_router.get("/leads/my_dashboard")
async def my_dashboard(current_user: dict = Depends(_require_user)):
    # Calls/leads only for now — orders aren't reliably assigned to
    # managers yet (claiming isn't in regular use), so a margin/orders
    # section here would just be empty or misleading either way.
    uid = current_user["id"]
    now = datetime.now(timezone.utc)

    months = []
    y, m = now.year, now.month
    for _ in range(6):
        months.append(f"{y}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    months.reverse()

    monthly = await asyncio.gather(*[_my_dashboard_month(uid, m) for m in months])
    monthly = list(monthly)

    best_month = max(monthly, key=lambda x: x["calls"]) if any(x["calls"] for x in monthly) else None
    total_calls = sum(x["calls"] for x in monthly)
    total_won = sum(x["won"] for x in monthly)

    return {
        "monthly": monthly,
        "best_month": best_month,
        "total_calls": total_calls,
        "total_won": total_won,
    }


# Registered here — after every specific /leads/... route above — on
# purpose. FastAPI matches routes in registration order, and this call
# defines GET /leads/{item_id} among others; if it ran earlier (it used to,
# right after /leads/calls_by_day), that catch-all shadowed any later
# single-segment /leads/<literal> route registered after it — e.g.
# /leads/leaderboard and /leads/my_dashboard both silently 404'd as "lead
# leaderboard/my_dashboard not found" instead of ever reaching their real
# handlers.
make_crud("leads", "leads", Lead, LeadUpdate, sync_to_sheets=True, user_filter=True, soft_delete=True)


@api_router.get("/admin/manager_stats/{manager_id}")
async def manager_stats_detail(manager_id: str, period: str = "week", current_user: dict = Depends(require_director)):
    start_str = _analytics_period_start(period).isoformat()
    logs = await db.call_logs.find(
        {"created_by": manager_id, "created_at": {"$gte": start_str}}, {"_id": 0}
    ).to_list(100000)
    leads = await db.leads.find(
        {"assigned_to": manager_id, "deleted": {"$ne": True}}, {"_id": 0}
    ).to_list(100000)
    return await _compute_leads_analytics(logs, leads)


async def _check_stale_managers():
    while True:
        await asyncio.sleep(3600)
        try:
            managers = await db.users.find({"role": "manager", "status": {"$ne": "suspended"}}).to_list(200)
            for m in managers:
                overdue = await db.leads.count_documents({
                    "assigned_to": m["id"], "next_call": {"$ne": None, "$lt": datetime.now(timezone.utc).isoformat()},
                    "stage": {"$nin": ["won", "lost", "no_contact"]},
                })
                if overdue >= 5:
                    await db.notifications.insert_one({
                        "id": str(uuid.uuid4()), "type": "stale_manager",
                        "message": f"{m['name']}: {overdue} просроченных перезвонов накопилось",
                        "created_at": datetime.now(timezone.utc).isoformat(), "read": False,
                    })
        except Exception as e:
            logging.getLogger(__name__).error(f"[stale_check] {e}")


@api_router.get("/notifications")
async def list_notifications(current_user: dict = Depends(require_director)):
    items = await db.notifications.find({"read": False}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return {"notifications": items}


@api_router.post("/notifications/{notif_id}/read")
async def mark_notification_read(notif_id: str, current_user: dict = Depends(require_director)):
    await db.notifications.update_one({"id": notif_id}, {"$set": {"read": True}})
    return {"ok": True}


@api_router.get("/users/{user_id}/activity")
async def get_user_activity(user_id: str, current_user: dict = Depends(require_director)):
    now_dt = datetime.now(timezone.utc)
    month_start = f"{now_dt.year}-{now_dt.month:02d}-01"
    orders_month = await db.orders.count_documents({
        "created_by": user_id,
        "created_at": {"$gte": month_start},
    })
    leads = await db.leads.find({"assigned_to": user_id}, {"_id": 0, "id": 1, "stage": 1}).to_list(10000)
    lead_ids = [l["id"] for l in leads]
    won_leads = sum(1 for l in leads if l.get("stage") == "won")
    total_leads = len(leads)
    conversion = round(won_leads / total_leads * 100) if total_leads else 0
    start_30 = (now_dt - timedelta(days=29)).strftime("%Y-%m-%d")
    if lead_ids:
        activity_docs = await db.lead_activity.find(
            {"lead_id": {"$in": lead_ids}, "date": {"$gte": start_30}}, {"_id": 0}
        ).to_list(10000)
    else:
        activity_docs = []
    calls_by_day: dict = {}
    for doc in activity_docs:
        dt = doc.get("date", "")
        if dt:
            calls_by_day[dt] = calls_by_day.get(dt, 0) + 1
    calls_total = sum(calls_by_day.values())
    activity_chart = [{"date": k, "count": v} for k, v in sorted(calls_by_day.items())]
    return {
        "orders_month": orders_month,
        "calls_total": calls_total,
        "total_leads": total_leads,
        "won_leads": won_leads,
        "conversion": conversion,
        "activity_chart": activity_chart,
    }


@api_router.get("/users/{user_id}/stats")
async def get_user_stats(user_id: str, month: Optional[str] = None, current_user: dict = Depends(require_director)):
    now_dt = datetime.now(timezone.utc)
    if month and len(month) == 7:
        y, m = int(month[:4]), int(month[5:7])
    else:
        y, m = now_dt.year, now_dt.month
    month_start = f"{y}-{m:02d}-01"
    month_end = f"{y + 1}-01-01" if m == 12 else f"{y}-{m + 1:02d}-01"
    all_user_orders = await db.orders.find(
        {"$or": [{"created_by": str(user_id)}, {"assigned_to": str(user_id)}], "deleted": {"$ne": True}},
        {"_id": 0, "client_rate": 1, "unload_date": 1, "created_at": 1},
    ).to_list(10000)
    orders_in_month = [
        o for o in all_user_orders
        if month_start <= (o.get("unload_date") or o.get("created_at", ""))[:10] < month_end
    ]
    orders_created = len(orders_in_month)
    revenue_month = sum(o.get("client_rate", 0) for o in orders_in_month)
    calls_made = await db.lead_activity.count_documents(
        {"user_id": user_id, "date": {"$gte": month_start, "$lt": month_end}}
    )
    leads = await db.leads.find({"assigned_to": user_id, "deleted": {"$ne": True}}, {"_id": 0, "stage": 1}).to_list(10000)
    won = sum(1 for l in leads if l.get("stage") == "won")
    total = len(leads)
    conversion = round(won / total * 100) if total else 0
    return {
        "orders_created": orders_created,
        "calls_made": calls_made,
        "conversion": conversion,
        "revenue_month": revenue_month,
        "total_leads": total,
        "won_leads": won,
    }


@api_router.get("/users/{user_id}/orders")
async def get_user_orders(user_id: str, month: Optional[str] = None, current_user: dict = Depends(require_director)):
    now_dt = datetime.now(timezone.utc)
    if month and len(month) == 7:
        y, m = int(month[:4]), int(month[5:7])
    else:
        y, m = now_dt.year, now_dt.month
    month_start = f"{y}-{m:02d}-01"
    month_end = f"{y + 1}-01-01" if m == 12 else f"{y}-{m + 1:02d}-01"
    all_user_orders = await db.orders.find(
        {"$or": [{"created_by": str(user_id)}, {"assigned_to": str(user_id)}], "deleted": {"$ne": True}},
        {"_id": 0, "id": 1, "order_number": 1, "client_name": 1, "route_from": 1,
         "route_to": 1, "client_rate": 1, "status": 1, "unload_date": 1, "created_at": 1},
    ).to_list(10000)
    orders = [
        o for o in all_user_orders
        if month_start <= (o.get("unload_date") or o.get("created_at", ""))[:10] < month_end
    ]
    orders.sort(key=lambda o: (o.get("unload_date") or o.get("created_at", "")), reverse=True)
    return orders


@api_router.get("/users/{user_id}/leads")
async def get_user_leads(user_id: str, current_user: dict = Depends(require_director)):
    leads = await db.leads.find(
        {"assigned_to": user_id, "deleted": {"$ne": True}},
        {"_id": 0, "id": 1, "name": 1, "company": 1, "stage": 1, "last_contact": 1, "phone": 1},
    ).sort("last_contact", -1).to_list(10000)
    return leads


# ====== Documents generation ======
@api_router.post("/orders/{order_id}/docs/{kind}")
async def generate_order_doc(order_id: str, kind: str, regenerate: bool = False):
    """Сгенерировать (или перегенерировать) документ для заявки.
    kind: 'client' | 'carrier' | 'act'
    Возвращает {url, kind}.
    """
    if kind not in ("client", "carrier", "act"):
        raise HTTPException(400, "kind должен быть client | carrier | act")
    if _docs_get_generator is None:
        raise HTTPException(500, "Docs generation недоступна")

    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")

    # Если уже есть и не просим перегенерировать — возвращаем существующий
    field = _docs_kind_to_field(kind)
    existing = order.get(field) or ""
    if existing and not regenerate:
        return {"url": existing, "kind": kind, "cached": True}

    # Подгружаем реквизиты клиента/перевозчика
    client_doc = None
    carrier_doc = None
    if order.get("client_id"):
        client_doc = await db.clients.find_one({"id": order["client_id"]}, {"_id": 0})
    if not client_doc and order.get("client_name"):
        client_doc = await db.clients.find_one({"name": order["client_name"]}, {"_id": 0})
    if order.get("carrier_id"):
        carrier_doc = await db.carriers.find_one({"id": order["carrier_id"]}, {"_id": 0})
    if not carrier_doc and order.get("carrier_name"):
        carrier_doc = await db.carriers.find_one({"company_name": order["carrier_name"]}, {"_id": 0})

    # Токен читаем здесь, на основном event loop, и передаём его в gen.generate
    # параметром — оно выполняется внутри asyncio.to_thread, и там нельзя
    # повторно await'ить Mongo (db привязан к основному loop, а не к тому,
    # что был бы создан внутри потока).
    token_doc = await db.oauth_tokens.find_one({"_id": "google"}, {"_id": 0})

    try:
        gen = _docs_get_generator()
        url = await asyncio.to_thread(gen.generate, kind, order, client_doc, carrier_doc, token_doc)
    except Exception as e:
        # Сообщение пользователю
        msg = str(e) or repr(e)
        try:
            if hasattr(e, 'reason'):
                msg = f"{msg} ({e.reason})"
        except Exception:
            pass
        logging.getLogger(__name__).error(f"docs generate failed: {e}", exc_info=True)
        raise HTTPException(500, f"Не удалось создать документ: {msg}")

    # Сохраняем URL в Mongo + пушим в таблицу
    await db.orders.update_one({"id": order_id}, {"$set": {field: url}})
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if updated and _sw_push_order:
        try:
            await _sw_push_order(updated)
        except Exception:
            pass
    return {"url": url, "kind": kind, "cached": False}


# ====== Google Sheets sync endpoints ======
@api_router.post("/sync/sheets")
async def sync_sheets_now():
    """ОТКЛЮЧЕНО. Выгрузка из CRM в Google Sheets отключена,
    чтобы случайно не перезаписать данные пользователя.
    Источник истины — Google Таблица. Импорт идёт только в одну сторону: Sheets → CRM.
    """
    raise HTTPException(
        409,
        "Выгрузка из CRM в Google Sheets отключена. Используйте импорт из таблицы (POST /api/sync/import_from_sheets).",
    )


@api_router.get("/sync/sheets/status")
async def sync_sheets_status():
    """Статус последней синхронизации."""
    if _get_sheets_sync is None:
        return {"ok": False, "message": "Sheets sync not available"}
    s = _get_sheets_sync()
    return s.last_status


@api_router.get("/sync/import_preview")
async def import_preview(
    collections: str = Query(""),
    current_user: dict = Depends(require_director),
):
    """Что ИЗМЕНИТСЯ при импорте из Google Таблицы — без записи. Директор
    смотрит список (изменённые поля заявок, новые заявки, кандидаты на
    авто-удаление) перед тем, как подтвердить массовую перезапись."""
    if _sheets_preview_import is None:
        raise HTTPException(500, "Импорт недоступен")
    targets = [c.strip() for c in collections.split(",") if c.strip()] or None
    return await _sheets_preview_import(db, targets=targets)


@api_router.post("/sync/import_from_sheets")
async def import_from_sheets(
    collections: str = Query(""),
    mode: str = Query("merge"),
    current_user: dict = Depends(require_director),
):
    """Импорт клиентов / перевозчиков / заказов / лидов из Google Таблицы в CRM.
    ОДНОСТОРОННЕ: только Sheets -> CRM. Таблица не изменяется.
    mode='merge' (по умолчанию) — не затирает уже заполненные в CRM поля.
    collections — через запятую (clients,carriers,orders,leads), пусто = все.
    Директор-only и требует предварительного просмотра через
    /sync/import_preview — эта массовая перезапись раньше срабатывала молча
    в фоне и несколько раз портила данные без возможности проверить заранее.
    """
    if _sheets_run_import is None:
        raise HTTPException(500, "Импорт недоступен")
    targets = [c.strip() for c in collections.split(",") if c.strip()] or None
    return await _sheets_run_import(db, targets=targets, mode=mode)


@api_router.get("/sync/import_status")
async def import_status():
    if _sheets_import_status is None:
        return {"ok": False, "message": "Импорт недоступен", "imported": {}}
    return _sheets_import_status()


# ====== Dashboard ======
def order_in_period(o: dict, period: str) -> bool:
    if period == "all":
        return True
    # ВАЖНО: считаем по дате ВЫГРУЗКИ (когда заказ фактически закрылся в этом месяце)
    ud = o.get("unload_date") or ""
    if not ud:
        # если нет даты выгрузки — пробуем дату загрузки, иначе created_at
        ud = o.get("load_date") or (o.get("created_at", "")[:10])
    return ud.startswith(period)


@api_router.get("/dashboard/day_orders")
async def dashboard_day_orders(date: str):
    all_orders = await db.orders.find({"deleted": {"$ne": True}}, {"_id": 0}).to_list(5000)
    day_list = []
    total_margin = 0.0
    for o in all_orders:
        order_date = (o.get("unload_date") or o.get("load_date") or "")[:10]
        if order_date == date:
            margin = o.get("client_rate", 0) - o.get("carrier_rate", 0)
            day_list.append({
                "id": str(o.get("id") or o.get("_id", "")),
                "order_number": o.get("order_number", "—"),
                "client_name": o.get("client_name", "—"),
                "margin": margin,
                "client_rate": o.get("client_rate", 0),
            })
            total_margin += margin
    return {
        "date": date,
        "orders": sorted(day_list, key=lambda x: x["margin"], reverse=True),
        "total_margin": total_margin,
        "orders_count": len(day_list),
    }


_DASHBOARD_PROJECTION = {
    "_id": 0, "client_rate": 1, "carrier_rate": 1, "client_paid": 1, "carrier_paid": 1,
    "client_name": 1, "carrier_name": 1, "status": 1, "unload_date": 1, "load_date": 1,
    "created_at": 1,
}


@api_router.get("/dashboard")
async def dashboard(period: str = "all"):
    # Only the handful of fields this endpoint actually reads — the full
    # order document (route text, doc URLs, payment history, ...) was being
    # pulled over the wire for every one of up to 5000 orders on every load.
    all_orders = await db.orders.find({"deleted": {"$ne": True}}, _DASHBOARD_PROJECTION).to_list(5000)
    orders = [o for o in all_orders if order_in_period(o, period)]

    carriers_count = await db.carriers.count_documents({"deleted": {"$ne": True}})
    leads_count = await db.leads.count_documents({"deleted": {"$ne": True}})

    # clients_count: new clients in the selected period; total for "all"
    import re as _re_dash
    if period == "all" or not _re_dash.match(r"^\d{4}-\d{2}$", period):
        clients_count = await db.clients.count_documents({"deleted": {"$ne": True}})
    else:
        _y, _m = int(period[:4]), int(period[5:7])
        _p_start = f"{_y}-{_m:02d}-01"
        _p_end = f"{_y + 1}-01-01" if _m == 12 else f"{_y}-{_m + 1:02d}-01"
        clients_count = await db.clients.count_documents({
            "created_at": {"$gte": _p_start, "$lt": _p_end},
            "deleted": {"$ne": True},
        })

    total_revenue = sum(o.get("client_rate", 0) for o in orders)
    total_cost = sum(o.get("carrier_rate", 0) for o in orders)
    total_margin = total_revenue - total_cost
    profit = total_margin * (1 - TAX_RATE)  # после 20% налога

    delivered = [o for o in orders if o.get("status") == "delivered"]
    active = [o for o in orders if o.get("status") in ("new", "in_progress")]

    unpaid_by_clients = sum(o.get("client_rate", 0) for o in orders if not o.get("client_paid", False))
    owed_to_carriers = sum(o.get("carrier_rate", 0) for o in orders if not o.get("carrier_paid", False))

    # === Должники (клиенты, которые нам должны) ===
    debtors_map: dict = {}
    for o in orders:
        if o.get("client_paid"):
            continue
        name = o.get("client_name") or "—"
        if name not in debtors_map:
            debtors_map[name] = {"name": name, "amount": 0, "orders": 0}
        debtors_map[name]["amount"] += o.get("client_rate", 0)
        debtors_map[name]["orders"] += 1
    debtors = sorted(debtors_map.values(), key=lambda x: x["amount"], reverse=True)

    # === Те, кому мы должны (перевозчики) ===
    creditors_map: dict = {}
    for o in orders:
        if o.get("carrier_paid"):
            continue
        name = o.get("carrier_name") or "—"
        if name not in creditors_map:
            creditors_map[name] = {"name": name, "amount": 0, "orders": 0}
        creditors_map[name]["amount"] += o.get("carrier_rate", 0)
        creditors_map[name]["orders"] += 1
    creditors = sorted(creditors_map.values(), key=lambda x: x["amount"], reverse=True)

    client_totals: dict = {}
    for o in orders:
        n = o.get("client_name") or "—"
        client_totals[n] = client_totals.get(n, 0) + o.get("client_rate", 0)
    top_clients = sorted([{"name": k, "revenue": v} for k, v in client_totals.items()],
                         key=lambda x: x["revenue"], reverse=True)

    client_margin_map: dict = {}
    for o in orders:
        n = o.get("client_name") or "—"
        if n not in client_margin_map:
            client_margin_map[n] = {"name": n, "orders_count": 0, "revenue": 0.0, "cost": 0.0}
        client_margin_map[n]["orders_count"] += 1
        client_margin_map[n]["revenue"] += o.get("client_rate", 0)
        client_margin_map[n]["cost"] += o.get("carrier_rate", 0)
    top_clients_margin = []
    for item in client_margin_map.values():
        margin = item["revenue"] - item["cost"]
        pct = round(margin / item["revenue"] * 100, 1) if item["revenue"] else 0
        top_clients_margin.append({**item, "margin": margin, "margin_percent": pct})
    top_clients_margin = sorted(top_clients_margin, key=lambda x: x["margin_percent"], reverse=True)

    status_breakdown = {
        "new": len([o for o in orders if o.get("status") == "new"]),
        "in_progress": len([o for o in orders if o.get("status") == "in_progress"]),
        "delivered": len(delivered),
        "cancelled": len([o for o in orders if o.get("status") == "cancelled"]),
    }

    # === Сравнение с предыдущим месяцем (только если period — конкретный месяц) ===
    prev_margin = None
    prev_revenue = None
    prev_period = None
    margin_change_pct = None
    import re as _re2
    if _re2.match(r"^\d{4}-\d{2}$", period):
        y, m = int(period[:4]), int(period[5:7])
        py, pm = (y - 1, 12) if m == 1 else (y, m - 1)
        prev_period = f"{py:04d}-{pm:02d}"
        prev_orders = [o for o in all_orders if order_in_period(o, prev_period)]
        prev_revenue = sum(o.get("client_rate", 0) for o in prev_orders)
        prev_cost = sum(o.get("carrier_rate", 0) for o in prev_orders)
        prev_margin = prev_revenue - prev_cost
        if prev_margin and abs(prev_margin) > 0.01:
            margin_change_pct = round((total_margin - prev_margin) / abs(prev_margin) * 100, 1)
        elif total_margin:
            margin_change_pct = None  # не было данных за прошлый месяц — без сравнения

    # Только валидные YYYY-MM (по дате выгрузки)
    import re as _re
    valid_ym = _re.compile(r"^\d{4}-\d{2}$")
    months = sorted({(o.get("unload_date") or o.get("load_date") or "")[:7]
                     for o in all_orders
                     if (o.get("unload_date") or o.get("load_date"))
                     and valid_ym.match((o.get("unload_date") or o.get("load_date") or "")[:7])},
                    reverse=True)[:24]

    return {
        "period": period,
        "available_months": months,
        "total_revenue": total_revenue,  # для совместимости
        "total_cost": total_cost,
        "total_margin": total_margin,
        "profit": profit,
        "tax_rate": TAX_RATE,
        "prev_period": prev_period,
        "prev_margin": prev_margin,
        "prev_revenue": prev_revenue,
        "margin_change_pct": margin_change_pct,
        "margin_percent": round((total_margin / total_revenue * 100), 1) if total_revenue else 0,
        "active_orders": len(active),
        "delivered_orders": len(delivered),
        "total_orders": len(orders),
        "unpaid_by_clients": unpaid_by_clients,
        "owed_to_carriers": owed_to_carriers,
        "clients_count": clients_count,
        "carriers_count": carriers_count,
        "leads_count": leads_count,
        "top_clients": top_clients,
        "top_clients_margin": top_clients_margin,
        "debtors": debtors,
        "creditors": creditors,
        "status_breakdown": status_breakdown,
        "chart_orders": [
            {"d": o.get("unload_date") or o.get("load_date"), "cr": o.get("client_rate", 0), "car": o.get("carrier_rate", 0)}
            for o in all_orders
            if o.get("unload_date") or o.get("load_date")
        ],
    }


# ====== Monthly Goals ======
class GoalsPayload(BaseModel):
    month: str
    profit_goal: float = 7000
    trips_goal: int = 45
    margin_goal: float = 230
    new_clients_goal: int = 9


@api_router.get("/goals")
async def get_goals(month: str, current_user: dict = Depends(_require_user)):
    try:
        y, m = int(month[:4]), int(month[5:7])
    except Exception:
        raise HTTPException(400, "Invalid month format, expected YYYY-MM")

    month_start_str = f"{y}-{m:02d}-01"
    month_end_str = f"{y + 1}-01-01" if m == 12 else f"{y}-{m + 1:02d}-01"

    doc = await db.monthly_goals.find_one({"month": month}, {"_id": 0})
    if not doc:
        py, pm = (y - 1, 12) if m == 1 else (y, m - 1)
        prev_doc = await db.monthly_goals.find_one({"month": f"{py:04d}-{pm:02d}"}, {"_id": 0})
        if prev_doc:
            doc = {**prev_doc, "month": month}
        else:
            doc = {"month": month, "profit_goal": 7000, "trips_goal": 45, "margin_goal": 230, "new_clients_goal": 9}

    orders = await db.orders.find({
        "unload_date": {"$gte": month_start_str, "$lt": month_end_str},
        "status": {"$ne": "cancelled"},
        "deleted": {"$ne": True},
    }).to_list(10000)

    profit_fact = sum((o.get("client_rate", 0) - o.get("carrier_rate", 0)) * 0.8 for o in orders)
    trips_fact = len(orders)
    margin_fact = round(profit_fact / trips_fact, 1) if trips_fact > 0 else 0

    new_clients_fact = await db.clients.count_documents({
        "created_at": {"$gte": month_start_str, "$lt": month_end_str},
        "deleted": {"$ne": True},
    })

    return {
        **doc,
        "profit_fact": round(profit_fact, 1),
        "trips_fact": trips_fact,
        "margin_fact": margin_fact,
        "new_clients_fact": new_clients_fact,
    }


@api_router.post("/goals")
async def save_goals(payload: GoalsPayload, current_user: dict = Depends(_require_user)):
    doc = payload.dict()
    await db.monthly_goals.replace_one({"month": payload.month}, doc, upsert=True)
    return doc


# ====== App Settings ======
_DEFAULT_RATE_MINIMUMS = [
    {"route": "Москва↔Минск",       "cost": 720,  "current": 50,  "threshold": 865,  "status": "red"},
    {"route": "СПб↔Минск",          "cost": 406,  "current": 75,  "threshold": 487,  "status": "red"},
    {"route": "Минск↔Москва",       "cost": 492,  "current": 150, "threshold": 590,  "status": "red"},
    {"route": "Жодино↔Москва",      "cost": 543,  "current": 250, "threshold": 651,  "status": "yellow"},
    {"route": "Ульяновск–Минск",    "cost": 1328, "current": 900, "threshold": 1594, "status": "yellow"},
    {"route": "Минск–Минск (лок.)", "cost": 130,  "current": 0,   "threshold": 200,  "status": "red"},
    {"route": "Смоленск–Минск",     "cost": 335,  "current": 160, "threshold": 402,  "status": "yellow"},
]

_DEFAULT_GOALS = {
    "margin_goal": 10000,
    "new_clients_goal": 9,
    "trips_goal": 45,
    "avg_margin_goal": 230,
}


@api_router.get("/settings")
async def get_settings(current_user: dict = Depends(_require_user)):
    doc = await db.app_settings.find_one({"_id": "main"})
    if not doc:
        return {**_DEFAULT_GOALS, "rate_minimums": _DEFAULT_RATE_MINIMUMS}
    return {
        "margin_goal": doc.get("margin_goal", _DEFAULT_GOALS["margin_goal"]),
        "new_clients_goal": doc.get("new_clients_goal", _DEFAULT_GOALS["new_clients_goal"]),
        "trips_goal": doc.get("trips_goal", _DEFAULT_GOALS["trips_goal"]),
        "avg_margin_goal": doc.get("avg_margin_goal", _DEFAULT_GOALS["avg_margin_goal"]),
        "rate_minimums": doc.get("rate_minimums", _DEFAULT_RATE_MINIMUMS),
    }


@api_router.put("/settings")
async def update_settings(request: Request, current_user: dict = Depends(_require_user)):
    payload = await request.json()
    await db.app_settings.update_one({"_id": "main"}, {"$set": payload}, upsert=True)
    return {"ok": True}


# ====== Analytics ======
MONTHS_RU = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн", "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]


@api_router.get("/analytics")
async def analytics(month: Optional[str] = None, current_user: dict = Depends(_require_user)):
    all_orders = await db.orders.find({"deleted": {"$ne": True}}, {"_id": 0}).to_list(5000)

    period = month or "all"
    orders = [o for o in all_orders if order_in_period(o, period)]

    def _margin(o: dict) -> float:
        return o.get("client_rate", 0) - o.get("carrier_rate", 0)

    revenue = sum(o.get("client_rate", 0) for o in orders)
    expenses = sum(o.get("carrier_rate", 0) for o in orders)
    margin = revenue - expenses
    margin_pct = round(margin / revenue * 100, 1) if revenue else 0
    trips_count = len(orders)
    avg_margin = round(margin / trips_count, 1) if trips_count else 0

    # Goals — для выбранного периода; если "all" — используем текущий месяц
    import re as _re_analytics
    cur_ym = datetime.now(timezone.utc).strftime("%Y-%m")
    goal_ym = period if _re_analytics.match(r"^\d{4}-\d{2}$", period) else cur_ym
    cur_orders = [o for o in all_orders if order_in_period(o, goal_ym)]
    margin_fact = sum(_margin(o) for o in cur_orders)
    trips_fact = len(cur_orders)
    avg_margin_fact = round(margin_fact / trips_fact, 1) if trips_fact else 0

    # new_clients_fact: count by created_at in clients collection, not order-derived
    _g_y, _g_m = int(goal_ym[:4]), int(goal_ym[5:7])
    _g_start = f"{_g_y}-{_g_m:02d}-01"
    _g_end = f"{_g_y + 1}-01-01" if _g_m == 12 else f"{_g_y}-{_g_m + 1:02d}-01"
    new_clients_fact = await db.clients.count_documents({
        "created_at": {"$gte": _g_start, "$lt": _g_end},
        "deleted": {"$ne": True},
    })

    # Clients top-10
    client_map: dict = {}
    for o in orders:
        n = o.get("client_name") or "—"
        if n not in client_map:
            client_map[n] = {"name": n, "trips": 0, "revenue": 0.0, "margin": 0.0}
        client_map[n]["trips"] += 1
        client_map[n]["revenue"] += o.get("client_rate", 0)
        client_map[n]["margin"] += _margin(o)
    clients_list = []
    for it in client_map.values():
        mp = round(it["margin"] / it["revenue"] * 100, 1) if it["revenue"] else 0
        clients_list.append({**it, "margin_pct": mp})
    clients_list = sorted(clients_list, key=lambda x: x["margin"], reverse=True)[:10]

    # Routes top-8
    route_map: dict = {}
    for o in orders:
        rf = (o.get("route_from") or "").strip()
        rt = (o.get("route_to") or "").strip()
        route = f"{rf}–{rt}" if rf and rt else rf or rt or "—"
        if route not in route_map:
            route_map[route] = {"route": route, "trips": 0, "total_margin": 0.0}
        route_map[route]["trips"] += 1
        route_map[route]["total_margin"] += _margin(o)
    routes_list = []
    for it in route_map.values():
        avg_m = round(it["total_margin"] / it["trips"], 1) if it["trips"] else 0
        routes_list.append({**it, "avg_margin": avg_m})
    routes_list = sorted(routes_list, key=lambda x: x["total_margin"], reverse=True)[:8]

    # Loss trips (маржа < 0)
    loss_trips = []
    for o in orders:
        m = _margin(o)
        if m < 0:
            rf = (o.get("route_from") or "").strip()
            rt = (o.get("route_to") or "").strip()
            route = f"{rf}–{rt}" if rf and rt else rf or rt or "—"
            loss_trips.append({
                "number": o.get("order_number", ""),
                "client": o.get("client_name") or "—",
                "route": route,
                "client_rate": o.get("client_rate", 0),
                "carrier_rate": o.get("carrier_rate", 0),
                "loss": m,
            })
    loss_trips = sorted(loss_trips, key=lambda x: x["loss"])

    # Monthly breakdown (по всем заявкам, не по периоду)
    monthly_map: dict = {}
    for o in all_orders:
        ym = (o.get("unload_date") or o.get("load_date") or o.get("created_at", ""))[:7]
        if not ym or len(ym) < 7:
            continue
        if ym not in monthly_map:
            monthly_map[ym] = {"month": ym, "revenue": 0.0, "expenses": 0.0, "margin": 0.0, "trips": 0}
        monthly_map[ym]["revenue"] += o.get("client_rate", 0)
        monthly_map[ym]["expenses"] += o.get("carrier_rate", 0)
        monthly_map[ym]["margin"] += _margin(o)
        monthly_map[ym]["trips"] += 1
    monthly_list = []
    for it in sorted(monthly_map.values(), key=lambda x: x["month"]):
        try:
            m_idx = int(it["month"][5:7]) - 1
            label = MONTHS_RU[m_idx]
        except Exception:
            label = it["month"]
        monthly_list.append({**it, "label": label})

    return {
        "period": period,
        "summary": {
            "revenue": revenue, "expenses": expenses, "margin": margin,
            "margin_pct": margin_pct, "trips_count": trips_count, "avg_margin": avg_margin,
        },
        "goals": {
            "margin_goal": 10000, "margin_fact": round(margin_fact, 1),
            "new_clients_goal": 9, "new_clients_fact": new_clients_fact,
            "trips_goal": 45, "trips_fact": trips_fact,
            "avg_margin_goal": 230, "avg_margin_fact": avg_margin_fact,
        },
        "monthly": monthly_list,
        "clients": clients_list,
        "routes": routes_list,
        "loss_trips": loss_trips,
    }


# ====== Global Search (Cmd+K palette) ======
@api_router.get("/search")
async def global_search(q: str = "", current_user: Optional[dict] = Depends(_get_user_from_token)):
    if not q or len(q.strip()) < 2:
        return {"results": []}
    rx = {"$regex": q.strip(), "$options": "i"}

    orders = await db.orders.find({
        "deleted": {"$ne": True},
        "$or": [{"order_number": rx}, {"client_name": rx}, {"carrier_name": rx}, {"route_from": rx}, {"route_to": rx}],
    }, {"_id": 0}).limit(8).to_list(8)

    clients = await db.clients.find({"deleted": {"$ne": True}, "name": rx}, {"_id": 0}).limit(8).to_list(8)
    carriers = await db.carriers.find({"deleted": {"$ne": True}, "company_name": rx}, {"_id": 0}).limit(8).to_list(8)
    leads = await db.leads.find({"deleted": {"$ne": True}, "$or": [{"name": rx}, {"phone": rx}]}, {"_id": 0}).limit(8).to_list(8)
    tasks = await db.tasks.find({"title": rx, "status": {"$ne": "done"}}, {"_id": 0}).limit(8).to_list(8)

    results = (
        [{"type": "order", "id": o["id"], "title": o.get("order_number", ""), "subtitle": f"{o.get('client_name','')} · {o.get('route_from','')} → {o.get('route_to','')}"} for o in orders] +
        [{"type": "client", "id": c["id"], "title": c.get("name", ""), "subtitle": c.get("phone", "") or "Клиент"} for c in clients] +
        [{"type": "carrier", "id": c["id"], "title": c.get("company_name", ""), "subtitle": c.get("phone", "") or "Перевозчик"} for c in carriers] +
        [{"type": "lead", "id": l["id"], "title": l.get("name", ""), "subtitle": l.get("phone", "") or "Лид"} for l in leads] +
        [{"type": "task", "id": t["id"], "title": t.get("title", ""), "subtitle": "Задача"} for t in tasks]
    )
    return {"results": results}


# ====== Seed ======
@api_router.post("/seed")
async def seed_data():
    await db.orders.delete_many({})
    await db.clients.delete_many({})
    await db.carriers.delete_many({})
    await db.leads.delete_many({})

    clients = [
        Client(name="ООО Логистик-Прайм", contact_person="Иванов И.И.", phone="+7 (495) 111-22-33", email="info@logistic-prime.ru",
               inn="7701234567", kpp="770101001", legal_address="г. Москва, ул. Тверская, д. 10",
               bank_name="ПАО Сбербанк", bank_account="40702810400000012345", bank_bik="044525225", bank_corr_account="30101810400000000225",
               payment_terms="10 дней", cargo_types="Электроника, бытовая техника", directions="МСК-СПб, МСК-НСК"),
        Client(name="ТД Северный Ветер", contact_person="Петрова О.С.", phone="+7 (812) 222-33-44", email="op@nordwind.ru",
               inn="7802345678", kpp="780201001", legal_address="г. СПб, Невский пр., д. 22",
               bank_name="АО Альфа-Банк", bank_account="40702810500000098765", bank_bik="044030786", bank_corr_account="30101810200000000593",
               payment_terms="7 дней", cargo_types="Продукты питания (реф)", directions="СПб-Урал"),
        Client(name="АО МеталлТорг", contact_person="Сидоров А.В.", phone="+7 (343) 333-44-55", email="logistics@metalltorg.ru",
               inn="6603456789", kpp="660301001", legal_address="г. Екатеринбург, ул. Ленина, д. 50",
               bank_name="ВТБ", bank_account="40702810700000054321", bank_bik="046577751", bank_corr_account="30101810700000000751",
               payment_terms="14 дней", cargo_types="Металлопрокат, арматура", directions="Урал-юг РФ"),
        Client(name="ИП Смирнов В.П.", contact_person="Смирнов В.П.", phone="+7 (903) 444-55-66", email="smirnov@mail.ru",
               inn="770345678901", payment_terms="3 дня", cargo_types="Хим. продукция в IBC"),
        Client(name="ООО Гранд Пром", contact_person="Кузнецова Е.Н.", phone="+7 (495) 555-66-77", email="grand@grandprom.ru",
               inn="7704567890", kpp="770401001", bank_name="Тинькофф", bank_account="40702810900000067890", bank_bik="044525974",
               payment_terms="5 дней", cargo_types="Запчасти", directions="ЦФО"),
    ]
    for c in clients: await db.clients.insert_one(c.dict())

    carriers = [
        Carrier(company_name="ИП Морозов А.Н.", driver_name="Морозов Андрей Николаевич", phone="+7 (905) 100-20-30",
                inn="503012345678", legal_address="МО, г. Подольск, ул. Кирова, д. 5",
                bank_name="Сбербанк", bank_account="40802810400001234567", bank_bik="044525225",
                vehicle_type="Тент", plate="А123БВ77", capacity_tons=20, capacity_m3=86, rating=4.8,
                cargo_types="Стандартные грузы", regions="ЦФО, СЗФО"),
        Carrier(company_name="ООО ТрансЛайн", driver_name="Волков Сергей Иванович", phone="+7 (916) 200-30-40",
                inn="7706123456", kpp="770601001", bank_name="ВТБ", bank_account="40702810800009876543", bank_bik="046577751",
                vehicle_type="Реф", plate="М456КН99", capacity_tons=20, capacity_m3=82, rating=4.9,
                cargo_types="Продукты, медикаменты", regions="Россия + Беларусь"),
        Carrier(company_name="ИП Гусев Д.Л.", driver_name="Гусев Дмитрий", phone="+7 (921) 300-40-50",
                inn="780234567890", vehicle_type="Изотерм", plate="К789ЕР78", capacity_tons=10, capacity_m3=45, rating=4.5,
                regions="СЗФО"),
        Carrier(company_name="АвтоПарк-Юг", driver_name="Романов Виктор", phone="+7 (988) 400-50-60",
                inn="2308765432", bank_name="Сбербанк", bank_account="40702810400005554433", bank_bik="040349602",
                vehicle_type="Тент", plate="Е321ОТ23", capacity_tons=20, capacity_m3=90, rating=4.7,
                regions="ЮФО, СКФО"),
        Carrier(company_name="ИП Беляев К.С.", driver_name="Беляев Константин", phone="+7 (962) 500-60-70",
                inn="660345678123", vehicle_type="Тент", plate="У654АХ66", capacity_tons=5, capacity_m3=24, rating=4.3,
                regions="УФО"),
    ]
    for c in carriers: await db.carriers.insert_one(c.dict())

    # Несколько заявок старше 15 дней с неоплатами для подсветки
    orders_data = [
        # старая неоплаченная (должна подсветиться красным)
        ("№2025-0120", clients[0], carriers[0], "Москва", "СПб", "г. Москва, Каширское ш., 23", "г. СПб, ул. Софийская, 14А", "2026-01-15", "2026-01-16", 95000, 75000, "delivered", False, False, True, False, False, False, "Электроника", 12),
        ("№2025-0118", clients[2], carriers[3], "Екатеринбург", "Краснодар", "г. Екатеринбург, ул. Шефская, 2В", "г. Краснодар, ул. Дзержинского, 100", "2026-01-10", "2026-01-13", 220000, 175000, "delivered", False, True, True, True, True, False, "Металлопрокат", 20),
        # свежие
        ("№2025-0142", clients[0], carriers[0], "Москва", "Санкт-Петербург", "г. Москва, Каширское ш., 23, склад №4", "г. СПб, ул. Софийская, 14А", "2026-02-12", "2026-02-13", 95000, 75000, "in_progress", True, False, True, True, False, False, "Электроника", 12),
        ("№2025-0141", clients[1], carriers[1], "Санкт-Петербург", "Екатеринбург", "г. СПб, пр. Обуховской Обороны, 271", "г. Екатеринбург, ул. Машинная, 31", "2026-02-10", "2026-02-12", 180000, 145000, "delivered", True, True, True, True, True, True, "Продукты питания (реф)", 18),
        ("№2025-0139", clients[3], carriers[2], "Москва", "Казань", "МО, Дмитровский р-н, склад", "г. Казань, ул. Гаврилова, 5", "2026-02-14", "2026-02-15", 65000, 48000, "new", False, False, False, False, False, False, "Хим. продукция", 8),
        ("№2025-0138", clients[4], carriers[4], "Нижний Новгород", "Москва", "г. Н.Новгород, ул. Кузбасская, 1", "г. Москва, МКАД 41 км", "2026-02-09", "2026-02-09", 38000, 28000, "delivered", True, True, True, True, True, True, "Запчасти", 4),
        ("№2025-0137", clients[0], carriers[1], "Москва", "Новосибирск", "г. Москва, Каширское ш., 23", "г. Новосибирск, ул. Петухова, 17", "2026-02-15", "2026-02-19", 320000, 260000, "in_progress", False, False, True, False, False, False, "Бытовая техника", 19),
        ("№2025-0136", clients[2], carriers[0], "Челябинск", "Москва", "г. Челябинск, ул. Производственная, 8Б", "г. Москва, склад МКАД 23", "2026-02-11", "2026-02-13", 145000, 115000, "in_progress", True, False, True, True, False, False, "Стройматериалы", 20),
    ]
    for od in orders_data:
        (num, cl, cr, rf, rt, raf, rat, ld, ud, c_rate, cr_rate, status, cp, crp, dtcs, dfcr, dtcrs, dfcrr, cargo, w) = od
        await db.orders.insert_one(Order(
            order_number=num, client_id=cl.id, client_name=cl.name,
            carrier_id=cr.id, carrier_name=cr.company_name,
            route_from=rf, route_to=rt, route_from_address=raf, route_to_address=rat,
            load_date=ld, unload_date=ud,
            driver_name=cr.driver_name, driver_phone=cr.phone,
            vehicle_type=cr.vehicle_type, vehicle_plate=cr.plate,
            client_rate=c_rate, carrier_rate=cr_rate,
            status=status, client_paid=cp, carrier_paid=crp,
            docs_to_client_sent=dtcs, docs_from_client_received=dfcr,
            docs_to_carrier_sent=dtcrs, docs_from_carrier_received=dfcrr,
            cargo=cargo, weight_tons=w,
        ).dict())

    leads = [
        Lead(name="Орлов Михаил", company="ООО Восток-Логистик", phone="+7 (495) 700-10-20", city="Москва", stage="new", next_call="2026-02-12", notes="Интерес — еженедельные перевозки МСК-СПб"),
        Lead(name="Захарова Анна", company="ТД Полюс", phone="+7 (812) 700-20-30", city="СПб", stage="kp_sent", last_contact="2026-02-08", next_call="2026-02-13", notes="Просили КП на реф направление"),
        Lead(name="Николаев Пётр", company="АО Стройка-Сервис", phone="+7 (343) 700-30-40", city="Екатеринбург", stage="thinking", last_contact="2026-02-09", next_call="2026-02-11", notes="Готовы к тестовому рейсу"),
        Lead(name="Григорьев Олег", company="ИП Григорьев", phone="+7 (961) 700-40-50", city="Краснодар", stage="won", last_contact="2026-02-07", notes="Стал клиентом"),
        Lead(name="Соколова Мария", company="ООО АгроТранс", phone="+7 (902) 700-50-60", city="Воронеж", stage="new", next_call="2026-02-14", notes="Сезон — март-октябрь"),
    ]
    for l in leads: await db.leads.insert_one(l.dict())

    return {"ok": True, "clients": len(clients), "carriers": len(carriers), "orders": len(orders_data), "leads": len(leads)}


# ====== Finance: Withdrawals ======
class FinanceWithdrawal(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    amount: float
    date: str
    note: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class FinanceWithdrawalPayload(BaseModel):
    amount: float
    date: str
    note: Optional[str] = ""


@api_router.get("/finance/withdrawals", response_model=List[FinanceWithdrawal])
async def list_withdrawals():
    docs = await db.finance_withdrawals.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
    return [FinanceWithdrawal(**d) for d in docs]


@api_router.post("/finance/withdrawals", response_model=FinanceWithdrawal)
async def create_withdrawal(payload: FinanceWithdrawalPayload):
    obj = FinanceWithdrawal(**payload.dict())
    await db.finance_withdrawals.insert_one(obj.dict())
    return obj


@api_router.delete("/finance/withdrawals/{wid}")
async def delete_withdrawal(wid: str):
    await db.finance_withdrawals.delete_one({"id": wid})
    return {"ok": True}


# ====== Finance: Transactions ======
class FinanceTransaction(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type: str  # "income" | "expense"
    amount: float
    date: str
    description: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class FinanceTransactionPayload(BaseModel):
    type: str
    amount: float
    date: str
    description: Optional[str] = ""


@api_router.get("/finance/transactions", response_model=List[FinanceTransaction])
async def list_finance_transactions():
    docs = await db.finance_transactions.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
    return [FinanceTransaction(**d) for d in docs]


@api_router.post("/finance/transactions", response_model=FinanceTransaction)
async def create_finance_transaction(payload: FinanceTransactionPayload):
    obj = FinanceTransaction(**payload.dict())
    await db.finance_transactions.insert_one(obj.dict())
    return obj


@api_router.delete("/finance/transactions/{tid}")
async def delete_finance_transaction(tid: str):
    await db.finance_transactions.delete_one({"id": tid})
    return {"ok": True}


# ====== Payments In (поступления от клиентов) ======
class PaymentIn(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    pp_number: str
    date: str
    amount: float
    client_id: Optional[str] = ""
    client_name: Optional[str] = ""
    order_id: Optional[str] = ""
    order_number: Optional[str] = ""
    notes: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class PaymentInPayload(BaseModel):
    pp_number: str
    date: str
    amount: float
    client_id: Optional[str] = ""
    client_name: Optional[str] = ""
    order_id: Optional[str] = ""
    order_number: Optional[str] = ""
    notes: Optional[str] = ""


@api_router.get("/payments/in", response_model=List[PaymentIn])
async def list_payments_in(client_id: Optional[str] = None, month: Optional[str] = None):
    q: dict = {}
    if client_id:
        client_doc = await db.clients.find_one({"id": client_id})
        if client_doc:
            client_name = client_doc.get("name", "")
            q["$or"] = [{"client_id": client_id}, {"client_name": client_name}]
            print(f"[payments/in] filter by id={client_id!r} OR name={client_name!r}")
        else:
            q["client_id"] = client_id
            print(f"[payments/in] client not found, filter by id={client_id!r}")
    if month:
        q["date"] = {"$regex": f"^{month}"}
    print(f"[payments/in] query={q}")
    docs = await db.payments_in.find(q, {"_id": 0}).sort("date", -1).to_list(5000)
    print(f"[payments/in] found {len(docs)} docs, sample client_ids={[d.get('client_id') for d in docs[:5]]}")
    return [PaymentIn(**d) for d in docs]


@api_router.post("/payments/in", response_model=PaymentIn)
async def create_payment_in(payload: PaymentInPayload):
    obj = PaymentIn(**payload.dict())
    await db.payments_in.insert_one(obj.dict())
    return obj


@api_router.put("/payments/in/{pid}", response_model=PaymentIn)
async def update_payment_in(pid: str, payload: PaymentInPayload):
    await db.payments_in.update_one({"id": pid}, {"$set": payload.dict()})
    doc = await db.payments_in.find_one({"id": pid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Not found")
    return PaymentIn(**doc)


@api_router.delete("/payments/in/{pid}")
async def delete_payment_in(pid: str):
    await db.payments_in.delete_one({"id": pid})
    return {"ok": True}


# ====== Payments Out (списания перевозчикам) ======
class PaymentOut(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    pp_number: str
    date: str
    amount: float
    carrier_id: Optional[str] = ""
    carrier_name: Optional[str] = ""
    order_id: Optional[str] = ""
    order_number: Optional[str] = ""
    notes: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class PaymentOutPayload(BaseModel):
    pp_number: str
    date: str
    amount: float
    carrier_id: Optional[str] = ""
    carrier_name: Optional[str] = ""
    order_id: Optional[str] = ""
    order_number: Optional[str] = ""
    notes: Optional[str] = ""


@api_router.get("/payments/out", response_model=List[PaymentOut])
async def list_payments_out(carrier_id: Optional[str] = None, month: Optional[str] = None):
    q: dict = {}
    if carrier_id:
        q["carrier_id"] = carrier_id
    if month:
        q["date"] = {"$regex": f"^{month}"}
    docs = await db.payments_out.find(q, {"_id": 0}).sort("date", -1).to_list(5000)
    return [PaymentOut(**d) for d in docs]


@api_router.post("/payments/out", response_model=PaymentOut)
async def create_payment_out(payload: PaymentOutPayload):
    obj = PaymentOut(**payload.dict())
    await db.payments_out.insert_one(obj.dict())
    return obj


@api_router.put("/payments/out/{pid}", response_model=PaymentOut)
async def update_payment_out(pid: str, payload: PaymentOutPayload):
    await db.payments_out.update_one({"id": pid}, {"$set": payload.dict()})
    doc = await db.payments_out.find_one({"id": pid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Not found")
    return PaymentOut(**doc)


@api_router.delete("/payments/out/{pid}")
async def delete_payment_out(pid: str):
    await db.payments_out.delete_one({"id": pid})
    return {"ok": True}


# ====== Reconciliation (акт сверки) ======
RECONCILIATION_TEMPLATE_ID = "14O_8Er-fY6ituNFyEvMClUdhgfhTOftMeo5vGdUUSdI"
RECONCILIATION_FOLDER_ID   = "1xXXH1_zjVStziS0txoDJoIgXnVizHZ3i"
OUR_NAME = "Александрович Е.А. ИП"


class ReconciliationRequest(BaseModel):
    type: str  # "client" | "carrier"
    counterparty_id: str
    counterparty_name: Optional[str] = None
    period: str  # "year" | "quarter" | "custom"
    year: Optional[int] = None
    quarter: Optional[int] = None
    date_from: Optional[str] = None
    date_to: Optional[str] = None


def _quarter_dates(year: int, quarter: int):
    q_starts = {1: "01-01", 2: "04-01", 3: "07-01", 4: "10-01"}
    q_ends   = {1: "03-31", 2: "06-30", 3: "09-30", 4: "12-31"}
    return f"{year}-{q_starts[quarter]}", f"{year}-{q_ends[quarter]}"


def _fmt_date(iso: str) -> str:
    """2026-01-12 → 12.01.2026"""
    try:
        parts = iso[:10].split("-")
        return f"{parts[2]}.{parts[1]}.{parts[0]}"
    except Exception:
        return iso


def _balance_str(balance: float, side: str, cp_name: str) -> str:
    """Format balance as 'составляет X,XX BYN в пользу ...' or 'нет'."""
    if abs(balance) < 0.005 or side == "none":
        return "нет"
    amt = f"{abs(balance):.2f}".replace(".", ",")
    favor = OUR_NAME if side == "them" else cp_name
    return f"составляет {amt} BYN в пользу {favor}"


def _find_placeholder_table(body_content: list):
    """Return (table_start_index, placeholder_row_index) or None."""
    for elem in body_content:
        if "table" not in elem:
            continue
        tbl = elem["table"]
        tbl_start = elem.get("startIndex", 0)
        for ri, row in enumerate(tbl.get("tableRows", [])):
            for cell in row.get("tableCells", []):
                for cp in cell.get("content", []):
                    for pe in cp.get("paragraph", {}).get("elements", []):
                        if "{{СТРОКА" in pe.get("textRun", {}).get("content", ""):
                            return tbl_start, ri
    return None


def _fill_table_rows(docs_svc, doc_id: str, left_rows: list, right_rows: list):
    """Insert data rows into the template table, then delete the placeholder row."""
    max_rows = max(len(left_rows), len(right_rows))

    doc = docs_svc.documents().get(documentId=doc_id).execute()
    body = doc.get("body", {}).get("content", [])
    info = _find_placeholder_table(body)
    if info is None:
        return

    tbl_start, ph_row_idx = info

    # Delete placeholder immediately if no data rows
    if max_rows == 0:
        docs_svc.documents().batchUpdate(documentId=doc_id, body={"requests": [
            {"deleteTableRow": {"tableCellLocation": {
                "tableStartLocation": {"index": tbl_start},
                "rowIndex": ph_row_idx, "columnIndex": 0
            }}}
        ]}).execute()
        return

    # Insert max_rows rows ABOVE the placeholder row
    ins_reqs = [
        {"insertTableRow": {"tableCellLocation": {
            "tableStartLocation": {"index": tbl_start},
            "rowIndex": ph_row_idx, "columnIndex": 0
        }, "insertBelow": False}}
        for _ in range(max_rows)
    ]
    docs_svc.documents().batchUpdate(documentId=doc_id, body={"requests": ins_reqs}).execute()

    # Re-read the document to get fresh indices after insertions
    doc = docs_svc.documents().get(documentId=doc_id).execute()
    body = doc.get("body", {}).get("content", [])
    tbl_elem = next((e for e in body if "table" in e), None)
    if tbl_elem is None:
        return
    tbl_start = tbl_elem.get("startIndex", 0)
    tbl = tbl_elem["table"]

    # Fill new rows (they are now at indices ph_row_idx .. ph_row_idx + max_rows - 1)
    # Table layout: col0=date_L, col1=doc_L, col2=amt_L, col3=separator, col4=date_R, col5=doc_R, col6=amt_R
    text_reqs = []
    for i in range(max_rows):
        row = tbl["tableRows"][ph_row_idx + i]
        cells = row.get("tableCells", [])
        ncols = len(cells)

        vals = [""] * ncols
        if i < len(left_rows):
            lr = left_rows[i]
            if ncols > 0: vals[0] = lr.get("date", "")
            if ncols > 1: vals[1] = lr.get("doc", "") or lr.get("pp_number", "") or lr.get("doc_number", "")
            if ncols > 2:
                raw = lr.get("sum") if "sum" in lr else lr.get("amount", 0)
                vals[2] = raw if isinstance(raw, str) else f"{raw:.2f}".replace(".", ",")
        if i < len(right_rows):
            rr = right_rows[i]
            if ncols > 4: vals[4] = rr.get("date", "")
            if ncols > 5: vals[5] = rr.get("doc", "") or rr.get("doc_number", "") or rr.get("pp_number", "")
            if ncols > 6:
                raw = rr.get("sum") if "sum" in rr else rr.get("amount", 0)
                vals[6] = raw if isinstance(raw, str) else f"{raw:.2f}".replace(".", ",")

        for ci, cell in enumerate(cells):
            text = vals[ci] if ci < len(vals) else ""
            if not text:
                continue
            cell_content = cell.get("content", [])
            if not cell_content:
                continue
            para_elems = cell_content[0].get("paragraph", {}).get("elements", [])
            if not para_elems:
                continue
            insert_idx = para_elems[0].get("startIndex", 0)
            text_reqs.append({"insertText": {
                "location": {"index": insert_idx},
                "text": text
            }})

    if text_reqs:
        # Sort descending so earlier insertions don't shift subsequent indices
        text_reqs.sort(key=lambda r: r["insertText"]["location"]["index"], reverse=True)
        docs_svc.documents().batchUpdate(documentId=doc_id, body={"requests": text_reqs}).execute()

    # Find and delete the placeholder row (search again after insertions)
    doc = docs_svc.documents().get(documentId=doc_id).execute()
    body = doc.get("body", {}).get("content", [])
    tbl_elem = next((e for e in body if "table" in e), None)
    if tbl_elem is None:
        return
    tbl_start = tbl_elem.get("startIndex", 0)
    tbl = tbl_elem["table"]
    ph_row_now = None
    for ri, row in enumerate(tbl.get("tableRows", [])):
        for cell in row.get("tableCells", []):
            for cp in cell.get("content", []):
                for pe in cp.get("paragraph", {}).get("elements", []):
                    if "{{СТРОКА" in pe.get("textRun", {}).get("content", ""):
                        ph_row_now = ri
                        break
                if ph_row_now is not None: break
            if ph_row_now is not None: break
        if ph_row_now is not None: break

    if ph_row_now is not None:
        docs_svc.documents().batchUpdate(documentId=doc_id, body={"requests": [
            {"deleteTableRow": {"tableCellLocation": {
                "tableStartLocation": {"index": tbl_start},
                "rowIndex": ph_row_now, "columnIndex": 0
            }}}
        ]}).execute()


async def _insert_reconciliation_table_rows(docs_svc, doc_id: str, left_rows: list, right_rows: list):
    """Async wrapper around _fill_table_rows for reconciliation template."""
    await asyncio.to_thread(_fill_table_rows, docs_svc, doc_id, left_rows, right_rows)


def _create_reconciliation_doc_sync(data: dict, token_doc: dict) -> tuple:
    from googleapiclient.discovery import build
    from oauth_google import make_user_credentials

    creds, new_token = make_user_credentials(token_doc)

    docs_svc  = build("docs",  "v1", credentials=creds)
    drive_svc = build("drive", "v3", credentials=creds)

    cp_name   = data["counterparty_name"]
    period_lb = data["period_label"]
    title     = f"Акт сверки – {cp_name} – {period_lb}"

    # Copy template into the reconciliation folder
    copied = drive_svc.files().copy(
        fileId=RECONCILIATION_TEMPLATE_ID,
        body={"name": title, "parents": [RECONCILIATION_FOLDER_ID]},
    ).execute()
    doc_id = copied["id"]

    ob         = data.get("opening_balance", 0.0)
    cb         = data.get("closing_balance", 0.0)
    left_total = data.get("left_total", 0)
    right_total= data.get("right_total", 0)

    cur_year      = str(datetime.now(timezone.utc).year)
    date_from_str = data.get("date_from", "")
    date_to_str   = data.get("date_to", "")

    ob_str         = f"{abs(ob):.2f} BYN" if abs(ob) >= 0.005 else "0,00 BYN"
    cb_str         = f"{abs(cb):.2f} BYN" if abs(cb) >= 0.005 else "0,00 BYN"
    balance_side   = data.get("balance_side_text", "")

    requests = [
        {"replaceAllText": {"containsText": {"text": "{{КонтрагентНазвание}}", "matchCase": True}, "replaceText": cp_name}},
        {"replaceAllText": {"containsText": {"text": "{{ДатаНачало}}",          "matchCase": True}, "replaceText": date_from_str}},
        {"replaceAllText": {"containsText": {"text": "{{ДатаКонец}}",           "matchCase": True}, "replaceText": date_to_str}},
        {"replaceAllText": {"containsText": {"text": "{{СальдоНачало}}",        "matchCase": True}, "replaceText": ob_str}},
        {"replaceAllText": {"containsText": {"text": "{{ИтогоЛевые}}",          "matchCase": True}, "replaceText": f"{left_total:.2f}".replace(".", ",")}},
        {"replaceAllText": {"containsText": {"text": "{{ИтогоПравые}}",         "matchCase": True}, "replaceText": f"{right_total:.2f}".replace(".", ",")}},
        {"replaceAllText": {"containsText": {"text": "{{СальдоКонец}}",         "matchCase": True}, "replaceText": cb_str}},
        {"replaceAllText": {"containsText": {"text": "{{СальдоСторона}}",       "matchCase": True}, "replaceText": balance_side}},
        {"replaceAllText": {"containsText": {"text": "{{ГодДок}}",              "matchCase": True}, "replaceText": cur_year}},
    ]

    docs_svc.documents().batchUpdate(
        documentId=doc_id,
        body={"requests": requests},
    ).execute()

    # Fill table rows (insertTableRow + insertText + deleteTableRow for placeholder)
    _fill_table_rows(docs_svc, doc_id, data.get("left_rows", []), data.get("right_rows", []))

    return f"https://docs.google.com/document/d/{doc_id}/edit", new_token


def _order_in_period_filter(date_from: str, date_to: str) -> dict:
    # load_date is blank on some real, paid orders (never filled in) —
    # falling back to created_at keeps them from silently vanishing out of
    # period-based reports like the reconciliation act.
    return {"$or": [
        {"load_date": {"$gte": date_from, "$lte": date_to}},
        {"load_date": {"$in": [None, ""]}, "created_at": {"$gte": date_from, "$lte": date_to + "T23:59:59"}},
    ]}


def _order_before_period_filter(date_from: str) -> dict:
    return {"$or": [
        {"load_date": {"$lt": date_from}},
        {"load_date": {"$in": [None, ""]}, "created_at": {"$lt": date_from}},
    ]}


@api_router.get("/client_pp_ledger/{client_id}")
async def get_client_pp_ledger(client_id: str):
    entries = await db.client_pp_ledger.find(
        {"client_id": client_id}, {"_id": 0}
    ).sort("pp_date", 1).to_list(1000)
    return {"entries": entries}


class PPLedgerEntry(BaseModel):
    pp_number: str
    pp_date: str
    amount: float


@api_router.post("/client_pp_ledger/{client_id}")
async def add_client_pp_entry(client_id: str, payload: PPLedgerEntry):
    client = await db.clients.find_one({"id": client_id})
    if not client:
        raise HTTPException(404, "Клиент не найден")

    entry = {
        "id": str(uuid.uuid4()),
        "client_id": client_id,
        "client_name": client.get("name", ""),
        "pp_number": payload.pp_number,
        "pp_date": payload.pp_date,
        "amount": payload.amount,
        "created_at": now_iso(),
    }
    await db.client_pp_ledger.insert_one(dict(entry))
    entry.pop("_id", None)
    return entry


@api_router.delete("/client_pp_ledger/{client_id}/{entry_id}")
async def delete_client_pp_entry(client_id: str, entry_id: str):
    result = await db.client_pp_ledger.delete_one({"id": entry_id, "client_id": client_id})
    return {"ok": True, "deleted": result.deleted_count}


@api_router.post("/reconciliation/generate")
async def generate_reconciliation(payload: ReconciliationRequest):
    now = datetime.now(timezone.utc)

    # Period → date_from / date_to
    if payload.period == "month":
        start_dt = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_dt   = now
    elif payload.period == "last_month":
        first_this = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_dt     = first_this - timedelta(days=1)
        start_dt   = end_dt.replace(day=1)
    elif payload.period == "quarter":
        q_month  = (now.month - 1) // 3 * 3 + 1
        start_dt = now.replace(month=q_month, day=1, hour=0, minute=0, second=0, microsecond=0)
        end_dt   = now
    elif payload.period == "year":
        start_dt = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end_dt   = now
    else:  # "all"
        start_dt = datetime(2020, 1, 1, tzinfo=timezone.utc)
        end_dt   = now

    date_from = start_dt.strftime("%Y-%m-%d")
    date_to   = end_dt.strftime("%Y-%m-%d")
    cid       = payload.counterparty_id

    def _fmt(d):
        try:
            s = str(d).split("T")[0][:10]
            y, m, day = s.split("-")
            return f"{day}.{m}.{y}"
        except Exception:
            return str(d)[:10] if d else ""

    if payload.type == "client":
        cp_doc = await db.clients.find_one({"id": cid, "deleted": {"$ne": True}}, {"_id": 0})
        if not cp_doc:
            raise HTTPException(404, "Клиент не найден")
        cp_name = payload.counterparty_name or cp_doc.get("name", "")

        orders = await db.orders.find({
            "$and": [{"$or": [{"client_id": cid}, {"client_name": cp_name}]}, _order_in_period_filter(date_from, date_to)],
            "deleted": {"$ne": True}, "status": {"$ne": "cancelled"},
        }).sort("load_date", 1).to_list(10000)

        orders_before = await db.orders.find({
            "$and": [{"$or": [{"client_id": cid}, {"client_name": cp_name}]}, _order_before_period_filter(date_from)],
            "deleted": {"$ne": True}, "status": {"$ne": "cancelled"},
        }, {"_id": 0, "id": 1, "client_rate": 1, "client_paid": 1}).to_list(10000)

        # Оплаты клиента для акта сверки ведутся в отдельном, накапливаемом
        # журнале client_pp_ledger — независимо от заявок и от payments_in,
        # чтобы платежи не терялись/не задваивались при правках заявок.
        ledger_before = await db.client_pp_ledger.find({
            "client_id": cid, "pp_date": {"$lt": date_from},
        }, {"_id": 0}).to_list(10000)
        paid_before = sum(float(e.get("amount") or 0) for e in ledger_before)
        opening_balance = sum(float(o.get("client_rate") or 0) for o in orders_before) - paid_before

        # Табличная шапка в шаблоне: левая колонка — под именем контрагента
        # (его платежи, ПП и своя дата), правая — под нашим ИП (наши заявки).
        ledger_entries = await db.client_pp_ledger.find({
            "client_id": cid, "pp_date": {"$gte": date_from, "$lte": date_to},
        }, {"_id": 0}).sort("pp_date", 1).to_list(10000)
        left_rows = [
            {"date": _fmt(e.get("pp_date")), "doc": f"ПП № {e.get('pp_number', '')}", "sum": f"{float(e.get('amount') or 0):.2f}"}
            for e in ledger_entries
        ]
        total_paid = sum(float(e.get("amount") or 0) for e in ledger_entries)

        right_rows = [
            {"date": _fmt(o.get("load_date")), "doc": o.get("order_number", ""),
             "sum": f"{float(o.get('client_rate') or 0):.2f}"}
            for o in orders
        ]
        total_charged = sum(float(o.get("client_rate") or 0) for o in orders)

    else:  # carrier
        cp_doc = await db.carriers.find_one({"id": cid, "deleted": {"$ne": True}}, {"_id": 0})
        if not cp_doc:
            raise HTTPException(404, "Перевозчик не найден")
        cp_name = payload.counterparty_name or cp_doc.get("company_name", cp_doc.get("name", ""))

        orders = await db.orders.find({
            "$and": [{"$or": [{"carrier_id": cid}, {"carrier_name": cp_name}]}, _order_in_period_filter(date_from, date_to)],
            "deleted": {"$ne": True}, "status": {"$ne": "cancelled"},
        }).sort("load_date", 1).to_list(10000)

        payments = await db.payments_out.find({
            "$or": [{"carrier_id": cid}, {"carrier_name": cp_name}],
            "date": {"$gte": date_from, "$lte": date_to},
        }).sort("date", 1).to_list(10000)

        orders_before = await db.orders.find({
            "$and": [{"$or": [{"carrier_id": cid}, {"carrier_name": cp_name}]}, _order_before_period_filter(date_from)],
            "deleted": {"$ne": True}, "status": {"$ne": "cancelled"},
        }, {"_id": 0, "id": 1, "carrier_rate": 1, "carrier_paid": 1}).to_list(10000)
        pmts_before = await db.payments_out.find({
            "$or": [{"carrier_id": cid}, {"carrier_name": cp_name}],
            "date": {"$lt": date_from},
        }, {"_id": 0, "amount": 1, "order_id": 1}).to_list(10000)

        paid_before_order_ids = {p.get("order_id") for p in pmts_before if p.get("order_id")}
        paid_before = (
            sum(float(o.get("carrier_rate") or 0) for o in orders_before
                if o.get("carrier_paid") and o.get("id") not in paid_before_order_ids)
            + sum(float(p.get("amount") or 0) for p in pmts_before if not p.get("order_id"))
        )
        opening_balance = sum(float(o.get("carrier_rate") or 0) for o in orders_before) - paid_before

        # Табличная шапка в шаблоне: левая колонка — под именем контрагента
        # (его платежи, ПП и своя дата), правая — под нашим ИП (наши заявки).
        paid_order_ids = {p.get("order_id") for p in payments if p.get("order_id")}
        raw_entries = []
        for o in orders:
            if not o.get("carrier_paid") or o.get("id") in paid_order_ids:
                continue
            raw_entries.append({
                "pp": (o.get("carrier_pp_number") or "").strip(),
                "date": o.get("carrier_pp_date") or o.get("carrier_paid_date") or o.get("load_date") or "",
                "sum": float(o.get("carrier_rate") or 0),
                "order_number": o.get("order_number", ""),
            })
        for p in payments:
            raw_entries.append({"pp": (p.get("pp_number") or "").strip(), "date": p.get("date") or "", "sum": float(p.get("amount") or 0), "order_number": p.get("order_number", "")})

        # Одна платёжка может закрывать несколько заявок сразу — если номер
        # ПП совпадает, схлопываем в одну строку с суммарной суммой
        # независимо от даты (см. клиентскую ветку выше — тот же случай).
        merged: dict = {}
        loose = []
        for e in raw_entries:
            if e["pp"]:
                key = e["pp"]
                row = merged.setdefault(key, {"pp": e["pp"], "date": e["date"], "sum": 0.0})
                if e["date"] and (not row["date"] or e["date"] < row["date"]):
                    row["date"] = e["date"]
                row["sum"] += e["sum"]
            else:
                loose.append(e)
        left_entries = [{"date": v["date"], "doc": f"ПП {v['pp']} от {_fmt(v['date'])}", "sum": v["sum"]} for v in merged.values()]
        left_entries += [{"date": e["date"], "doc": (f"Оплата (заявка {e['order_number']})" if e["order_number"] else "Оплата"), "sum": e["sum"]} for e in loose]
        left_entries.sort(key=lambda e: e["date"])
        left_rows = [{"date": _fmt(e["date"]), "doc": e["doc"], "sum": f"{e['sum']:.2f}"} for e in left_entries]

        right_rows = [
            {"date": _fmt(o.get("load_date")), "doc": o.get("order_number", ""),
             "sum": f"{float(o.get('carrier_rate') or 0):.2f}"}
            for o in orders
        ]
        total_charged = sum(float(o.get("carrier_rate") or 0) for o in orders)
        total_paid    = sum(e["sum"] for e in left_entries)

    closing_balance = opening_balance + total_charged - total_paid
    # closing_balance > 0 значит "начислено больше, чем оплачено" — для клиента
    # это долг клиента перед нами, для перевозчика — наш долг перед ним.
    # Знак один и тот же, а то, кому он "в пользу", у клиента и перевозчика
    # противоположный.
    if payload.type == "client":
        balance_side = OUR_NAME if closing_balance > 0 else cp_name
    else:
        balance_side = cp_name if closing_balance > 0 else OUR_NAME
    period_label    = f"с {_fmt(date_from)} по {_fmt(date_to)}"

    print(f"[rec] type={payload.type} name={cp_name} period={date_from}–{date_to} "
          f"left={len(left_rows)} right={len(right_rows)} ob={opening_balance:.2f} cb={closing_balance:.2f}")

    result = {
        "counterparty_name":  cp_name,
        "period_label":       period_label,
        "date_from":          _fmt(date_from),
        "date_to":            _fmt(date_to),
        "opening_balance":    round(opening_balance, 2),
        "left_rows":          left_rows,
        "right_rows":         right_rows,
        "left_total":         round(total_paid, 2),
        "right_total":        round(total_charged, 2),
        "closing_balance":    round(closing_balance, 2),
        "balance_side_text":  f"в пользу {balance_side}",
    }

    doc_url = None
    doc_error = None
    try:
        token_doc = await db.oauth_tokens.find_one({"_id": "google"}, {"_id": 0})
        if not token_doc:
            doc_error = "Google OAuth не подключён — перейдите в Настройки и авторизуйте Google"
        else:
            doc_url, new_token = await asyncio.to_thread(_create_reconciliation_doc_sync, result, token_doc)
            if new_token:
                await db.oauth_tokens.update_one(
                    {"_id": "google"},
                    {"$set": {"access_token": new_token, "updated_at": datetime.now(timezone.utc).isoformat()}},
                )
    except Exception as _de:
        doc_error = str(_de)
        logging.getLogger(__name__).error(f"reconciliation doc failed: {_de}", exc_info=True)

    if doc_url:
        await db.reconciliation_history.insert_one({
            "counterparty_id":   cid,
            "counterparty_name": cp_name,
            "type":              payload.type,
            "period_label":      period_label,
            "doc_url":           doc_url,
            "url":               doc_url,
            "created_at":        datetime.now(timezone.utc),
        })

    return {**result, "doc_url": doc_url, "url": doc_url, "doc_error": doc_error}


@api_router.get("/reconciliation/history")
async def get_reconciliation_history(counterparty_id: Optional[str] = None, type: Optional[str] = None):
    if not counterparty_id:
        return []
    q: dict = {"counterparty_id": counterparty_id}
    if type:
        q["type"] = type
    cursor = db.reconciliation_history.find(q, {"_id": 0}).sort("created_at", -1)
    items = await cursor.to_list(200)
    for item in items:
        if isinstance(item.get("created_at"), datetime):
            item["created_at"] = item["created_at"].strftime("%d.%m.%Y")
    return items


# ====== Notes ======
class Note(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    text: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class NotePayload(BaseModel):
    title: str
    text: Optional[str] = ""


@api_router.get("/notes", response_model=List[Note])
async def list_notes():
    docs = await db.notes.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return [Note(**d) for d in docs]


@api_router.post("/notes", response_model=Note)
async def create_note(payload: NotePayload):
    obj = Note(**payload.dict())
    await db.notes.insert_one(obj.dict())
    return obj


@api_router.put("/notes/{note_id}", response_model=Note)
async def update_note(note_id: str, payload: NotePayload):
    await db.notes.update_one({"id": note_id}, {"$set": payload.dict()})
    doc = await db.notes.find_one({"id": note_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Note not found")
    return Note(**doc)


@api_router.delete("/notes/{note_id}")
async def delete_note(note_id: str):
    await db.notes.delete_one({"id": note_id})
    return {"ok": True}


# ====== Trash ======
_TRASH_LABELS = {
    "orders": "Заявка",
    "clients": "Клиент",
    "carriers": "Перевозчик",
    "leads": "Лид",
}


@api_router.get("/trash")
async def list_trash(current_user: dict = Depends(_require_user)):
    result = []
    now = datetime.now(timezone.utc)
    for cname, type_label in _TRASH_LABELS.items():
        docs = await db[cname].find({"deleted": True}, {"_id": 0}).sort("deleted_at", -1).to_list(1000)
        print(f"[list_trash] {cname}: found {len(docs)} deleted docs")
        for d in docs:
            if cname == "orders":
                num = d.get("order_number") or d.get("id", "")
                parts = [f"№{num}"]
                if d.get("route_from") and d.get("route_to"):
                    parts.append(f"{d['route_from']} → {d['route_to']}")
                if d.get("client_name"):
                    parts.append(d["client_name"])
                label = " · ".join(parts)
            else:
                label = d.get("name") or d.get("company_name") or d.get("title") or d.get("id", "")
            raw_deleted_at = d.get("deleted_at")
            deleted_at_str = None
            days_left = None
            if raw_deleted_at is not None:
                try:
                    if isinstance(raw_deleted_at, datetime):
                        deleted_dt = raw_deleted_at.replace(tzinfo=timezone.utc) if raw_deleted_at.tzinfo is None else raw_deleted_at
                    else:
                        deleted_dt = datetime.fromisoformat(str(raw_deleted_at))
                        if deleted_dt.tzinfo is None:
                            deleted_dt = deleted_dt.replace(tzinfo=timezone.utc)
                    deleted_at_str = deleted_dt.isoformat()
                    days_left = max(0, 30 - (now - deleted_dt).days)
                except Exception as e:
                    print(f"[list_trash] failed to parse deleted_at={raw_deleted_at!r}: {e}")
                    deleted_at_str = str(raw_deleted_at)
            result.append({
                "id": d.get("id"),
                "collection": cname,
                "type": type_label,
                "label": label,
                "deleted_at": deleted_at_str,
                "days_left": days_left,
            })
    print(f"[list_trash] total: {len(result)} items")
    result.sort(key=lambda x: x.get("deleted_at") or "", reverse=True)
    return result


@api_router.post("/trash/restore/{collection}/{item_id}")
async def restore_trash(collection: str, item_id: str, current_user: dict = Depends(_require_user)):
    if collection not in _TRASH_LABELS:
        raise HTTPException(400, "Неизвестная коллекция")
    result = await db[collection].update_one(
        {"id": item_id, "deleted": True},
        {"$unset": {"deleted": "", "deleted_at": ""}},
    )
    if result.matched_count == 0:
        raise HTTPException(404, "Запись не найдена в корзине")
    return {"ok": True}


@api_router.post("/trash/purge")
async def purge_trash(current_user: dict = Depends(require_director)):
    total = 0
    for cname in _TRASH_LABELS:
        r = await db[cname].delete_many({"deleted": True})
        total += r.deleted_count
    return {"ok": True, "deleted_count": total}


# ====== Backup ======
@api_router.get("/backup/list")
async def list_backups(current_user: dict = Depends(require_director)):
    docs = await db.backups.find({}, {"_id": 0, "collections": 0}).sort("created_at", -1).to_list(50)
    return docs


@api_router.post("/backup/create")
async def create_backup_manual(current_user: dict = Depends(require_director)):
    await _create_backup(reason="manual")
    return {"ok": True}


class RestoreRequest(BaseModel):
    confirm_word: str


@api_router.post("/backup/restore/{backup_id}")
async def restore_backup(backup_id: str, payload: RestoreRequest, current_user: dict = Depends(require_director)):
    if payload.confirm_word.strip().upper() != "ВОССТАНОВИТЬ":
        raise HTTPException(400, "Введите слово ВОССТАНОВИТЬ для подтверждения")

    doc = await db.backups.find_one({"id": backup_id})
    if not doc:
        raise HTTPException(404, "Бэкап не найден")

    # Снимок текущего состояния перед восстановлением — на случай если откат тоже ошибка
    await _create_backup(reason=f"pre_restore_{backup_id}")

    snapshot = doc.get("collections", {})
    restored: dict = {}
    for cname, docs_list in snapshot.items():
        if not isinstance(docs_list, list):
            continue
        await db[cname].delete_many({})
        if docs_list:
            await db[cname].insert_many(docs_list)
        restored[cname] = len(docs_list)
    return {"ok": True, "restored_at": now_iso(), "restored": restored}


@api_router.get("/")
async def root():
    return {"message": "Logistics CRM API"}


@api_router.get("/ping")
async def ping():
    return {"ok": True}


@api_router.post("/admin/restore")
async def admin_restore():
    all_perms = {
        "can_view_finance": True,
        "can_view_all_orders": True,
        "can_view_all_clients": True,
        "can_view_all_leads": True,
        "can_create_orders": True,
    }
    result = await db.users.update_one(
        {"login": "admin"},
        {"$set": {"role": "admin", "permissions": all_perms}},
    )
    if result.matched_count == 0:
        raise HTTPException(404, "Пользователь admin не найден")
    user = await db.users.find_one({"login": "admin"}, {"_id": 0, "password_hash": 0})
    return {"ok": True, "user": user}


app.add_middleware(CORSMiddleware, allow_credentials=True, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
app.include_router(api_router)

@app.get("/health")
async def health():
    return {"ok": True}

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 10000))
    uvicorn.run("server:app", host="0.0.0.0", port=port)
